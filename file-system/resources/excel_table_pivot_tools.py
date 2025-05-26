"""
Advanced Excel Table and PivotTable manipulation tools for the MCP server.
This module leverages xlwings for rich PivotTable functionality and openpyxl for base operations.
"""
import os
import logging
from typing import Any, Dict, List, Optional, Union, Tuple
from datetime import datetime

# Attempt to import xlwings, critical for PivotTable functionality
try:
    import xlwings as xw
    HAS_XLWINGS = True
except ImportError:
    HAS_XLWINGS = False
    print("xlwings library not installed. Full PivotTable functionality will be limited. pip install xlwings")

# Attempt to import openpyxl, used for table operations and base file handling
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.styles import Font, Border, PatternFill, Side, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("openpyxl library not installed. Some table functionalities will be limited. pip install openpyxl")

# Attempt to import pandas, used as a helper or fallback
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    print("pandas library not installed. Some data manipulation fallbacks will be unavailable. pip install pandas")


logger = logging.getLogger(__name__)

# --- Helper Functions ---
def _check_dependencies_for_pivot_tables() -> bool:
    if not HAS_XLWINGS:
        logger.error("xlwings is required for PivotTable operations but is not installed.")
        return False
    if not HAS_OPENPYXL: # Though xlwings handles files, openpyxl might be used for initial checks or data prep
        logger.warning("openpyxl is recommended for full Excel integration.")
    return True

def _get_excel_app_and_workbook(filepath: str, visible: bool = False) -> Tuple[Optional[xw.App], Optional[xw.Book]]:
    """Starts an Excel application instance and opens the workbook."""
    if not HAS_XLWINGS:
        return None, None
    try:
        app = xw.App(visible=visible, add_book=False)
        # Ensure the file exists before trying to open, or xlwings might create it
        if not os.path.exists(filepath):
            logger.error(f"File not found for xlwings operation: {filepath}")
            app.quit()
            return None, None
        wb = app.books.open(filepath)
        return app, wb
    except Exception as e:
        logger.error(f"Failed to start Excel or open workbook '{filepath}' with xlwings: {e}")
        if 'app' in locals() and app:
            app.quit()
        return None, None

def _close_excel_app(app: Optional[xw.App], wb: Optional[xw.Book], save: bool = True):
    """Closes the workbook and quits the Excel application."""
    if app and wb:
        try:
            if save:
                wb.save()
            wb.close()
        except Exception as e:
            logger.error(f"Error during workbook save/close: {e}")
        finally:
            try:
                app.quit()
            except Exception as e:
                logger.error(f"Error quitting Excel app: {e}")
    elif app:
        try:
            app.quit()
        except Exception as e:
            logger.error(f"Error quitting Excel app: {e}")

# --- Excel Table Tools (using openpyxl for creation, xlwings for dynamic parts) ---

def create_excel_table(
    filepath: str,
    sheet_name: str,
    data_range: str, # e.g., "A1:D10"
    table_name: str,
    table_style: str = "TableStyleMedium9"
) -> Dict[str, Any]:
    """
    Creates a formatted Excel table from a specified range.
    Uses openpyxl for table creation.
    """
    if not HAS_OPENPYXL:
        return {"success": False, "error": "openpyxl is required to create tables."}

    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            return {"success": False, "error": f"Sheet '{sheet_name}' not found."}
        ws = wb[sheet_name]

        # Remove existing table if it has the same name to prevent errors
        if table_name in ws.tables:
            del ws.tables[table_name]
            logger.info(f"Removed existing table '{table_name}' to recreate.")

        table = Table(displayName=table_name, ref=data_range)
        style = TableStyleInfo(
            name=table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)
        wb.save(filepath)
        return {
            "success": True,
            "message": f"Table '{table_name}' created in sheet '{sheet_name}' for range '{data_range}'.",
        }
    except Exception as e:
        logger.error(f"Failed to create Excel table: {e}")
        return {"success": False, "error": f"Failed to create Excel table: {str(e)}"}

def sort_excel_table(
    filepath: str,
    sheet_name: str,
    table_name: str,
    sort_column_name: str, # Name of the column header to sort by
    sort_order: str = "ascending" # "ascending" or "descending"
) -> Dict[str, Any]:
    """
    Sorts an Excel table by a specified column.
    Uses xlwings for dynamic sorting as openpyxl's capabilities are limited here.
    """
    if not _check_dependencies_for_pivot_tables(): # xlwings needed
        return {"success": False, "error": "xlwings is required for sorting tables."}

    app, wb = _get_excel_app_and_workbook(filepath)
    if not wb:
        return {"success": False, "error": "Failed to open workbook with xlwings."}

    try:
        ws = wb.sheets[sheet_name]
        if table_name not in [tbl.name for tbl in ws.tables]:
            return {"success": False, "error": f"Table '{table_name}' not found in sheet '{sheet_name}'."}
        
        table = ws.tables[table_name]
        
        # Find the column index from its name
        header_row_range = table.header_row_range
        column_names = [cell.value for cell in header_row_range]
        if sort_column_name not in column_names:
            return {"success": False, "error": f"Column '{sort_column_name}' not found in table '{table_name}'. Available columns: {column_names}"}
        
        # xlwings table sort column is 1-based index within the table's range
        sort_column_index_in_table = column_names.index(sort_column_name) + 1

        # xlSortOrder constants: 1 for Ascending, 2 for Descending
        xl_sort_order = 1 if sort_order.lower() == "ascending" else 2
        
        # Access the underlying ListObject API for sorting
        list_object = table.api
        list_object.Sort.SortFields.Clear()
        list_object.Sort.SortFields.Add(
            Key=table.range.columns[sort_column_index_in_table-1].api, # Key is a Range object
            SortOn=0,  # xlSortOnValues
            Order=xl_sort_order,
            DataOption=0  # xlSortNormal
        )
        list_object.Sort.Header = 1 # xlYes (Indicates that the table has headers)
        list_object.Sort.Apply()
        
        _close_excel_app(app, wb, save=True)
        return {
            "success": True,
            "message": f"Table '{table_name}' in sheet '{sheet_name}' sorted by column '{sort_column_name}' ({sort_order})."
        }
    except Exception as e:
        logger.error(f"Failed to sort Excel table: {e}")
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Failed to sort Excel table: {str(e)}"}

def filter_excel_table(
    filepath: str,
    sheet_name: str,
    table_name: str,
    column_name: str, # Name of the column header to filter
    criteria1: str, # For simple filters. For multiple values, use list in criteria1.
    operator: str = "equals", # "equals", "contains", "beginswith", "endswith", "greaterthan", "lessthan", "between"
    criteria2: Optional[str] = None # For "between" operator
) -> Dict[str, Any]:
    """
    Filters an Excel table by a specified column and criteria.
    Uses xlwings for dynamic filtering.
    """
    if not _check_dependencies_for_pivot_tables(): # xlwings needed
        return {"success": False, "error": "xlwings is required for filtering tables."}

    app, wb = _get_excel_app_and_workbook(filepath)
    if not wb:
        return {"success": False, "error": "Failed to open workbook with xlwings."}

    try:
        ws = wb.sheets[sheet_name]
        if table_name not in [tbl.name for tbl in ws.tables]:
            return {"success": False, "error": f"Table '{table_name}' not found in sheet '{sheet_name}'."}
        
        table = ws.tables[table_name]
        
        # Find the column index from its name
        header_row_range = table.header_row_range
        column_names = [cell.value for cell in header_row_range]
        if column_name not in column_names:
            return {"success": False, "error": f"Column '{column_name}' not found in table '{table_name}'. Available columns: {column_names}"}
        
        # xlwings table filter column is 1-based index within the table's range
        filter_column_index_in_table = column_names.index(column_name) + 1

        # Clear existing filters on the table's range
        if table.api.AutoFilter:
            table.api.AutoFilter.ShowAllData()

        # Apply AutoFilter. For multiple criteria, Excel uses an array of strings for Criteria1.
        # xlwings handles this by accepting a list for criteria1.
        # Operator mapping (simplified for now, can be expanded)
        # xlFilterValues = 7 (for list of values)
        # For text filters like "contains", "beginswith", etc., Criteria1 format is specific.
        
        actual_criteria1 = criteria1
        xl_operator = 0 # Default: xlAnd, not directly used by AutoFilter method like this
        
        # Basic operator mapping for common cases (can be made more robust)
        op_lower = operator.lower()
        if op_lower == "equals":
             actual_criteria1 = f"={criteria1}"
        elif op_lower == "contains":
            actual_criteria1 = f"*{criteria1}*"
        elif op_lower == "beginswith":
            actual_criteria1 = f"{criteria1}*"
        elif op_lower == "endswith":
            actual_criteria1 = f"*{criteria1}"
        elif op_lower == "greaterthan":
            actual_criteria1 = f">{criteria1}"
        elif op_lower == "lessthan":
            actual_criteria1 = f"<{criteria1}"
        elif op_lower == "between" and criteria2:
            # For 'between', Excel's AutoFilter typically needs two criteria
            # This is a simplified handling; true "between" might need specific xlwings features or API calls
            actual_criteria1 = f">{criteria1}"
            # A second filter call might be needed, or a more complex Criteria1
            # For now, let's assume criteria1 handles the lower bound and we'd need another for upper.
            # A robust solution might involve xlFilterValues with a list of all matching items.
            # This example only applies the first part.
            pass 
            
        # Apply the filter
        table.range.api.AutoFilter(Field=filter_column_index_in_table, Criteria1=actual_criteria1)
        
        _close_excel_app(app, wb, save=True)
        return {
            "success": True,
            "message": f"Table '{table_name}' filtered by column '{column_name}' with criteria '{criteria1}'."
        }
    except Exception as e:
        logger.error(f"Failed to filter Excel table: {e}")
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Failed to filter Excel table: {str(e)}"}


# --- PivotTable Tools (primarily using xlwings) ---

# Function mapping for value fields
# xlConsolidationFunction constants
XL_SUM = -4157
XL_COUNT = -4112
XL_AVERAGE = -4106
XL_MAX = -4136
XL_MIN = -4139
XL_PRODUCT = -4149
XL_COUNT_NUMS = -4113
XL_STDEV = -4155
XL_STDEVP = -4156
XL_VAR = -4164
XL_VARP = -4165

# xlPivotFieldCalculation constants for "Show Value As"
XL_NORMAL = 0 # Or sometimes seen as xw.constants.PivotFieldCalculation.xlNormal / pywintypes. också -4143
XL_DIFFERENCE_FROM = 2
XL_PERCENT_DIFFERENCE_FROM = 4
XL_PERCENT_OF = 3
XL_PERCENT_OF_GRAND_TOTAL = 8 # % of Grand Total (same as xlPercentOfTotal)
XL_PERCENT_OF_ROW_TOTAL = 7 # often xlPercentOfRow
XL_PERCENT_OF_COLUMN_TOTAL = 6 # often xlPercentOfColumn
XL_PERCENT_OF_PARENT_ROW = 12
XL_PERCENT_OF_PARENT_COLUMN = 13
XL_PERCENT_OF_PARENT = 14 # xlPercentOfParentTotal?
XL_RUNNING_TOTAL_IN = 5
XL_RANK_SMALLEST_TO_LARGEST = 11 # xw.constants.PivotFieldCalculation.xlRankAscending
XL_RANK_LARGEST_TO_SMALLEST = 10 # xw.constants.PivotFieldCalculation.xlRankDescending
XL_INDEX = 15 # xw.constants.PivotFieldCalculation.xlIndex

FUNCTION_MAP = {
    "sum": XL_SUM, "count": XL_COUNT, "average": XL_AVERAGE,
    "max": XL_MAX, "min": XL_MIN, "product": XL_PRODUCT,
    "countnumbers": XL_COUNT_NUMS, "stdev": XL_STDEV, "stdevp": XL_STDEVP,
    "var": XL_VAR, "varp": XL_VARP,
}

SHOW_VALUE_AS_MAP = {
    "normal": XL_NORMAL,
    "differencefrom": XL_DIFFERENCE_FROM,
    "%differencefrom": XL_PERCENT_DIFFERENCE_FROM,
    "%of": XL_PERCENT_OF,
    "%ofgrandtotal": XL_PERCENT_OF_GRAND_TOTAL, # Excel often calls this % of Total
    "%ofcolumntotal": XL_PERCENT_OF_COLUMN_TOTAL,
    "%ofrowtotal": XL_PERCENT_OF_ROW_TOTAL,
    "%ofparentrowtotal": XL_PERCENT_OF_PARENT_ROW,
    "%ofparentcolumntotal": XL_PERCENT_OF_PARENT_COLUMN,
    "%ofparenttotal": XL_PERCENT_OF_PARENT, # Parent item total
    "runningtotalin": XL_RUNNING_TOTAL_IN,
    "ranksmallesttolargest": XL_RANK_SMALLEST_TO_LARGEST,
    "ranklargesttosmallest": XL_RANK_LARGEST_TO_SMALLEST,
    "index": XL_INDEX,
}

def create_pivot_table(
    filepath: str,
    source_sheet_name: str,
    source_data_range: str, # e.g., "A1:G100"
    target_sheet_name: str,
    target_cell_address: str, # e.g., "A3"
    pivot_table_name: str,
    row_fields: Optional[List[str]] = None,
    column_fields: Optional[List[str]] = None,
    value_fields: Optional[List[Dict[str, Any]]] = None, 
    # Example value_fields:
    # [
    #   {"field": "Sales", "function": "Sum", "custom_name": "Total Sales"},
    #   {"field": "Revenue", "function": "Sum", "custom_name": "% of Revenue", 
    #    "show_value_as": "%OfGrandTotal"},
    #   {"field": "Profit", "function": "Sum", "custom_name": "Profit Diff from Prev Month",
    #    "show_value_as": "DifferenceFrom", "base_field_name": "Month", "base_item_value": "(previous)"}
    # ]
    filter_fields: Optional[List[str]] = None,
    pivot_style: str = "PivotStyleMedium9"
) -> Dict[str, Any]:
    """
    Creates an Excel PivotTable with specified rows, columns, values, and filters.
    """
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings is required for PivotTable operations."}

    app, wb = _get_excel_app_and_workbook(filepath)
    if not wb:
        return {"success": False, "error": "Failed to open workbook with xlwings."}

    try:
        src_ws = wb.sheets[source_sheet_name]
        
        # Ensure target sheet exists, create if not
        if target_sheet_name not in [s.name for s in wb.sheets]:
            wb.sheets.add(target_sheet_name)
        tgt_ws = wb.sheets[target_sheet_name]

        # Define the source data for the PivotCache
        # For xlwings, it's often easier to pass the sheet and range separately or use full address string
        source_data_ref = f"'{source_sheet_name}'!{source_data_range}"
        
        # Create PivotCache
        # xlDatabase = 1
        pivot_cache = wb.api.PivotCaches().Create(SourceType=1, SourceData=source_data_ref)

        # Create PivotTable
        # TableDestination needs to be a Range object in the API, or a string like "SheetName!R1C1"
        # Using R1C1 for robustness with xlwings API for TableDestination
        # Convert A1 to R1C1 for target_cell_address
        col_letter = ''.join(filter(str.isalpha, target_cell_address))
        row_num = int(''.join(filter(str.isdigit, target_cell_address)))
        col_num = column_index_from_string(col_letter)
        target_r1c1 = f"R{row_num}C{col_num}"
        
        # Ensure pivot table name is unique for the sheet or workbook
        # xlwings doesn't directly expose PivotTables collection on a sheet easily, use API
        existing_pivot_names = [pt.Name for pt in tgt_ws.api.PivotTables()]
        if pivot_table_name in existing_pivot_names:
            # Attempt to delete existing pivot table
            try:
                tgt_ws.api.PivotTables(pivot_table_name).TableRange2.Clear() # Clears the pivot table range
                logger.info(f"Cleared existing PivotTable '{pivot_table_name}' to recreate.")
            except Exception as pt_clear_error:
                logger.warning(f"Could not clear existing PivotTable '{pivot_table_name}': {pt_clear_error}. Recreation might fail or create with a default name.")
                # Excel might rename it automatically if name conflicts and old one can't be removed this way
        
        pivot_table_obj = pivot_cache.CreatePivotTable(
            TableDestination=tgt_ws.range(target_cell_address).api, # Pass Range object's API
            TableName=pivot_table_name
        )
        
        # Constants for field orientation
        XL_ROW_FIELD = 1
        XL_COLUMN_FIELD = 2
        XL_PAGE_FIELD = 3 # Filter field
        XL_DATA_FIELD = 4

        # Configure Row Fields
        if row_fields:
            for i, field_name in enumerate(row_fields):
                pf = pivot_table_obj.PivotFields(field_name)
                pf.Orientation = XL_ROW_FIELD
                pf.Position = i + 1
        
        # Configure Column Fields
        if column_fields:
            for i, field_name in enumerate(column_fields):
                pf = pivot_table_obj.PivotFields(field_name)
                pf.Orientation = XL_COLUMN_FIELD
                pf.Position = i + 1

        # Configure Value Fields
        if value_fields:
            for i, val_field_config in enumerate(value_fields):
                field_name = val_field_config["field"]
                func_name = val_field_config.get("function", "Sum").lower()
                custom_name = val_field_config.get("custom_name")
                
                show_value_as = val_field_config.get("show_value_as", "normal").lower()
                base_field_name = val_field_config.get("base_field_name")
                base_item_value = val_field_config.get("base_item_value")

                pf_source = pivot_table_obj.PivotFields(field_name)
                
                # Add as data field - this creates a new PivotField in Data area
                # The caption is set here.
                data_field = pivot_table_obj.AddDataField(
                    Field=pf_source, 
                    Caption=custom_name or f"{func_name.capitalize()} of {field_name}", 
                    Function=FUNCTION_MAP.get(func_name, XL_SUM)
                )
                
                # Now, apply "Show Value As" settings if not "normal"
                if show_value_as != "normal":
                    calculation_type = SHOW_VALUE_AS_MAP.get(show_value_as)
                    if calculation_type is not None:
                        data_field.Calculation = calculation_type
                        if base_field_name and calculation_type in [XL_DIFFERENCE_FROM, XL_PERCENT_OF, XL_PERCENT_DIFFERENCE_FROM, XL_RUNNING_TOTAL_IN]:
                            try:
                                data_field.BaseField = pivot_table_obj.PivotFields(base_field_name).Name # Use .Name to pass string
                            except Exception as bf_e:
                                logger.warning(f"Could not set BaseField '{base_field_name}' for '{data_field.Caption}': {bf_e}")
                        if base_item_value and calculation_type in [XL_DIFFERENCE_FROM, XL_PERCENT_OF, XL_PERCENT_DIFFERENCE_FROM]:
                            # For "(previous)" or "(next)", xlwings/Excel API handles it by setting BaseItem appropriately
                            if base_item_value.lower() == "(previous)":
                                data_field.api.BaseItem = "(previous)" # More direct API interaction often works
                            elif base_item_value.lower() == "(next)":
                                data_field.api.BaseItem = "(next)"
                            else: # Specific item name
                                try:
                                    data_field.BaseItem = base_item_value
                                except Exception as bi_e:
                                     logger.warning(f"Could not set BaseItem '{base_item_value}' for '{data_field.Caption}': {bi_e}")
                    else:
                        logger.warning(f"Unsupported 'show_value_as' type: {val_field_config.get('show_value_as')}")
        
        # Configure Filter Fields
        if filter_fields:
            for i, field_name in enumerate(filter_fields):
                pf = pivot_table_obj.PivotFields(field_name)
                pf.Orientation = XL_PAGE_FIELD
                pf.Position = i + 1
        
        # Apply Style
        if pivot_style:
            pivot_table_obj.TableStyle2 = pivot_style # Use TableStyle2 for modern styles

        # Refresh (might not be necessary on creation but good practice)
        pivot_table_obj.PivotCache().Refresh()

        _close_excel_app(app, wb, save=True)
        return {
            "success": True,
            "message": f"PivotTable '{pivot_table_name}' created in sheet '{target_sheet_name}' at '{target_cell_address}'.",
            "name": pivot_table_obj.Name # Return actual name as Excel might change it
        }
    except Exception as e:
        logger.error(f"Failed to create PivotTable: {e}")
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Failed to create PivotTable: {str(e)}"}

def modify_pivot_table_fields(
    filepath: str,
    sheet_name: str,
    pivot_table_name: str,
    add_row_fields: Optional[List[str]] = None,
    add_column_fields: Optional[List[str]] = None,
    add_value_fields: Optional[List[Dict[str, Any]]] = None, # Updated to Dict[str, Any]
    add_filter_fields: Optional[List[str]] = None,
    remove_fields: Optional[List[str]] = None
) -> Dict[str, Any]:
    """Modifies the field layout of an existing PivotTable."""
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings is required for PivotTable operations."}

    app, wb = _get_excel_app_and_workbook(filepath)
    if not wb:
        return {"success": False, "error": "Failed to open workbook with xlwings."}

    try:
        ws = wb.sheets[sheet_name]
        # Access PivotTable via API
        try:
            pivot_table_obj = ws.api.PivotTables(pivot_table_name)
        except Exception:
             return {"success": False, "error": f"PivotTable '{pivot_table_name}' not found in sheet '{sheet_name}'."}

        # Constants and function map as in create_pivot_table
        XL_ROW_FIELD, XL_COLUMN_FIELD, XL_PAGE_FIELD, XL_DATA_FIELD = 1, 2, 3, 4
        function_map = {
            "sum": -4157, "count": -4112, "average": -4106, "max": -4136, "min": -4139
        }

        # Remove Fields
        if remove_fields:
            for field_name in remove_fields:
                try:
                    # This removes the field from any area it's in
                    pivot_table_obj.PivotFields(field_name).Orientation = 0 # xlHidden
                except Exception as e_remove:
                    logger.warning(f"Could not remove field '{field_name}': {e_remove}")
        
        # Add/Update Row Fields
        if add_row_fields:
            for i, field_name in enumerate(add_row_fields):
                pf = pivot_table_obj.PivotFields(field_name)
                pf.Orientation = XL_ROW_FIELD
                # Position might need to be managed carefully if fields are reordered, not just added
                # pf.Position = pivot_table_obj.RowFields.Count + 1 # Simplistic positioning

        # Add/Update Column Fields
        if add_column_fields:
            for i, field_name in enumerate(add_column_fields):
                pf = pivot_table_obj.PivotFields(field_name)
                pf.Orientation = XL_COLUMN_FIELD
                # pf.Position = pivot_table_obj.ColumnFields.Count + 1

        # Add/Update Value Fields
        if add_value_fields:
            for val_field_config in add_value_fields:
                field_name = val_field_config["field"]
                func_name = val_field_config.get("function", "Sum").lower()
                custom_name = val_field_config.get("custom_name")
                show_value_as = val_field_config.get("show_value_as", "normal").lower()
                base_field_name = val_field_config.get("base_field_name")
                base_item_value = val_field_config.get("base_item_value")

                pf_to_add = pivot_table_obj.PivotFields(field_name)
                
                # Check if this field (as a data field with this function and name) already exists to avoid duplicates
                # This check is simplistic. Excel allows same source field multiple times with different calcs/names.
                # A more robust check would see if a DataField with the target `custom_name` exists.
                already_exists = False
                target_caption = custom_name or f"{func_name.capitalize()} of {field_name}"
                for df_existing in pivot_table_obj.DataFields:
                    if df_existing.Name == target_caption: # Name is the caption
                        already_exists = True
                        logger.info(f"Value field with caption '{target_caption}' already exists. Skipping addition.")
                        break
                
                if not already_exists:
                    data_field = pivot_table_obj.AddDataField(
                        Field=pf_to_add, 
                        Caption=target_caption, 
                        Function=FUNCTION_MAP.get(func_name, XL_SUM)
                    )
                    # Apply "Show Value As"
                    if show_value_as != "normal":
                        calculation_type = SHOW_VALUE_AS_MAP.get(show_value_as)
                        if calculation_type is not None:
                            data_field.Calculation = calculation_type
                            if base_field_name and calculation_type in [XL_DIFFERENCE_FROM, XL_PERCENT_OF, XL_PERCENT_DIFFERENCE_FROM, XL_RUNNING_TOTAL_IN]:
                                data_field.BaseField = pivot_table_obj.PivotFields(base_field_name).Name
                            if base_item_value and calculation_type in [XL_DIFFERENCE_FROM, XL_PERCENT_OF, XL_PERCENT_DIFFERENCE_FROM]:
                                if base_item_value.lower() == "(previous)": data_field.api.BaseItem = "(previous)"
                                elif base_item_value.lower() == "(next)": data_field.api.BaseItem = "(next)"
                                else: data_field.BaseItem = base_item_value
                        else:
                            logger.warning(f"Unsupported 'show_value_as' type: {val_field_config.get('show_value_as')} for field {target_caption}")


        # Add/Update Filter Fields
        if add_filter_fields:
            for i, field_name in enumerate(add_filter_fields):
                pf = pivot_table_obj.PivotFields(field_name)
                pf.Orientation = XL_PAGE_FIELD
                # pf.Position = pivot_table_obj.PageFields.Count + 1
        
        pivot_table_obj.PivotCache().Refresh()
        _close_excel_app(app, wb, save=True)
        return {"success": True, "message": f"PivotTable '{pivot_table_name}' fields updated."}
    except Exception as e:
        logger.error(f"Failed to modify PivotTable fields: {e}")
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Failed to modify PivotTable fields: {str(e)}"}

def sort_pivot_table_field(
    filepath: str,
    sheet_name: str,
    pivot_table_name: str,
    field_name: str, # The field in Row/Column area whose items are to be sorted
    sort_on_field: str, # The DataField (value field) to sort by, or the field_name itself for label sort
    sort_order: str = "ascending", # "ascending" or "descending"
    sort_type: str = "data" # "data" (sort by values) or "label" (sort by item names A-Z)
) -> Dict[str, Any]:
    """Sorts a PivotTable field by its own labels or by values of a data field."""
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings is required for PivotTable operations."}

    app, wb = _get_excel_app_and_workbook(filepath)
    if not wb:
        return {"success": False, "error": "Failed to open workbook with xlwings."}

    try:
        ws = wb.sheets[sheet_name]
        try:
            pivot_table_obj = ws.api.PivotTables(pivot_table_name)
        except Exception:
             return {"success": False, "error": f"PivotTable '{pivot_table_name}' not found."}

        pf_to_sort = pivot_table_obj.PivotFields(field_name)

        # xlAscending = 1, xlDescending = 2
        xl_order = 1 if sort_order.lower() == "ascending" else 2
        
        # For AutoSort, OrderField is the name of the "Sum of Sales" type field if sorting by data.
        # If sorting by labels, it's the field_name itself.
        actual_sort_on_field = sort_on_field
        if sort_type.lower() == "label":
            actual_sort_on_field = field_name # Correct for label sort if API expects the field itself
            # Some APIs might require passing the field name itself, others a specific flag.
            # xlwings/pywin32 AutoSort takes the data field name for 'OrderField'
            # If sorting by labels, the data field context might not be relevant or handled differently.
            # For label sort, usually, you sort the PivotField itself.
            pf_to_sort.AutoSort(xl_order, field_name) # Sorting field_name by its own items
        else: # sort_type == "data"
            # Ensure sort_on_field is a valid DataField caption, e.g., "Sum of Sales"
            is_valid_data_field = False
            for df in pivot_table_obj.DataFields:
                if df.Name == sort_on_field: # df.Name is usually the caption e.g. "Sum of Sales"
                    is_valid_data_field = True
                    break
            if not is_valid_data_field:
                valid_data_fields = [df.Name for df in pivot_table_obj.DataFields]
                return {"success": False, "error": f"Data field '{sort_on_field}' not found for sorting. Available: {valid_data_fields}"}
            pf_to_sort.AutoSort(xl_order, actual_sort_on_field)

        pivot_table_obj.PivotCache().Refresh()
        _close_excel_app(app, wb, save=True)
        return {"success": True, "message": f"PivotTable field '{field_name}' sorted by '{actual_sort_on_field}' ({sort_order})."}
    except Exception as e:
        logger.error(f"Failed to sort PivotTable field: {e}")
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Failed to sort PivotTable field: {str(e)}"}

def filter_pivot_table_items(
    filepath: str,
    sheet_name: str,
    pivot_table_name: str,
    field_name: str,
    visible_items: Optional[List[str]] = None, # Items to make visible, if None, shows all
    hidden_items: Optional[List[str]] = None, # Items to make hidden
    filter_type: str = "value" # "value", "label", "date" - for more advanced filters later
) -> Dict[str, Any]:
    """Filters items in a PivotTable field. Shows/hides specific items."""
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings is required for PivotTable operations."}

    app, wb = _get_excel_app_and_workbook(filepath)
    if not wb:
        return {"success": False, "error": "Failed to open workbook with xlwings."}
    try:
        ws = wb.sheets[sheet_name]
        try:
            pivot_table_obj = ws.api.PivotTables(pivot_table_name)
        except Exception:
             return {"success": False, "error": f"PivotTable '{pivot_table_name}' not found."}

        pf = pivot_table_obj.PivotFields(field_name)
        
        # Clear any existing filters on this specific field to start fresh
        pf.ClearAllFilters()
        
        # Enable manual filtering by items
        pf.EnableMultiplePageItems = True # Important for item-level visibility

        if visible_items is not None:
            # First, hide all items (if we are specifying what to make visible)
            # This is necessary because directly setting Visible=True for an item doesn't auto-hide others
            # when EnableMultiplePageItems = True and some items are already filtered.
            if len(visible_items) > 0 : # Only hide all if we intend to show specific ones
                all_items = [item.Name for item in pf.PivotItems()]
                for item_name_to_hide in all_items:
                    if item_name_to_hide not in visible_items:
                        try:
                            pf.PivotItems(item_name_to_hide).Visible = False
                        except Exception as e_hide:
                            logger.warning(f"Could not hide item '{item_name_to_hide}': {e_hide}")
            
            # Then, make specified items visible
            for item_name in visible_items:
                try:
                    pf.PivotItems(item_name).Visible = True
                except Exception as e_show:
                    logger.warning(f"Could not make item '{item_name}' visible (it may not exist): {e_show}")
        elif hidden_items is not None:
            # Ensure all are visible first if we are only specifying items to hide
            # This logic might need refinement based on desired behavior (additive vs. exclusive hide)
            # For simplicity now: just hide the specified items.
            for item_name in hidden_items:
                try:
                    pf.PivotItems(item_name).Visible = False
                except Exception as e_hide:
                    logger.warning(f"Could not hide item '{item_name}' (it may not exist): {e_hide}")
        else: # No items specified means show all
            pf.ClearAllFilters() # Simplest way to show all

        pivot_table_obj.PivotCache().Refresh() # Refresh to apply filter changes
        _close_excel_app(app, wb, save=True)
        return {"success": True, "message": f"Filter applied to PivotTable field '{field_name}'."}
    except Exception as e:
        logger.error(f"Failed to filter PivotTable items: {e}")
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Failed to filter PivotTable items: {str(e)}"}

def set_pivot_table_value_field_calculation(
    filepath: str,
    sheet_name: str,
    pivot_table_name: str,
    value_field_caption: str, # Current display name/caption of the value field (e.g., 'Sum of Sales')
    new_function: Optional[str] = None, # e.g., "Average", "Count", "Min", "Max"
    new_custom_name: Optional[str] = None, # New display name for the field
    show_value_as: Optional[str] = None, # e.g., "%OfGrandTotal", "DifferenceFrom"
    base_field_name: Optional[str] = None, # For calculations like "DifferenceFrom" or "%Of"
    base_item_value: Optional[str] = None  # Specific item for BaseField, or "(previous)", "(next)"
) -> Dict[str, Any]:
    """
    Changes the summary function, custom name, and/or "Show Values As" calculation 
    for an existing value field in a PivotTable.
    """
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings is required for PivotTable operations."}

    app, wb = _get_excel_app_and_workbook(filepath)
    if not wb:
        return {"success": False, "error": "Failed to open workbook with xlwings."}
    try:
        ws = wb.sheets[sheet_name]
        try:
            pivot_table_obj = ws.api.PivotTables(pivot_table_name)
        except Exception:
             return {"success": False, "error": f"PivotTable '{pivot_table_name}' not found."}

        data_field_to_change = None
        for df in pivot_table_obj.DataFields:
            if df.Name == value_field_caption: # df.Name is the caption, e.g. "Sum of Sales"
                data_field_to_change = df
                break
        
        if not data_field_to_change:
            valid_fields = [df.Name for df in pivot_table_obj.DataFields]
            return {"success": False, "error": f"Value field '{value_field_caption}' not found. Available: {valid_fields}"}

        # Change summary function
        if new_function:
            xl_function = FUNCTION_MAP.get(new_function.lower())
            if xl_function is None:
                return {"success": False, "error": f"Unsupported summary function: '{new_function}'. Supported: {list(FUNCTION_MAP.keys())}"}
            data_field_to_change.Function = xl_function
        
        # Change "Show Values As"
        effective_show_value_as = show_value_as.lower() if show_value_as else "normal" # Assume normal if not specified or if clearing
        
        calculation_type = SHOW_VALUE_AS_MAP.get(effective_show_value_as)
        if calculation_type is None: # Includes if show_value_as was None and defaulted to normal, or if invalid
            if show_value_as and show_value_as.lower() != "normal": # only warn if it was an actual attempt to set invalid calc
                 logger.warning(f"Unsupported or invalid 'show_value_as' type: {show_value_as}. Resetting to Normal.")
            data_field_to_change.Calculation = XL_NORMAL # Reset to normal if invalid or explicitly set to normal
        else:
            data_field_to_change.Calculation = calculation_type
            if base_field_name and calculation_type in [XL_DIFFERENCE_FROM, XL_PERCENT_OF, XL_PERCENT_DIFFERENCE_FROM, XL_RUNNING_TOTAL_IN]:
                try:
                    # Ensure BaseField is set using the Name property of the PivotField object
                    data_field_to_change.BaseField = pivot_table_obj.PivotFields(base_field_name).Name
                except Exception as bf_e:
                    msg = f"Could not set BaseField '{base_field_name}' for '{data_field_to_change.Name}': {bf_e}. Calculation might be incorrect."
                    logger.error(msg)
                    _close_excel_app(app, wb, save=False)
                    return {"success": False, "error": msg}

            if base_item_value and calculation_type in [XL_DIFFERENCE_FROM, XL_PERCENT_OF, XL_PERCENT_DIFFERENCE_FROM]:
                try:
                    if base_item_value.lower() == "(previous)":
                        data_field_to_change.api.BaseItem = "(previous)" # Using .api for special string values
                    elif base_item_value.lower() == "(next)":
                        data_field_to_change.api.BaseItem = "(next)"
                    else: # Specific item name
                        data_field_to_change.BaseItem = base_item_value
                except Exception as bi_e:
                    msg = f"Could not set BaseItem '{base_item_value}' for '{data_field_to_change.Name}': {bi_e}. Calculation might be incorrect."
                    logger.error(msg)
                    _close_excel_app(app, wb, save=False)
                    return {"success": False, "error": msg}
            elif calculation_type != XL_NORMAL and calculation_type in [XL_DIFFERENCE_FROM, XL_PERCENT_OF, XL_PERCENT_DIFFERENCE_FROM] and not (base_field_name and base_item_value):
                 # If a calculation requiring base field/item is set, but they are not provided, Excel might error or use defaults.
                 # For DifferenceFrom/PercentDifferenceFrom, BaseField and BaseItem are crucial.
                 # For PercentOf, only BaseField is strictly needed if BaseItem is not set (implies % of total for that base field item).
                 # For RunningTotalIn, only BaseField is needed.
                 if calculation_type in [XL_DIFFERENCE_FROM, XL_PERCENT_DIFFERENCE_FROM] and not (base_field_name and base_item_value):
                    logger.warning(f"Calculation '{show_value_as}' for field '{data_field_to_change.Name}' typically requires both BaseField and BaseItem.")
                 elif calculation_type == XL_PERCENT_OF and not base_field_name:
                     logger.warning(f"Calculation '{show_value_as}' for field '{data_field_to_change.Name}' typically requires BaseField.")


        # Change custom name (Caption)
        if new_custom_name:
            data_field_to_change.Name = new_custom_name # Name property is the caption for DataFields

        pivot_table_obj.PivotCache().Refresh()
        _close_excel_app(app, wb, save=True)
        final_caption = data_field_to_change.Name
        return {"success": True, "message": f"Value field '{value_field_caption}' (now '{final_caption}') calculation updated."}
    except Exception as e:
        logger.error(f"Failed to set PivotTable value field calculation: {e}")
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Failed to set PivotTable value field calculation for '{value_field_caption}': {str(e)}"}

def refresh_pivot_table(
    filepath: str,
    sheet_name: str,
    pivot_table_name: str
) -> Dict[str, Any]:
    """Refreshes a PivotTable to reflect changes in its source data."""
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings is required for PivotTable operations."}

    app, wb = _get_excel_app_and_workbook(filepath)
    if not wb:
        return {"success": False, "error": "Failed to open workbook with xlwings."}
    try:
        ws = wb.sheets[sheet_name]
        try:
            pivot_table_obj = ws.api.PivotTables(pivot_table_name)
        except Exception:
             return {"success": False, "error": f"PivotTable '{pivot_table_name}' not found."}
        
        pivot_table_obj.PivotCache().Refresh()
        
        _close_excel_app(app, wb, save=True)
        return {"success": True, "message": f"PivotTable '{pivot_table_name}' refreshed."}
    except Exception as e:
        logger.error(f"Failed to refresh PivotTable: {e}")
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Failed to refresh PivotTable: {str(e)}"}

# --- TODO: More advanced PivotTable features ---
# - Calculated Fields/Items
# - Slicers
# - PivotTable Layouts (Compact, Outline, Tabular) - (Basic version added in create/modify)
# - Grand Totals / Subtotals configuration
# - Formatting specific parts of the PivotTable
# - Changing PivotTable Data Source

def add_pivot_table_calculated_field(
    filepath: str,
    sheet_name: str,
    pivot_table_name: str,
    field_name: str,
    formula: str
) -> Dict[str, Any]:
    """Adds a calculated field to a PivotTable.
    Example: field_name='Profit', formula='=Revenue-Cost'
    """
    app, wb = None, None
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings or pywin32 is not installed. Please install them for PivotTable operations."}
    try:
        app, wb = _get_excel_app_and_workbook(filepath)
        if not wb:
            return {"success": False, "error": f"Failed to open workbook: {filepath}"}

        sheet = wb.sheets[sheet_name]
        pivot_table_obj = sheet.pivot_tables[pivot_table_name]
        
        # Check if field already exists
        try:
            if pivot_table_obj.pivotfields(field_name):
                return {"success": False, "error": f"Calculated field '{field_name}' already exists in PivotTable '{pivot_table_name}'."}
        except: # Field does not exist, can proceed
            pass

        pivot_table_obj.calculated_fields().add(name=field_name, formula=formula)
        
        _close_excel_app(app, wb, save=True)
        return {"success": True, "message": f"Calculated field '{field_name}' added to PivotTable '{pivot_table_name}'."}
    except Exception as e:
        logger.error(f"Failed to add calculated field: {e}")
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Failed to add calculated field to PivotTable '{pivot_table_name}': {str(e)}"}

def add_pivot_table_calculated_item(
    filepath: str,
    sheet_name: str,
    pivot_table_name: str,
    base_field_name: str, # The PivotField where the calculated item will be added
    item_name: str, # Name of the new calculated item
    formula: str # Formula for the calculated item, e.g., "='USA' + 'Canada'"
) -> Dict[str, Any]:
    """Adds a calculated item to a specific PivotField in a PivotTable."""
    app, wb = None, None
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings or pywin32 is not installed. Please install them for PivotTable operations."}
    try:
        app, wb = _get_excel_app_and_workbook(filepath)
        if not wb:
            return {"success": False, "error": f"Failed to open workbook: {filepath}"}

        sheet = wb.sheets[sheet_name]
        pivot_table_obj = sheet.pivot_tables[pivot_table_name]
        
        pivot_field = pivot_table_obj.pivotfields(base_field_name)
        if not pivot_field:
            return {"success": False, "error": f"PivotField '{base_field_name}' not found in PivotTable '{pivot_table_name}'."}

        # Check if item already exists
        try:
            if pivot_field.pivotitems(item_name):
                 return {"success": False, "error": f"Calculated item '{item_name}' already exists in field '{base_field_name}'."}
        except: # Item does not exist
            pass
            
        pivot_field.calculated_items().add(name=item_name, formula=formula)
        
        _close_excel_app(app, wb, save=True)
        return {"success": True, "message": f"Calculated item '{item_name}' added to field '{base_field_name}' in PivotTable '{pivot_table_name}'."}
    except Exception as e:
        logger.error(f"Failed to add calculated item: {e}")
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Failed to add calculated item to PivotTable '{pivot_table_name}': {str(e)}"}

def create_pivot_table_slicer(
    filepath: str,
    sheet_name: str, # Sheet where slicer will be placed
    pivot_table_name: str, # Name of the PivotTable to connect the slicer to
    slicer_field_name: str, # Field from PivotTable to use for slicer
    slicer_name: Optional[str] = None, # Optional name for the slicer object
    top: Optional[float] = None, # Position from top (in points)
    left: Optional[float] = None, # Position from left (in points)
    width: Optional[float] = None, # Width of the slicer (in points)
    height: Optional[float] = None # Height of the slicer (in points)
) -> Dict[str, Any]:
    """Creates a slicer for a PivotTable field."""
    app, wb = None, None
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings or pywin32 is not installed. Please install them for PivotTable operations."}
    try:
        app, wb = _get_excel_app_and_workbook(filepath)
        if not wb:
            return {"success": False, "error": f"Failed to open workbook: {filepath}"}

        # Find the PivotTable by searching all sheets if sheet_name is ambiguous or PivotTable is on different sheet
        pt_sheet = None
        pivot_table_obj = None
        for s in wb.sheets:
            try:
                pt = s.pivot_tables[pivot_table_name]
                if pt:
                    pivot_table_obj = pt
                    pt_sheet = s # sheet where pivot table actually exists
                    break
            except:
                continue
        
        if not pivot_table_obj:
            return {"success": False, "error": f"PivotTable '{pivot_table_name}' not found in the workbook."}

        slicer_cache = wb.api.SlicerCaches.Add2(pivot_table_obj.api, slicer_field_name)
        
        # Sheet where the slicer will be placed
        target_sheet_api = wb.sheets[sheet_name].api

        slicer_obj_name = slicer_name if slicer_name else f"Slicer_{slicer_field_name.replace(' ', '_')}"

        slicer = slicer_cache.Slicers.Add(
            SlicerDestination=target_sheet_api,
            Name=slicer_obj_name,
            Caption=slicer_field_name # Caption is what user sees
        )
        
        # Set optional properties
        if top is not None: slicer.Top = top
        if left is not None: slicer.Left = left
        if width is not None: slicer.Width = width
        if height is not None: slicer.Height = height
        
        _close_excel_app(app, wb, save=True)
        return {"success": True, "message": f"Slicer '{slicer_obj_name}' for field '{slicer_field_name}' created and linked to PivotTable '{pivot_table_name}' on sheet '{sheet_name}'.", "slicer_name": slicer.Name}
    except Exception as e:
        logger.error(f"Failed to create slicer: {e}")
        _close_excel_app(app, wb, save=False)
        # Attempt to provide more specific error messages
        if "SlicerCaches.Add" in str(e) or "PivotTable field" in str(e).lower():
            return {"success": False, "error": f"Failed to create slicer for PivotTable '{pivot_table_name}': Could not connect to field '{slicer_field_name}'. Ensure field exists and is suitable for a slicer. Details: {str(e)}"}
        return {"success": False, "error": f"Failed to create slicer for PivotTable '{pivot_table_name}': {str(e)}"}

def modify_pivot_table_slicer(
    filepath: str,
    sheet_name: str, # Sheet where the slicer is located
    slicer_name: str,
    selected_items: Optional[List[str]] = None, # List of item names to select. If None, no change. If empty list, deselects all.
    slicer_style: Optional[str] = None, # e.g., "SlicerStyleLight1"
    caption: Optional[str] = None,
    top: Optional[float] = None,
    left: Optional[float] = None,
    width: Optional[float] = None,
    height: Optional[float] = None,
    number_of_columns: Optional[int] = None
) -> Dict[str, Any]:
    """Modifies properties of an existing PivotTable slicer."""
    app, wb = None, None
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings or pywin32 is not installed. Please install them for PivotTable operations."}
    try:
        app, wb = _get_excel_app_and_workbook(filepath)
        if not wb:
            return {"success": False, "error": f"Failed to open workbook: {filepath}"}

        sheet = wb.sheets[sheet_name]
        slicer = None
        try:
            slicer = sheet.api.Slicers(slicer_name)
        except Exception:
            return {"success": False, "error": f"Slicer '{slicer_name}' not found on sheet '{sheet_name}'."}

        if selected_items is not None:
            slicer.SlicerCache.ClearAllFilters() # Clear existing selections
            if selected_items: # If list is not empty, select specified items
                for item_name in selected_items:
                    try:
                        slicer.SlicerCache.SlicerItems(item_name).Selected = True
                    except Exception as item_e:
                        logger.warning(f"Could not select item '{item_name}' in slicer '{slicer_name}': {item_e}")
                        # Optionally return partial success or accumulate errors
        
        if slicer_style: slicer.Style = slicer_style
        if caption: slicer.Caption = caption
        if top is not None: slicer.Top = top
        if left is not None: slicer.Left = left
        if width is not None: slicer.Width = width
        if height is not None: slicer.Height = height
        if number_of_columns is not None: slicer.NumberOfColumns = number_of_columns
        
        _close_excel_app(app, wb, save=True)
        return {"success": True, "message": f"Slicer '{slicer_name}' on sheet '{sheet_name}' modified."}
    except Exception as e:
        logger.error(f"Failed to modify slicer: {e}")
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Failed to modify slicer '{slicer_name}': {str(e)}"}

def set_pivot_table_layout(
    filepath: str,
    sheet_name: str,
    pivot_table_name: str,
    layout_type: str, # "compact", "outline", "tabular"
    repeat_all_item_labels: Optional[bool] = None # For outline/tabular
) -> Dict[str, Any]:
    """Sets the report layout for a PivotTable."""
    app, wb = None, None
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings or pywin32 is not installed. Please install them for PivotTable operations."}
    try:
        app, wb = _get_excel_app_and_workbook(filepath)
        if not wb:
            return {"success": False, "error": f"Failed to open workbook: {filepath}"}

        sheet = wb.sheets[sheet_name]
        pivot_table_obj = sheet.pivot_tables[pivot_table_name]
        
        layout_map = {
            "compact": 1, # xlCompactRow
            "outline": 2, # xlOutlineRow
            "tabular": 3  # xlTabularRow
        }
        
        if layout_type.lower() not in layout_map:
            return {"success": False, "error": f"Invalid layout_type: '{layout_type}'. Must be one of {list(layout_map.keys())}."}

        pivot_table_obj.api.RowAxisLayout(layout_map[layout_type.lower()])

        if repeat_all_item_labels is not None and layout_type.lower() in ["outline", "tabular"]:
            for pf in pivot_table_obj.pivotfields(): # Iterate through all pivot fields
                if pf.orientation == xw.constants.PivotFieldOrientation.xlRowField: #Check if it is a row field
                     pf.api.RepeatLabels = repeat_all_item_labels
        
        _close_excel_app(app, wb, save=True)
        return {"success": True, "message": f"Layout of PivotTable '{pivot_table_name}' set to '{layout_type}'.", "layout_set": layout_type}
    except Exception as e:
        logger.error(f"Failed to set PivotTable layout: {e}")
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Failed to set layout for PivotTable '{pivot_table_name}': {str(e)}"}

def configure_pivot_table_totals(
    filepath: str,
    sheet_name: str,
    pivot_table_name: str,
    grand_totals_for_rows: Optional[bool] = None, # True to show, False to hide
    grand_totals_for_columns: Optional[bool] = None,
    subtotals_settings: Optional[List[Dict[str, Any]]] = None # Enhanced! e.g. [
    # {"field_name": "Category", "show": False}, # Hide all subtotals
    # {"field_name": "Region", "show": True, "functions": ["Sum", "Average"]} # Show specific subtotals only
    # ]
) -> Dict[str, Any]:
    """
    Configures grand totals and subtotals for a PivotTable.
    
    Args:
        filepath: Path to the Excel workbook
        sheet_name: Name of the worksheet containing the PivotTable
        pivot_table_name: Name of the PivotTable
        grand_totals_for_rows: True to show grand totals for rows, False to hide
        grand_totals_for_columns: True to show grand totals for columns, False to hide
        subtotals_settings: List of dictionaries specifying subtotal settings for fields:
          - field_name: Name of the field
          - show: True to show subtotals, False to hide all subtotals
          - functions: Optional list of specific functions for subtotals (Sum, Count, Average, etc.)
                      If not provided but show=True, uses the default (same as the data field)
    
    Returns:
        Dictionary with operation result
    """
    app, wb = None, None
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings or pywin32 is not installed. Please install them for PivotTable operations."}
        
    try:
        app, wb = _get_excel_app_and_workbook(filepath)
        if not wb:
            return {"success": False, "error": f"Failed to open workbook: {filepath}"}
            
        sheet = wb.sheets[sheet_name]
        try:
            pivot_table_obj = sheet.api.PivotTables(pivot_table_name)
        except Exception:
            return {"success": False, "error": f"PivotTable '{pivot_table_name}' not found."}
            
        # Configure grand totals
        if grand_totals_for_rows is not None:
            pivot_table_obj.RowGrand = grand_totals_for_rows
            
        if grand_totals_for_columns is not None:
            pivot_table_obj.ColumnGrand = grand_totals_for_columns
        
        # Map of function names to their index in the Excel API's Subtotals array
        # In Excel API, Subtotals is a 1-indexed array with 12 elements for different functions
        subtotal_function_indices = {
            "automatic": 1,  # Use same function as the data field
            "sum": 2,
            "count": 3,
            "average": 4,
            "max": 5,
            "min": 6,
            "product": 7,
            "countnumbers": 8,
            "stddev": 9,
            "stddevp": 10,
            "var": 11,
            "varp": 12
        }
            
        # Configure subtotals
        if subtotals_settings:
            modified_fields = []
            
            for setting in subtotals_settings:
                field_name = setting["field_name"]
                show = setting.get("show", True)  # Default to showing subtotals
                functions = setting.get("functions", [])  # Optional list of specific subtotal functions
                
                try:
                    pivot_field = pivot_table_obj.PivotFields(field_name)
                    
                    # Check if this is a row/column field (only these can have subtotals)
                    # if not (pivot_field.Orientation == 1 or pivot_field.Orientation == 2):  # XL_ROW_FIELD or XL_COLUMN_FIELD
                    #     logger.warning(f"Field '{field_name}' is not a row or column field. Subtotals are only applicable to row/column fields.")
                    #     continue
                    
                    if not show:
                        # Hide all subtotals for this field
                        pivot_field.api.Subtotals = [False] * 12  # All False = no subtotals
                        modified_fields.append(f"{field_name} (all subtotals hidden)")
                    elif not functions:
                        # Show default subtotal (Automatic - same function as data field)
                        # First, reset all to False
                        pivot_field.api.Subtotals = [False] * 12
                        # Then set Automatic to True (it's the first item in the array, index 1)
                        subtotals_array = pivot_field.api.Subtotals
                        subtotals_array[0] = True  # Index 0 in Python = Index 1 in VBA (Automatic subtotal)
                        pivot_field.api.Subtotals = subtotals_array
                        modified_fields.append(f"{field_name} (automatic subtotal)")
                    else:
                        # Show specific subtotal functions
                        # First, reset all to False
                        subtotals_array = [False] * 12
                        
                        for function in functions:
                            function_lower = function.lower()
                            if function_lower in subtotal_function_indices:
                                # Convert to 0-based index for Python
                                index = subtotal_function_indices[function_lower] - 1
                                subtotals_array[index] = True
                            else:
                                logger.warning(f"Unsupported subtotal function: {function}. Supported functions: {list(subtotal_function_indices.keys())}")
                        
                        # Apply the customized subtotals array
                        pivot_field.api.Subtotals = subtotals_array
                        modified_fields.append(f"{field_name} (custom subtotals: {', '.join(functions)})")
                        
                except Exception as e:
                    logger.warning(f"Failed to configure subtotals for field '{field_name}': {str(e)}")
            
            # Refresh PivotTable to apply changes
            pivot_table_obj.PivotCache().Refresh()
            _close_excel_app(app, wb, save=True)
            
            return {
                "success": True,
                "message": f"PivotTable '{pivot_table_name}' totals configuration updated. Modified fields: {modified_fields}",
                "modified_fields": modified_fields
            }
        else:
            # If only grand totals were modified
            pivot_table_obj.PivotCache().Refresh()
            _close_excel_app(app, wb, save=True)
            
            grand_total_changes = []
            if grand_totals_for_rows is not None:
                grand_total_changes.append(f"row grand totals {'shown' if grand_totals_for_rows else 'hidden'}")
            if grand_totals_for_columns is not None:
                grand_total_changes.append(f"column grand totals {'shown' if grand_totals_for_columns else 'hidden'}")
                
            return {
                "success": True,
                "message": f"PivotTable '{pivot_table_name}' grand totals updated: {', '.join(grand_total_changes)}"
            }
            
    except Exception as e:
        logger.error(f"Failed to configure PivotTable totals: {e}")
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Failed to configure PivotTable totals: {str(e)}"}

def format_pivot_table_part(
    filepath: str,
    sheet_name: str,
    pivot_table_name: str,
    part_to_format: str, # "data_body_range", "row_header_range", "column_header_range", "page_field_range", "grand_total_range"
    font_bold: Optional[bool] = None,
    font_italic: Optional[bool] = None,
    font_size: Optional[int] = None,
    font_color_rgb: Optional[Tuple[int, int, int]] = None, # e.g., (255, 0, 0) for red
    bg_color_rgb: Optional[Tuple[int, int, int]] = None,
    horizontal_alignment: Optional[str] = None # "left", "center", "right", "general"
) -> Dict[str, Any]:
    """Applies formatting to specific parts of a PivotTable."""
    app, wb = None, None
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings or pywin32 is not installed. Please install them for PivotTable operations."}
    try:
        app, wb = _get_excel_app_and_workbook(filepath)
        if not wb:
            return {"success": False, "error": f"Failed to open workbook: {filepath}"}

        sheet = wb.sheets[sheet_name]
        pivot_table_obj = sheet.pivot_tables[pivot_table_name]
        
        target_range_api = None
        part_map = {
            "data_body_range": pivot_table_obj.api.DataBodyRange,
            "row_header_range": pivot_table_obj.api.RowRange, # Includes headers and items
            "column_header_range": pivot_table_obj.api.ColumnRange, # Includes headers and items
            # PageFields are a bit trickier as they are separate controls typically
            # "grand_total_range": Accessing specific grand total cells requires more specific logic
        }

        if part_to_format.lower() in part_map:
            target_range_api = part_map[part_to_format.lower()]
        elif part_to_format.lower() == "page_field_range":
             # Iterate page fields and format their labels if possible
            # This is a simplification, real page field formatting can be complex
            formatted_count = 0
            for pf_api in pivot_table_obj.api.PageFields:
                # Try to format the cell containing the page field if it's simple
                # This might not cover all scenarios of page field placement.
                try:
                    # Attempt to get a range associated with the page field - this is non-standard
                    # Excel typically doesn't expose a direct range for the whole page field control container.
                    # We might need to infer the cell from pivot_table_obj.api.PageRangeCells if available,
                    # or iterate slicers if page fields are implemented as slicers.
                    # For now, we'll log a warning.
                    logger.warning(f"Direct formatting of 'page_field_range' for '{pf_api.Caption}' is complex and might not apply as expected via this simplified function.")

                except Exception as pf_e:
                     logger.warning(f"Could not format page field '{pf_api.Caption}': {pf_e}")
            if formatted_count > 0:
                 _close_excel_app(app, wb, save=True)
                 return {"success": True, "message": f"Attempted formatting for page fields in PivotTable '{pivot_table_name}'."}
            else:
                 _close_excel_app(app, wb, save=False)
                 return {"success": False, "error": f"Could not reliably format 'page_field_range' for PivotTable '{pivot_table_name}'. Consider formatting cells directly."}


        else:
            return {"success": False, "error": f"Invalid 'part_to_format': {part_to_format}. Supported parts: {list(part_map.keys())}."}

        if not target_range_api:
            return {"success": False, "error": f"Could not get range for '{part_to_format}' in PivotTable '{pivot_table_name}'."}

        if font_bold is not None: target_range_api.Font.Bold = font_bold
        if font_italic is not None: target_range_api.Font.Italic = font_italic
        if font_size is not None: target_range_api.Font.Size = font_size
        if font_color_rgb: target_range_api.Font.Color = xw.utils.rgb_to_int(font_color_rgb)
        if bg_color_rgb: target_range_api.Interior.Color = xw.utils.rgb_to_int(bg_color_rgb)
        
        if horizontal_alignment:
            align_map = {"left": -4131, "center": -4108, "right": -4152, "general": 1} # xlLeft, xlCenter, xlRight, xlGeneral
            if horizontal_alignment.lower() in align_map:
                target_range_api.HorizontalAlignment = align_map[horizontal_alignment.lower()]
        
        _close_excel_app(app, wb, save=True)
        return {"success": True, "message": f"Formatting applied to '{part_to_format}' of PivotTable '{pivot_table_name}'."}
    except Exception as e:
        logger.error(f"Failed to format PivotTable part: {e}")
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Failed to format '{part_to_format}' for PivotTable '{pivot_table_name}': {str(e)}"}

def change_pivot_table_data_source(
    filepath: str,
    sheet_name: str,
    pivot_table_name: str,
    new_source_data: str # e.g., "Sheet1!A1:H200" or "TableName"
) -> Dict[str, Any]:
    """Changes the data source for an existing PivotTable."""
    app, wb = None, None
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings or pywin32 is not installed. Please install them for PivotTable operations."}
    try:
        app, wb = _get_excel_app_and_workbook(filepath)
        if not wb:
            return {"success": False, "error": f"Failed to open workbook: {filepath}"}

        sheet = wb.sheets[sheet_name]
        pivot_table_obj = sheet.pivot_tables[pivot_table_name]
        
        # The SourceData property expects a string like "'Sheet Name'!R1C1:R100C5" or "TableName"
        # If the new_source_data is a sheet range, ensure it's correctly formatted.
        # xlwings/pywin32 might handle typical "Sheet1!A1:H200" format directly.
        
        pivot_table_obj.api.SourceData = new_source_data
        pivot_table_obj.pivotcache().refresh() # Refresh after changing source
        
        _close_excel_app(app, wb, save=True)
        return {"success": True, "message": f"Data source for PivotTable '{pivot_table_name}' changed to '{new_source_data}' and refreshed."}
    except Exception as e:
        logger.error(f"Failed to change PivotTable data source: {e}")
        _close_excel_app(app, wb, save=False)
        # More specific error if source data is invalid
        if "invalid source" in str(e).lower() or "reference is not valid" in str(e).lower():
            return {"success": False, "error": f"Failed to change data source for PivotTable '{pivot_table_name}': The new source data '{new_source_data}' is invalid or not found. Details: {str(e)}"}
        return {"success": False, "error": f"Failed to change data source for PivotTable '{pivot_table_name}': {str(e)}"}

def group_pivot_field_items(
    filepath: str,
    sheet_name: str,
    pivot_table_name: str,
    field_name: str,
    group_type: str, # "date", "numeric", "selection"
    start_value: Optional[Any] = None, # For numeric grouping
    end_value: Optional[Any] = None,   # For numeric grouping
    interval: Optional[float] = None,  # For numeric grouping
    date_parts: Optional[Dict[str, bool]] = None, # For date grouping, e.g., {"years": True, "quarters": True, "months": True}
    selected_items: Optional[List[str]] = None # For selection grouping (text items)
) -> Dict[str, Any]:
    """
    Groups items in a PivotTable field based on specified criteria.
    
    Args:
        filepath: Path to the Excel workbook
        sheet_name: Name of the worksheet containing the PivotTable
        pivot_table_name: Name of the PivotTable
        field_name: Name of the field to group
        group_type: Type of grouping - "date", "numeric", or "selection"
        start_value: For numeric grouping, the start value of the range
        end_value: For numeric grouping, the end value of the range
        interval: For numeric grouping, the interval size
        date_parts: For date grouping, which date parts to include (years, quarters, months, days)
        selected_items: For selection grouping, the items to group together
    
    Returns:
        Dictionary with operation result
    """
    app, wb = None, None
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings or pywin32 is not installed. Please install them for PivotTable operations."}
    
    try:
        app, wb = _get_excel_app_and_workbook(filepath)
        if not wb:
            return {"success": False, "error": f"Failed to open workbook: {filepath}"}

        sheet = wb.sheets[sheet_name]
        try:
            pivot_table_obj = sheet.api.PivotTables(pivot_table_name)
        except Exception:
            return {"success": False, "error": f"PivotTable '{pivot_table_name}' not found."}

        # Find the field to group
        try:
            pivot_field = pivot_table_obj.PivotFields(field_name)
        except Exception:
            return {"success": False, "error": f"Field '{field_name}' not found in PivotTable '{pivot_table_name}'."}

        # Set the correct group operation based on the group_type
        if group_type.lower() == "date":
            # Date grouping - group by Years, Quarters, Months, etc.
            # Default to grouping by all date parts if none specified
            if not date_parts:
                date_parts = {
                    "seconds": False, 
                    "minutes": False, 
                    "hours": False, 
                    "days": True, 
                    "months": True, 
                    "quarters": True, 
                    "years": True
                }
                
            # Convert date_parts to integers (0 or 1) for the GroupBy API
            # In the VBA API, the order is [Seconds, Minutes, Hours, Days, Months, Quarters, Years]
            # But in xlwings, we may need to use the named constants (checking to be safe)
            try:
                # Create an array of boolean values in the right order expected by Excel API
                group_array = [
                    date_parts.get("seconds", False),  # 0
                    date_parts.get("minutes", False),  # 1
                    date_parts.get("hours", False),    # 2
                    date_parts.get("days", False),     # 3
                    date_parts.get("months", False),   # 4
                    date_parts.get("quarters", False), # 5
                    date_parts.get("years", False)     # 6
                ]
                
                # Group by date - the API expects True/False for each date part
                # The True/False array is converted to 1's and 0's by the interop layer
                pivot_field.api.Group(GroupBy=group_array)
                
                # Get which date parts were used
                grouped_parts = [name for name, value in date_parts.items() if value]
                
                _close_excel_app(app, wb, save=True)
                return {
                    "success": True, 
                    "message": f"Field '{field_name}' grouped by date: {', '.join(grouped_parts)}.",
                    "group_type": "date",
                    "grouped_parts": grouped_parts
                }
                
            except Exception as e:
                _close_excel_app(app, wb, save=False)
                return {
                    "success": False, 
                    "error": f"Failed to group field '{field_name}' by date: {str(e)}"
                }
                
        elif group_type.lower() == "numeric":
            # Numeric grouping - group by numeric ranges
            if start_value is None or end_value is None or interval is None:
                _close_excel_app(app, wb, save=False)
                return {
                    "success": False, 
                    "error": "For numeric grouping, start_value, end_value, and interval must be provided."
                }
                
            try:
                # Convert to float for numeric grouping
                start_val = float(start_value)
                end_val = float(end_value)
                interval_val = float(interval)
                
                # Group using the numeric parameters
                pivot_field.api.Group(Start=start_val, End=end_val, By=interval_val)
                
                _close_excel_app(app, wb, save=True)
                return {
                    "success": True, 
                    "message": f"Field '{field_name}' grouped by numeric ranges: start={start_val}, end={end_val}, interval={interval_val}.",
                    "group_type": "numeric",
                    "start": start_val,
                    "end": end_val,
                    "interval": interval_val
                }
                
            except Exception as e:
                _close_excel_app(app, wb, save=False)
                return {
                    "success": False, 
                    "error": f"Failed to group field '{field_name}' by numeric ranges: {str(e)}"
                }
                
        elif group_type.lower() == "selection":
            # Selection grouping - group selected items
            if not selected_items or len(selected_items) < 2:
                _close_excel_app(app, wb, save=False)
                return {
                    "success": False, 
                    "error": "For selection grouping, at least two items must be provided in selected_items."
                }
                
            try:
                # Get all pivot items in the field
                field_items = [item.Name for item in pivot_field.PivotItems]
                
                # Check if all selected items exist in the field
                invalid_items = [item for item in selected_items if item not in field_items]
                if invalid_items:
                    _close_excel_app(app, wb, save=False)
                    return {
                        "success": False, 
                        "error": f"Items not found in field '{field_name}': {invalid_items}. Available items: {field_items}"
                    }
                
                # In Excel API, we need to select the items before grouping them
                # This requires direct manipulation of Field.PivotItems collection
                for item_name in selected_items:
                    # Get the PivotItem object
                    item = pivot_field.PivotItems(item_name)
                    # Set the item as selected
                    item.api.Selected = True
                
                # Group the selected items
                pivot_field.api.CreateGroup()
                
                # Reset selection
                for item_name in field_items:
                    if item_name in selected_items:
                        continue # Skip the items we just grouped
                    try:
                        item = pivot_field.PivotItems(item_name)
                        item.api.Selected = False
                    except Exception:
                        # Item might no longer exist or be accessible after grouping
                        pass
                
                _close_excel_app(app, wb, save=True)
                return {
                    "success": True, 
                    "message": f"Items {selected_items} grouped in field '{field_name}'.",
                    "group_type": "selection",
                    "grouped_items": selected_items
                }
                
            except Exception as e:
                _close_excel_app(app, wb, save=False)
                return {
                    "success": False, 
                    "error": f"Failed to group selected items in field '{field_name}': {str(e)}"
                }
                
        else:
            # Invalid group type
            _close_excel_app(app, wb, save=False)
            return {
                "success": False, 
                "error": f"Invalid group_type: {group_type}. Must be 'date', 'numeric', or 'selection'."
            }
            
    except Exception as e:
        logger.error(f"Failed to group pivot field items: {e}")
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Failed to group pivot field items: {str(e)}"}


def ungroup_pivot_field_items(
    filepath: str,
    sheet_name: str,
    pivot_table_name: str,
    field_name: str,
    group_name: Optional[str] = None  # If provided, ungroup only this specific group
) -> Dict[str, Any]:
    """
    Removes grouping from a PivotTable field.
    
    Args:
        filepath: Path to the Excel workbook
        sheet_name: Name of the worksheet containing the PivotTable
        pivot_table_name: Name of the PivotTable
        field_name: Name of the field to ungroup
        group_name: Optional name of a specific group to ungroup. If None, ungroups all.
    
    Returns:
        Dictionary with operation result
    """
    app, wb = None, None
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings or pywin32 is not installed. Please install them for PivotTable operations."}
    
    try:
        app, wb = _get_excel_app_and_workbook(filepath)
        if not wb:
            return {"success": False, "error": f"Failed to open workbook: {filepath}"}

        sheet = wb.sheets[sheet_name]
        try:
            pivot_table_obj = sheet.api.PivotTables(pivot_table_name)
        except Exception:
            return {"success": False, "error": f"PivotTable '{pivot_table_name}' not found."}

        # Find the field to ungroup
        try:
            pivot_field = pivot_table_obj.PivotFields(field_name)
        except Exception:
            return {"success": False, "error": f"Field '{field_name}' not found in PivotTable '{pivot_table_name}'."}

        # Check if the field is actually grouped
        is_grouped = False
        try:
            # Check if field has an IsCalculated property (indicates it might be grouped)
            # or if it has a non-empty Groups collection
            is_grouped = hasattr(pivot_field, "IsGroupField") and pivot_field.IsGroupField
        except Exception:
            pass
            
        if not is_grouped:
            _close_excel_app(app, wb, save=False)
            return {
                "success": False, 
                "error": f"Field '{field_name}' is not grouped."
            }
            
        if group_name:
            # Ungroup a specific group
            try:
                # Find the group among PivotItems
                found = False
                for item in pivot_field.PivotItems:
                    if item.Name == group_name:
                        item.api.Ungroup()
                        found = True
                        break
                        
                if not found:
                    _close_excel_app(app, wb, save=False)
                    return {
                        "success": False, 
                        "error": f"Group '{group_name}' not found in field '{field_name}'."
                    }
                    
                _close_excel_app(app, wb, save=True)
                return {
                    "success": True, 
                    "message": f"Group '{group_name}' was ungrouped in field '{field_name}'."
                }
                
            except Exception as e:
                _close_excel_app(app, wb, save=False)
                return {
                    "success": False, 
                    "error": f"Failed to ungroup '{group_name}' in field '{field_name}': {str(e)}"
                }
                
        else:
            # Ungroup the entire field (all groups)
            try:
                # In Excel, ungrouping at the field level removes all grouping
                pivot_field.api.Ungroup()
                
                _close_excel_app(app, wb, save=True)
                return {
                    "success": True, 
                    "message": f"All groups removed from field '{field_name}'."
                }
                
            except Exception as e:
                _close_excel_app(app, wb, save=False)
                return {
                    "success": False, 
                    "error": f"Failed to ungroup field '{field_name}': {str(e)}"
                }
                
    except Exception as e:
        logger.error(f"Failed to ungroup pivot field: {e}")
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Failed to ungroup pivot field: {str(e)}"}

def apply_pivot_table_conditional_formatting(
    filepath: str,
    sheet_name: str,
    pivot_table_name: str,
    formatting_scope: str, # "data_field", "field_items", "grand_totals", "subtotals"
    field_name: str, # For data_field: the value field caption (e.g. "Sum of Sales")
                    # For field_items: the field to format items for (e.g. "Region")
    condition_type: str, # "top_bottom", "greater_than", "less_than", "between", "equal_to", "contains", "date_occurring"
    condition_parameters: Dict[str, Any] = None, # e.g. {"rank": 5, "type": "top", "percent": True} for top 5%
                                                # or {"value": 1000} for ">1000"
                                                # or {"start": 1000, "end": 2000} for "between"
    format_settings: Dict[str, Any] = None, # Formatting to apply: {"bold": True, "bg_color": (255,0,0), etc}
    specific_items: Optional[List[str]] = None # For field_items, optionally limit to these specific items
) -> Dict[str, Any]:
    """
    Apply advanced conditional formatting specifically designed for PivotTables.
    
    This function handles PivotTable-specific formatting with awareness of the dynamic
    structure of PivotTables. It can target specific value fields, field items, and totals.
    
    Args:
        filepath: Path to the Excel workbook
        sheet_name: Name of the worksheet containing the PivotTable
        pivot_table_name: Name of the PivotTable
        formatting_scope: Where to apply formatting - "data_field", "field_items", "grand_totals", or "subtotals"
        field_name: Name of field to format (interpretation depends on formatting_scope)
        condition_type: Type of condition for formatting - "top_bottom", "greater_than", "less_than", 
                        "between", "equal_to", "contains", "date_occurring"
        condition_parameters: Dictionary with parameters for the condition:
                             For "top_bottom": {"rank": 5, "type": "top"|"bottom", "percent": True|False}
                             For "greater_than"/"less_than": {"value": 1000}
                             For "between": {"start": 1000, "end": 2000}
                             For "equal_to": {"value": "East"}
                             For "contains": {"text": "America"}
        format_settings: Dictionary with formatting to apply:
                        {"bold": True, "bg_color_rgb": (255,0,0), "font_color_rgb": (255,255,255),
                         "font_size": 12, "font_italic": False, "horizontal_alignment": "center"}
        specific_items: For field_items scope, optionally limit to specific items in the field
    
    Returns:
        Dictionary with operation result
    """
    app, wb = None, None
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings or pywin32 is not installed. Please install them for PivotTable operations."}
    
    try:
        app, wb = _get_excel_app_and_workbook(filepath)
        if not wb:
            return {"success": False, "error": f"Failed to open workbook: {filepath}"}
            
        sheet = wb.sheets[sheet_name]
        try:
            pivot_table_obj = sheet.api.PivotTables(pivot_table_name)
        except Exception:
            return {"success": False, "error": f"PivotTable '{pivot_table_name}' not found."}
            
        # Validate and prepare format settings
        format_props = {}
        if format_settings:
            if "bold" in format_settings:
                format_props["bold"] = bool(format_settings["bold"])
                
            if "font_italic" in format_settings:
                format_props["italic"] = bool(format_settings["font_italic"])
                
            if "font_size" in format_settings:
                format_props["size"] = int(format_settings["font_size"])
                
            if "font_color_rgb" in format_settings:
                rgb = format_settings["font_color_rgb"]
                if isinstance(rgb, (tuple, list)) and len(rgb) == 3:
                    # Convert RGB tuple to hex color string
                    hex_color = f"{rgb[0]:02x}{rgb[1]:02x}{rgb[2]:02x}"
                    format_props["color"] = hex_color
                    
            if "bg_color_rgb" in format_settings:
                rgb = format_settings["bg_color_rgb"]
                if isinstance(rgb, (tuple, list)) and len(rgb) == 3:
                    # Create a PatternFill object with the RGB color
                    from openpyxl.styles import PatternFill
                    bg_hex = f"{rgb[0]:02x}{rgb[1]:02x}{rgb[2]:02x}"
                    format_props["fill"] = PatternFill(
                        start_color=bg_hex,
                        end_color=bg_hex,
                        fill_type="solid"
                    )
                    
            if "horizontal_alignment" in format_settings:
                from openpyxl.styles import Alignment
                halign = format_settings["horizontal_alignment"]
                if halign in ["left", "center", "right", "general"]:
                    format_props["alignment"] = Alignment(horizontal=halign)
        
        # Function to apply formatting to a range
        def apply_formatting_to_range(range_obj):
            if not format_props:
                return
                
            # Convert the xlwings Range to the underlying Excel Range object
            excel_range = range_obj.api
            
            # Apply font formatting
            if "bold" in format_props:
                excel_range.Font.Bold = format_props["bold"]
                
            if "italic" in format_props:
                excel_range.Font.Italic = format_props["italic"]
                
            if "size" in format_props:
                excel_range.Font.Size = format_props["size"]
                
            if "color" in format_props:
                # Apply font color (hex string)
                excel_range.Font.Color = int(f"0x{format_props['color']}", 16)
                
            # Apply fill (background color)
            if "fill" in format_props:
                # Extract the hex color from our fill
                fill_color = format_props["fill"].start_color.rgb
                if fill_color.startswith("FF"):  # Excel uses ARGB with FF for solid
                    fill_color = fill_color[2:]  # Remove FF prefix if present
                excel_range.Interior.Color = int(f"0x{fill_color}", 16)
                
            # Apply alignment
            if "alignment" in format_props:
                if format_props["alignment"].horizontal == "left":
                    excel_range.HorizontalAlignment = -4131  # xlLeft
                elif format_props["alignment"].horizontal == "center":
                    excel_range.HorizontalAlignment = -4108  # xlCenter
                elif format_props["alignment"].horizontal == "right":
                    excel_range.HorizontalAlignment = -4152  # xlRight
                else:  # "general" or default
                    excel_range.HorizontalAlignment = -4130  # xlGeneral
        
        # Process different formatting scopes
        if formatting_scope == "data_field":
            # Find the data field by caption
            try:
                data_field = None
                for df in pivot_table_obj.DataFields:
                    if df.Name == field_name:  # Name is the caption for DataFields
                        data_field = df
                        break
                
                if not data_field:
                    _close_excel_app(app, wb, save=False)
                    return {
                        "success": False,
                        "error": f"Data field with caption '{field_name}' not found in PivotTable '{pivot_table_name}'."
                    }
                
                # Get the DataBodyRange (the range containing all data values)
                data_body_range = pivot_table_obj.DataBodyRange
                if not data_body_range:
                    _close_excel_app(app, wb, save=False)
                    return {
                        "success": False,
                        "error": f"PivotTable '{pivot_table_name}' has no data body range. It might be empty."
                    }
                
                # Convert to xlwings Range object
                data_range = sheet.range(data_body_range.Address)
                
                # Create the condition expression based on condition_type
                if condition_type == "top_bottom":
                    # Parameters: {"rank": 5, "type": "top", "percent": True}
                    if not condition_parameters or "rank" not in condition_parameters:
                        _close_excel_app(app, wb, save=False)
                        return {"success": False, "error": "Missing 'rank' parameter for top_bottom condition."}
                    
                    rank = condition_parameters["rank"]
                    is_percent = condition_parameters.get("percent", False)
                    is_top = condition_parameters.get("type", "top").lower() == "top"
                    
                    if is_top:
                        if is_percent:
                            # Top N%
                            condition_formula = data_range.api.FormatConditions.AddTopPercentile(
                                Rank=rank, Type=True  # True for Top
                            )
                        else:
                            # Top N Items
                            condition_formula = data_range.api.FormatConditions.AddTop10(
                                Rank=rank, Type=True  # True for Top
                            )
                    else:
                        if is_percent:
                            # Bottom N%
                            condition_formula = data_range.api.FormatConditions.AddBottomPercentile(
                                Rank=rank, Type=False  # False for Bottom
                            )
                        else:
                            # Bottom N Items
                            condition_formula = data_range.api.FormatConditions.AddTop10(
                                Rank=rank, Type=False  # False for Bottom
                            )
                            
                elif condition_type in ["greater_than", "less_than", "equal_to"]:
                    # Parameters: {"value": 1000}
                    if not condition_parameters or "value" not in condition_parameters:
                        _close_excel_app(app, wb, save=False)
                        return {"success": False, "error": f"Missing 'value' parameter for {condition_type} condition."}
                    
                    value = condition_parameters["value"]
                    
                    # Map condition types to Excel FormatCondition operators
                    operator_map = {
                        "greater_than": 5,  # xlGreater
                        "less_than": 6,     # xlLess
                        "equal_to": 3       # xlEqual
                    }
                    
                    # Apply cell value condition
                    condition_formula = data_range.api.FormatConditions.Add(
                        Type=1,  # xlCellValue
                        Operator=operator_map[condition_type],
                        Formula1=value
                    )
                    
                elif condition_type == "between":
                    # Parameters: {"start": 1000, "end": 2000}
                    if not condition_parameters or "start" not in condition_parameters or "end" not in condition_parameters:
                        _close_excel_app(app, wb, save=False)
                        return {"success": False, "error": "Missing 'start' or 'end' parameters for between condition."}
                    
                    start = condition_parameters["start"]
                    end = condition_parameters["end"]
                    
                    # Apply between condition
                    condition_formula = data_range.api.FormatConditions.Add(
                        Type=1,  # xlCellValue
                        Operator=1,  # xlBetween
                        Formula1=start,
                        Formula2=end
                    )
                    
                elif condition_type == "contains":
                    # Parameters: {"text": "America"}
                    if not condition_parameters or "text" not in condition_parameters:
                        _close_excel_app(app, wb, save=False)
                        return {"success": False, "error": "Missing 'text' parameter for contains condition."}
                    
                    text = condition_parameters["text"]
                    
                    # Apply text contains condition
                    condition_formula = data_range.api.FormatConditions.Add(
                        Type=2,  # xlTextString
                        String=text,
                        TextOperator=8  # xlContains
                    )
                    
                elif condition_type == "date_occurring":
                    # Parameters: {"period": "today"}
                    if not condition_parameters or "period" not in condition_parameters:
                        _close_excel_app(app, wb, save=False)
                        return {"success": False, "error": "Missing 'period' parameter for date_occurring condition."}
                    
                    period = condition_parameters["period"].lower()
                    
                    # Map period to Excel TimePeriod constants
                    period_map = {
                        "today": 0,        # xlToday
                        "yesterday": 1,    # xlYesterday
                        "tomorrow": 2,     # xlTomorrow
                        "last7days": 3,    # xlLast7Days
                        "thismonth": 4,    # xlThisMonth
                        "lastmonth": 5,    # xlLastMonth
                        "nextmonth": 6,    # xlNextMonth
                        "thisweek": 7,     # xlThisWeek
                        "lastweek": 8,     # xlLastWeek
                        "nextweek": 9      # xlNextWeek
                    }
                    
                    if period not in period_map:
                        _close_excel_app(app, wb, save=False)
                        return {
                            "success": False,
                            "error": f"Invalid date period '{period}'. Supported values: {list(period_map.keys())}"
                        }
                    
                    # Apply date occurring condition
                    condition_formula = data_range.api.FormatConditions.Add(
                        Type=10,  # xlDateOccurring
                        DateOperator=period_map[period]
                    )
                    
                else:
                    _close_excel_app(app, wb, save=False)
                    return {
                        "success": False,
                        "error": f"Invalid condition_type: {condition_type}"
                    }
                
                # Apply formatting to the condition formula
                # For font formatting
                if "bold" in format_props:
                    condition_formula.Font.Bold = format_props["bold"]
                if "italic" in format_props:
                    condition_formula.Font.Italic = format_props["italic"]
                if "size" in format_props:
                    condition_formula.Font.Size = format_props["size"]
                if "color" in format_props:
                    condition_formula.Font.Color = int(f"0x{format_props['color']}", 16)
                
                # For background color
                if "fill" in format_props:
                    fill_color = format_props["fill"].start_color.rgb
                    if fill_color.startswith("FF"):
                        fill_color = fill_color[2:]
                    condition_formula.Interior.Color = int(f"0x{fill_color}", 16)
                
                # For alignment
                if "alignment" in format_props:
                    if format_props["alignment"].horizontal == "left":
                        condition_formula.HorizontalAlignment = -4131  # xlLeft
                    elif format_props["alignment"].horizontal == "center":
                        condition_formula.HorizontalAlignment = -4108  # xlCenter
                    elif format_props["alignment"].horizontal == "right":
                        condition_formula.HorizontalAlignment = -4152  # xlRight
                
                _close_excel_app(app, wb, save=True)
                return {
                    "success": True,
                    "message": f"Conditional formatting applied to data field '{field_name}' in PivotTable '{pivot_table_name}'.",
                    "scope": "data_field",
                    "condition": condition_type
                }
                
            except Exception as df_error:
                _close_excel_app(app, wb, save=False)
                return {
                    "success": False,
                    "error": f"Error applying conditional formatting to data field: {str(df_error)}"
                }
                
        elif formatting_scope == "field_items":
            # Format specific items in a field
            try:
                # Find the field
                try:
                    pivot_field = pivot_table_obj.PivotFields(field_name)
                except Exception:
                    _close_excel_app(app, wb, save=False)
                    return {
                        "success": False,
                        "error": f"Field '{field_name}' not found in PivotTable '{pivot_table_name}'."
                    }
                
                # Get all items in the field
                field_items = pivot_field.PivotItems
                
                # Filter to specific items if requested
                items_to_format = []
                if specific_items:
                    for item_name in specific_items:
                        try:
                            item = pivot_field.PivotItems(item_name)
                            items_to_format.append(item)
                        except Exception:
                            logger.warning(f"Item '{item_name}' not found in field '{field_name}'.")
                else:
                    # Format all items
                    items_to_format = [item for item in field_items]
                
                if not items_to_format:
                    _close_excel_app(app, wb, save=False)
                    return {
                        "success": False,
                        "error": f"No valid items found to format in field '{field_name}'."
                    }
                
                # Apply formatting to each item's data range
                formatted_items = []
                for item in items_to_format:
                    try:
                        # Get the Range object for this item's data cells
                        item_range = item.DataRange
                        
                        if item_range:
                            # Convert to xlwings Range for formatting
                            xl_range = sheet.range(item_range.Address)
                            
                            # Apply formatting
                            apply_formatting_to_range(xl_range)
                            formatted_items.append(item.Name)
                    except Exception as item_error:
                        logger.warning(f"Error formatting item '{item.Name}': {str(item_error)}")
                
                _close_excel_app(app, wb, save=True)
                return {
                    "success": True,
                    "message": f"Formatting applied to {len(formatted_items)} items in field '{field_name}'.",
                    "formatted_items": formatted_items
                }
                
            except Exception as fi_error:
                _close_excel_app(app, wb, save=False)
                return {
                    "success": False,
                    "error": f"Error applying formatting to field items: {str(fi_error)}"
                }
                
        elif formatting_scope in ["grand_totals", "subtotals"]:
            # Format grand totals or subtotals
            try:
                if formatting_scope == "grand_totals":
                    # Get the grand totals ranges
                    has_row_totals = pivot_table_obj.RowGrand
                    has_col_totals = pivot_table_obj.ColumnGrand
                    
                    if not (has_row_totals or has_col_totals):
                        _close_excel_app(app, wb, save=False)
                        return {
                            "success": False,
                            "error": f"Grand totals are not enabled in PivotTable '{pivot_table_name}'."
                        }
                    
                    formatted_parts = []
                    
                    if field_name.lower() in ["row", "rows"] and has_row_totals:
                        # Format row grand totals
                        try:
                            # In the PivotTable's DataBodyRange, the last row is the Grand Total row
                            data_body = pivot_table_obj.DataBodyRange
                            if data_body:
                                last_row = data_body.Rows(data_body.Rows.Count)
                                # Format the row
                                xl_range = sheet.range(last_row.Address)
                                apply_formatting_to_range(xl_range)
                                formatted_parts.append("row grand totals")
                        except Exception as row_error:
                            logger.warning(f"Error formatting row grand totals: {str(row_error)}")
                    
                    if field_name.lower() in ["column", "columns"] and has_col_totals:
                        # Format column grand totals
                        try:
                            # The last column of the DataBodyRange is the Grand Total column
                            data_body = pivot_table_obj.DataBodyRange
                            if data_body:
                                last_col = data_body.Columns(data_body.Columns.Count)
                                # Format the column
                                xl_range = sheet.range(last_col.Address)
                                apply_formatting_to_range(xl_range)
                                formatted_parts.append("column grand totals")
                        except Exception as col_error:
                            logger.warning(f"Error formatting column grand totals: {str(col_error)}")
                    
                    if not formatted_parts:
                        _close_excel_app(app, wb, save=False)
                        return {
                            "success": False,
                            "error": f"No grand totals were formatted. Ensure field_name is 'row' or 'column' and grand totals are enabled."
                        }
                    
                    _close_excel_app(app, wb, save=True)
                    return {
                        "success": True,
                        "message": f"Formatting applied to {', '.join(formatted_parts)} in PivotTable '{pivot_table_name}'.",
                        "formatted_parts": formatted_parts
                    }
                    
                else:  # subtotals
                    # Find the field
                    try:
                        pivot_field = pivot_table_obj.PivotFields(field_name)
                    except Exception:
                        _close_excel_app(app, wb, save=False)
                        return {
                            "success": False,
                            "error": f"Field '{field_name}' not found in PivotTable '{pivot_table_name}'."
                        }
                    
                    # Check if this field has subtotals enabled
                    has_subtotals = False
                    try:
                        subtotals_array = pivot_field.api.Subtotals
                        has_subtotals = any(subtotals_array)
                    except Exception:
                        pass
                    
                    if not has_subtotals:
                        _close_excel_app(app, wb, save=False)
                        return {
                            "success": False,
                            "error": f"Field '{field_name}' does not have subtotals enabled."
                        }
                    
                    # Getting the exact range of subtotal cells is complex and requires iterating 
                    # through PivotItems and checking for their LabelRange vs DataRange
                    # This is a simplified approach that may not work in all PivotTable layouts
                    formatted_count = 0
                    try:
                        for item in pivot_field.PivotItems:
                            # Try to get the data range for this item
                            data_range = item.api.DataRange
                            label_range = item.api.LabelRange
                            
                            if data_range and label_range:
                                # The subtotal cell might be at the end of the data range for this item
                                # This is a heuristic approach; exact behavior depends on the PivotTable layout
                                subtotal_row = data_range.Rows(data_range.Rows.Count)
                                if "Total" in subtotal_row.Cells(1, 1).Text:
                                    # This looks like a subtotal row, format it
                                    xl_range = sheet.range(subtotal_row.Address)
                                    apply_formatting_to_range(xl_range)
                                    formatted_count += 1
                    except Exception as sub_error:
                        logger.warning(f"Error trying to format subtotals: {str(sub_error)}")
                    
                    if formatted_count == 0:
                        _close_excel_app(app, wb, save=False)
                        return {
                            "success": False,
                            "error": f"Could not identify subtotal cells for field '{field_name}'. This may require a different approach."
                        }
                    
                    _close_excel_app(app, wb, save=True)
                    return {
                        "success": True,
                        "message": f"Applied formatting to approximately {formatted_count} subtotal cells for field '{field_name}'.",
                    }
            
            except Exception as totals_error:
                _close_excel_app(app, wb, save=False)
                return {
                    "success": False,
                    "error": f"Error formatting {formatting_scope}: {str(totals_error)}"
                }
                
        else:
            _close_excel_app(app, wb, save=False)
            return {
                "success": False,
                "error": f"Invalid formatting_scope: {formatting_scope}. Must be 'data_field', 'field_items', 'grand_totals', or 'subtotals'."
            }
            
    except Exception as e:
        logger.error(f"Failed to apply PivotTable conditional formatting: {e}")
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Failed to apply PivotTable conditional formatting: {str(e)}"}

def create_timeline_slicer(
    filepath: str,
    sheet_name: str,  # Sheet where timeline will be placed
    pivot_table_name: str,  # Name of the PivotTable to connect the timeline to
    date_field_name: str,  # Date field from PivotTable to use for timeline
    timeline_name: Optional[str] = None,  # Optional name for the timeline object
    top: Optional[float] = None,  # Position from top (in points)
    left: Optional[float] = None,  # Position from left (in points)
    width: Optional[float] = None,  # Width of the timeline (in points)
    height: Optional[float] = None,  # Height of the timeline (in points)
    time_level: str = "months"  # Default timeline level: "days", "months", "quarters", "years"
) -> Dict[str, Any]:
    """
    Creates a timeline slicer for date fields in a PivotTable, providing a specialized date-filtering interface.
    
    A timeline slicer is similar to a regular slicer but is specifically designed for date fields, offering
    intuitive filtering by days, months, quarters, and years.
    
    Args:
        filepath: Path to the Excel workbook.
        sheet_name: Sheet name where the timeline will be placed.
        pivot_table_name: Name of the PivotTable to connect the timeline to.
        date_field_name: Name of the date field from the PivotTable to use for the timeline.
        timeline_name: Optional custom name for the timeline. If None, a default name will be generated.
        top: Optional position from top of sheet in points.
        left: Optional position from left of sheet in points.
        width: Optional width of the timeline in points.
        height: Optional height of the timeline in points.
        time_level: Default time level to display ("days", "months", "quarters", "years").
        
    Returns:
        Dictionary with success status and timeline details.
    """
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings is required for timeline slicers but is not installed."}

    app, wb = _get_excel_app_and_workbook(filepath)
    if not wb:
        return {"success": False, "error": f"Failed to open workbook '{filepath}'"}
    
    try:
        # Get the worksheet
        try:
            sheet = wb.sheets[sheet_name]
        except Exception as sheet_error:
            _close_excel_app(app, wb, save=False)
            return {"success": False, "error": f"Sheet '{sheet_name}' not found: {str(sheet_error)}"}
        
        # Find the PivotTable
        try:
            pivot_table = None
            for pt in sheet.api.Parent.PivotTables():
                if pt.Name == pivot_table_name:
                    pivot_table = pt
                    break
            
            if not pivot_table:
                # Try to find the PivotTable on other sheets
                for ws in wb.sheets:
                    try:
                        for pt in ws.api.PivotTables():
                            if pt.Name == pivot_table_name:
                                pivot_table = pt
                                break
                        if pivot_table:
                            break
                    except:
                        continue
            
            if not pivot_table:
                _close_excel_app(app, wb, save=False)
                return {"success": False, "error": f"PivotTable '{pivot_table_name}' not found in workbook."}
        except Exception as pivot_error:
            _close_excel_app(app, wb, save=False)
            return {"success": False, "error": f"Error finding PivotTable: {str(pivot_error)}"}
        
        # Verify the field exists and is a date field
        try:
            # Find the field in the pivot cache
            cache_fields = pivot_table.PivotCache().PivotFields()
            field_exists = False
            is_date_field = False
            
            for i in range(1, cache_fields.Count + 1):
                field = cache_fields.Item(i)
                if field.Name == date_field_name:
                    field_exists = True
                    # Check if it's a date field by examining the data type
                    # This is a bit tricky as we need to check the actual data type
                    try:
                        # Try to get data from the first item
                        if pivot_table.PivotCache().SourceData:
                            src_range = wb.app.api.Range(pivot_table.PivotCache().SourceData)
                            for col in range(1, src_range.Columns.Count + 1):
                                if src_range.Cells(1, col).Value == date_field_name:
                                    # Check a few sample values to determine if they're dates
                                    for row in range(2, min(10, src_range.Rows.Count) + 1):
                                        cell_value = src_range.Cells(row, col).Value
                                        if cell_value:
                                            # Check if the cell is formatted as a date
                                            is_date_field = isinstance(cell_value, datetime) or \
                                                           "date" in src_range.Cells(row, col).NumberFormat.lower()
                                            break
                                    break
                    except:
                        # If we can't determine from source data, we'll proceed anyway
                        # and let Excel handle potential errors
                        pass
                    break
            
            if not field_exists:
                _close_excel_app(app, wb, save=False)
                return {"success": False, "error": f"Field '{date_field_name}' not found in PivotTable."}
            
            if not is_date_field:
                logger.warning(f"Field '{date_field_name}' may not be a date field. Timeline creation might fail.")
        except Exception as field_error:
            logger.warning(f"Error checking date field: {str(field_error)}. Proceeding anyway.")
        
        # Create the timeline
        try:
            # Default timeline name if not provided
            if not timeline_name:
                timeline_name = f"Timeline_{date_field_name}_{pivot_table_name}"
            
            # Create timeline (API differs slightly from regular slicer)
            timeline = sheet.api.Parent.TimeLines.Add(
                SlicerDestination=sheet.api,
                Source=pivot_table,
                Field=date_field_name,
                Name=timeline_name
            )
            
            # Set timeline options
            timeline.ShowHeader = True
            
            # Map time level options to Excel constants
            time_level_map = {
                "days": 0,     # xlTimelinePeriodTypeDays
                "months": 1,   # xlTimelinePeriodTypeMonths
                "quarters": 2, # xlTimelinePeriodTypeQuarters
                "years": 3     # xlTimelinePeriodTypeYears
            }
            
            # Set the timeline level if valid
            if time_level in time_level_map:
                timeline.Level = time_level_map[time_level]
            
            # Set position and size if provided
            if top is not None and left is not None:
                timeline.Top = top
                timeline.Left = left
            
            if width is not None:
                timeline.Width = width
            
            if height is not None:
                timeline.Height = height
            
            # Adjust timeline style to match Excel's default
            try:
                timeline.Style = "TimelineStyleLight1"  # Default style
            except:
                pass  # Ignore style errors
            
            _close_excel_app(app, wb, save=True)
            return {
                "success": True,
                "timeline_name": timeline_name,
                "pivot_table": pivot_table_name,
                "sheet_name": sheet_name,
                "date_field": date_field_name,
                "time_level": time_level,
                "position": {"top": timeline.Top, "left": timeline.Left},
                "size": {"width": timeline.Width, "height": timeline.Height},
                "message": f"Timeline slicer '{timeline_name}' created for field '{date_field_name}' in PivotTable '{pivot_table_name}'."
            }
            
        except Exception as timeline_error:
            _close_excel_app(app, wb, save=False)
            return {"success": False, "error": f"Failed to create timeline slicer: {str(timeline_error)}"}
            
    except Exception as e:
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Unexpected error creating timeline slicer: {str(e)}"}

def connect_slicer_to_pivot_tables(
    filepath: str,
    sheet_name: str,  # Sheet where the slicer is located
    slicer_name: str,  # Name of the slicer to connect
    pivot_table_names: List[str]  # List of PivotTable names to connect to the slicer
) -> Dict[str, Any]:
    """
    Connects a slicer to multiple PivotTables, allowing simultaneous filtering across them.
    
    This function is useful for dashboards where multiple PivotTables need to be filtered
    by the same criteria.
    
    Args:
        filepath: Path to the Excel workbook.
        sheet_name: Sheet name where the slicer is located.
        slicer_name: Name of the slicer to connect to multiple PivotTables.
        pivot_table_names: List of PivotTable names to connect to the slicer.
        
    Returns:
        Dictionary with success status and connection details.
    """
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings is required for slicer connections but is not installed."}

    app, wb = _get_excel_app_and_workbook(filepath)
    if not wb:
        return {"success": False, "error": f"Failed to open workbook '{filepath}'"}
    
    try:
        # Get the worksheet where the slicer is located
        try:
            sheet = wb.sheets[sheet_name]
        except Exception as sheet_error:
            _close_excel_app(app, wb, save=False)
            return {"success": False, "error": f"Sheet '{sheet_name}' not found: {str(sheet_error)}"}
        
        # Find the slicer
        slicer = None
        try:
            # First check in the current sheet
            for sl in sheet.api.Slicers:
                if sl.Name == slicer_name:
                    slicer = sl
                    break
            
            # If not found, search through all sheets
            if not slicer:
                for ws in wb.sheets:
                    try:
                        for sl in ws.api.Slicers:
                            if sl.Name == slicer_name:
                                slicer = sl
                                break
                        if slicer:
                            break
                    except:
                        continue
            
            if not slicer:
                # Try looking for a timeline with the same name
                for ws in wb.sheets:
                    try:
                        for tl in ws.api.Parent.TimeLines:
                            if tl.Name == slicer_name:
                                slicer = tl
                                break
                        if slicer:
                            break
                    except:
                        continue
            
            if not slicer:
                _close_excel_app(app, wb, save=False)
                return {"success": False, "error": f"Slicer or Timeline '{slicer_name}' not found in workbook."}
        except Exception as slicer_error:
            _close_excel_app(app, wb, save=False)
            return {"success": False, "error": f"Error finding slicer: {str(slicer_error)}"}
        
        # Find all requested PivotTables
        pivot_tables = {}
        missing_tables = []
        
        for pt_name in pivot_table_names:
            found = False
            # Search through all sheets
            for ws in wb.sheets:
                try:
                    for pt in ws.api.PivotTables():
                        if pt.Name == pt_name:
                            pivot_tables[pt_name] = pt
                            found = True
                            break
                    if found:
                        break
                except:
                    continue
            
            if not found:
                missing_tables.append(pt_name)
        
        if missing_tables:
            _close_excel_app(app, wb, save=False)
            return {
                "success": False, 
                "error": f"The following PivotTables were not found: {', '.join(missing_tables)}"
            }
        
        # Connect the slicer to the PivotTables
        try:
            is_timeline = hasattr(slicer, 'PivotConnection')  # TimeLines use PivotConnection
            connected_tables = []
            
            for pt_name, pt in pivot_tables.items():
                try:
                    if is_timeline:
                        # For timeline
                        slicer.PivotConnection.PivotTables.AddPivotTable(pt)
                    else:
                        # For regular slicer
                        slicer.SlicerCache.PivotTables.AddPivotTable(pt)
                    connected_tables.append(pt_name)
                except Exception as connect_error:
                    logger.warning(f"Could not connect '{pt_name}' to slicer: {str(connect_error)}")
            
            if not connected_tables:
                _close_excel_app(app, wb, save=False)
                return {"success": False, "error": "Failed to connect any PivotTables to the slicer."}
            
            _close_excel_app(app, wb, save=True)
            # Determine if it's a Timeline or regular Slicer
            slicer_type = "Timeline" if is_timeline else "Slicer" 
            return {
                "success": True,
                "slicer_name": slicer_name,
                "connected_pivot_tables": connected_tables,
                "slicer_type": slicer_type,
                "message": f"Successfully connected {len(connected_tables)} PivotTables to the {slicer_name} {slicer_type}."
            }
            
        except Exception as connection_error:
            _close_excel_app(app, wb, save=False)
            return {"success": False, "error": f"Failed to connect slicer to PivotTables: {str(connection_error)}"}
            
    except Exception as e:
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Unexpected error connecting slicer to PivotTables: {str(e)}"}

def setup_power_pivot_data_model(
    filepath: str,
    data_sources: List[Dict[str, Any]],  # List of data sources to add
    relationships: Optional[List[Dict[str, Any]]] = None  # List of relationships between tables
) -> Dict[str, Any]:
    """
    Sets up the Power Pivot data model by importing external data sources and establishing
    relationships between tables in the Excel workbook.
    
    Power Pivot is an Excel add-in that enables advanced data modeling capabilities, 
    including working with multiple data sources, defining relationships, and creating
    calculated measures/columns using DAX (Data Analysis Expressions).
    
    Args:
        filepath: Path to the Excel workbook.
        data_sources: List of data sources to add to the model. Each source is a dictionary 
                     containing properties like:
                     - source_type: 'excel', 'csv', 'database', etc.
                     - location: path or connection string
                     - target_table_name: name for the table in the model
                     - properties: source-specific properties
        relationships: Optional list of relationships to define between tables.
                      Each relationship is a dictionary containing:
                      - from_table: name of the parent table
                      - from_column: column in the parent table
                      - to_table: name of the child table
                      - to_column: column in the child table
                      - active: boolean whether this is an active relationship
        
    Returns:
        Dictionary with information about the created data model.
    """
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings is required for Power Pivot but is not installed."}

    app, wb = _get_excel_app_and_workbook(filepath)
    if not wb:
        return {"success": False, "error": f"Failed to open workbook '{filepath}'"}
    
    try:
        # Check if Power Pivot add-in is available and loaded
        addins_available = False
        try:
            # Try to access PowerPivot functionality directly
            wb.api.Model
            addins_available = True
        except:
            # If direct access fails, try to load the add-in
            try:
                # Try to load Power Pivot add-in
                app.api.COMAddIns("PowerPivotExcelClientAddIn.Connect").Connect = True
                addins_available = True
            except:
                _close_excel_app(app, wb, save=False)
                return {"success": False, "error": "Power Pivot add-in is not available or could not be loaded. Make sure it's installed and enabled."}
        
        if not addins_available:
            _close_excel_app(app, wb, save=False)
            return {"success": False, "error": "Power Pivot add-in could not be accessed after loading attempts."}
                
        # Access the data model
        model = wb.api.Model
        
        # Process each data source
        imported_tables = []
        for source in data_sources:
            try:
                source_type = source.get("source_type", "").lower()
                location = source.get("location", "")
                target_table_name = source.get("target_table_name", "")
                properties = source.get("properties", {})
                
                if not location or not target_table_name:
                    logger.warning(f"Skipping data source - missing required properties: {source}")
                    continue
                
                # Handle different source types
                if source_type == "excel":
                    # Import from Excel
                    connection_string = f"Provider=Microsoft.ACE.OLEDB.12.0;Data Source={location};Extended Properties='Excel 12.0;HDR=YES'"
                    # Get target sheet
                    sheet_name = properties.get("sheet_name", "Sheet1")
                    query = f"SELECT * FROM [{sheet_name}$]"
                    
                    # Create connection and table
                    model.AddConnection(
                        ConnectionName=f"Connection_{target_table_name}",
                        ConnectionString=connection_string,
                        Description=f"Connection to {location}"
                    )
                    model.CreateQueryTable(
                        QueryName=target_table_name,
                        ConnectionName=f"Connection_{target_table_name}",
                        QueryString=query
                    )
                    imported_tables.append({
                        "name": target_table_name,
                        "source_type": source_type,
                        "source": location,
                        "details": {"sheet": sheet_name}
                    })
                    
                elif source_type == "csv":
                    # Import from CSV
                    connection_string = f"Provider=Microsoft.ACE.OLEDB.12.0;Data Source={os.path.dirname(location)};Extended Properties='text;HDR=Yes;FMT=Delimited'"
                    file_name = os.path.basename(location)
                    query = f"SELECT * FROM [{file_name}]"
                    
                    # Create connection and table
                    model.AddConnection(
                        ConnectionName=f"Connection_{target_table_name}",
                        ConnectionString=connection_string,
                        Description=f"Connection to {location}"
                    )
                    model.CreateQueryTable(
                        QueryName=target_table_name,
                        ConnectionName=f"Connection_{target_table_name}",
                        QueryString=query
                    )
                    imported_tables.append({
                        "name": target_table_name,
                        "source_type": source_type,
                        "source": location,
                        "details": {}
                    })
                    
                elif source_type == "database":
                    # Import from database
                    conn_str = properties.get("connection_string", "")
                    if not conn_str:
                        logger.warning(f"Skipping database source - missing connection string: {source}")
                        continue
                    
                    query = properties.get("query", "")
                    if not query:
                        logger.warning(f"Skipping database source - missing query: {source}")
                        continue
                    
                    # Create connection and table
                    model.AddConnection(
                        ConnectionName=f"Connection_{target_table_name}",
                        ConnectionString=conn_str,
                        Description=f"Database connection for {target_table_name}"
                    )
                    model.CreateQueryTable(
                        QueryName=target_table_name,
                        ConnectionName=f"Connection_{target_table_name}",
                        QueryString=query
                    )
                    imported_tables.append({
                        "name": target_table_name,
                        "source_type": source_type,
                        "source": "database",
                        "details": {"connection_type": properties.get("connection_type", "Unknown")}
                    })
                    
                elif source_type == "worksheet_table":
                    # Use an existing worksheet table
                    sheet_name = properties.get("sheet_name", "")
                    table_name = properties.get("table_name", "")
                    
                    if not sheet_name or not table_name:
                        logger.warning(f"Skipping worksheet table source - missing sheet_name or table_name: {source}")
                        continue
                    
                    # Find the table in the workbook
                    sheet = wb.sheets[sheet_name]
                    for tbl in sheet.api.ListObjects:
                        if tbl.Name == table_name:
                            # Add table to the data model
                            tbl.TableObject.AddToModel(target_table_name)
                            imported_tables.append({
                                "name": target_table_name,
                                "source_type": source_type,
                                "source": f"{sheet_name}.{table_name}",
                                "details": {"sheet": sheet_name, "table": table_name}
                            })
                            break
                    else:
                        logger.warning(f"Table {table_name} not found in sheet {sheet_name}")
                        
                else:
                    logger.warning(f"Unsupported source type: {source_type}")
                    
            except Exception as src_error:
                logger.error(f"Error processing data source {source.get('target_table_name')}: {str(src_error)}")
        
        # Create relationships if provided
        created_relationships = []
        if relationships:
            for rel in relationships:
                try:
                    from_table = rel.get("from_table")
                    from_column = rel.get("from_column")
                    to_table = rel.get("to_table")
                    to_column = rel.get("to_column")
                    active = rel.get("active", True)
                    
                    if not all([from_table, from_column, to_table, to_column]):
                        logger.warning(f"Skipping relationship - missing required properties: {rel}")
                        continue
                    
                    # Create the relationship
                    new_rel = model.CreateRelationship(
                        ForeignKeyTable=to_table,
                        ForeignKeyColumn=to_column,
                        PrimaryKeyTable=from_table,
                        PrimaryKeyColumn=from_column,
                        Active=active
                    )
                    
                    created_relationships.append({
                        "from_table": from_table,
                        "from_column": from_column,
                        "to_table": to_table,
                        "to_column": to_column,
                        "active": active
                    })
                    
                except Exception as rel_error:
                    logger.error(f"Error creating relationship: {str(rel_error)}")
        
        # Save and close
        _close_excel_app(app, wb, save=True)
        
        return {
            "success": True,
            "imported_tables": imported_tables,
            "created_relationships": created_relationships,
            "message": f"Successfully set up Power Pivot data model with {len(imported_tables)} tables and {len(created_relationships)} relationships."
        }
        
    except Exception as e:
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Error setting up Power Pivot data model: {str(e)}"}

def create_power_pivot_measure(
    filepath: str,
    measure_name: str,
    dax_formula: str,  # DAX formula for the measure
    table_name: str,  # Table where the measure will be displayed
    display_folder: Optional[str] = None,  # Display folder for organizing measures (optional)
    format_string: Optional[str] = None  # Format string for the measure (optional)
) -> Dict[str, Any]:
    """
    Creates a new calculated measure in the Power Pivot data model using DAX (Data Analysis Expressions).
    
    Measures are used in PivotTables to perform calculations on the data model, providing advanced
    aggregation capabilities beyond what regular Excel formulas can do.
    
    Args:
        filepath: Path to the Excel workbook.
        measure_name: Name for the new measure.
        dax_formula: DAX formula that defines the measure's calculation.
        table_name: Name of the table where the measure will be displayed.
        display_folder: Optional folder path to organize measures in the field list.
        format_string: Optional format string for the measure (e.g., "$#,##0.00" for currency).
        
    Returns:
        Dictionary with information about the created measure.
    """
    if not _check_dependencies_for_pivot_tables():
        return {"success": False, "error": "xlwings is required for Power Pivot measures but is not installed."}

    app, wb = _get_excel_app_and_workbook(filepath)
    if not wb:
        return {"success": False, "error": f"Failed to open workbook '{filepath}'"}
    
    try:
        # Check if Power Pivot add-in is available and enabled
        try:
            # Try to access PowerPivot functionality directly
            model = wb.api.Model
        except:
            try:
                # Try to load Power Pivot add-in
                app.api.COMAddIns("PowerPivotExcelClientAddIn.Connect").Connect = True
                model = wb.api.Model
            except:
                _close_excel_app(app, wb, save=False)
                return {"success": False, "error": "Power Pivot add-in is not available or could not be loaded."}
        
        # Verify the table exists in the model
        table_exists = False
        for table in model.Tables:
            if table.Name == table_name:
                table_exists = True
                break
        
        if not table_exists:
            _close_excel_app(app, wb, save=False)
            return {"success": False, "error": f"Table '{table_name}' not found in the data model."}
        
        # Create the measure
        try:
            mdx = "CREATEKPI("
            mdx += f"'{table_name}'[{measure_name}], "
            mdx += f"'{dax_formula}', "
            
            # Add optional properties if provided
            if display_folder:
                mdx += f"'{display_folder}', "
            else:
                mdx += "'', "
                
            if format_string:
                mdx += f"'{format_string}'"
            else:
                mdx += "''"
                
            mdx += ")"
            
            # Execute the MDX command to create the measure
            model.ExecuteCommand(mdx)
            
            _close_excel_app(app, wb, save=True)
            
            return {
                "success": True,
                "measure_name": measure_name,
                "table": table_name,
                "dax_formula": dax_formula,
                "display_folder": display_folder,
                "format_string": format_string,
                "message": f"Successfully created Power Pivot measure '{measure_name}' in table '{table_name}'."
            }
            
        except Exception as measure_error:
            _close_excel_app(app, wb, save=False)
            return {"success": False, "error": f"Failed to create measure: {str(measure_error)}"}
            
    except Exception as e:
        _close_excel_app(app, wb, save=False)
        return {"success": False, "error": f"Unexpected error creating Power Pivot measure: {str(e)}"}