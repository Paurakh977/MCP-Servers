"""Excel workbook and worksheet manipulation tools for the MCP server.

This module provides functions for creating, reading, and manipulating Excel workbooks,
worksheets, ranges, and formatting. It builds upon openpyxl for Excel operations.
"""

import os
import logging
import json
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional, Union, Tuple
import re
import time

from openpyxl import Workbook, load_workbook
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Border, PatternFill, Side, Alignment
from openpyxl.styles.colors import Color
from openpyxl.worksheet.datavalidation import DataValidation

# Set up logging
logger = logging.getLogger(__name__)

# Constants for formula handling
ARRAY_FORMULA_TYPES = ["XLOOKUP", "FILTER", "UNIQUE", "SORT", "SORTBY", "SEQUENCE", "TRANSPOSE"]
ARRAY_REFERENCE_PATTERN = r'([A-Z]+[0-9]+)#'  # Matches references like A2# used in spilled array references
EXTERNAL_REF_PATTERN = r'\[([^\]]+)\]'

# Exceptions
class ExcelError(Exception):
    """Base exception for Excel operations"""
    pass

class WorkbookError(ExcelError):
    """Error related to workbook operations"""
    pass

class SheetError(ExcelError):
    """Error related to sheet operations"""
    pass

class RangeError(ExcelError):
    """Error related to range operations"""
    pass

class ValidationError(ExcelError):
    """Error related to validation"""
    pass

# Add after imports section (around line 20)
try:
    from formulas import Parser
    from formulas.errors import FormulaError
    has_formula_engine = True
except ImportError:
    has_formula_engine = False
    print("Formulas library not installed. For formula evaluation: pip install formulas")

# Helper Functions
def parse_cell_reference(cell_ref: str) -> Tuple[int, int]:
    """Parse a cell reference (e.g. 'A1') into row and column indices"""
    if not cell_ref or not isinstance(cell_ref, str):
        raise ValidationError(f"Invalid cell reference: {cell_ref}")
        
    # Extract column letters and row numbers
    col_str = ''.join(c for c in cell_ref if c.isalpha())
    row_str = ''.join(c for c in cell_ref if c.isdigit())
    
    if not col_str or not row_str:
        raise ValidationError(f"Invalid cell reference format: {cell_ref}")
        
    try:
        col_idx = column_index_from_string(col_str)
        row_idx = int(row_str)
        return row_idx, col_idx
    except (ValueError, KeyError) as e:
        raise ValidationError(f"Invalid cell reference: {cell_ref} - {str(e)}")

def parse_cell_range(start_cell: str, end_cell: Optional[str] = None) -> Tuple[int, int, Optional[int], Optional[int]]:
    """Parse a cell range into row and column indices"""
    start_row, start_col = parse_cell_reference(start_cell)
    
    if end_cell:
        end_row, end_col = parse_cell_reference(end_cell)
        return start_row, start_col, end_row, end_col
    else:
        return start_row, start_col, None, None

def format_range_string(start_row: int, start_col: int, end_row: int, end_col: int) -> str:
    """Format a range string from row and column indices"""
    return f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

def find_file_in_allowed_dirs(filename: str, allowed_dirs: List[str]) -> Optional[str]:
    """Find a file in the allowed directories"""
    for base_dir in allowed_dirs:
        for root, _, files in os.walk(base_dir):
            matching_files = [f for f in files if filename.lower() in f.lower()]
            if matching_files:
                return os.path.join(root, matching_files[0])
    return None

# Workbook Operations
def create_excel_workbook(filepath: str, sheet_name: str = "Sheet1") -> Dict[str, Any]:
    """Create a new Excel workbook with optional custom sheet name"""
    try:
        wb = Workbook()
        # Rename default sheet
        if "Sheet" in wb.sheetnames:
            sheet = wb["Sheet"]
            sheet.title = sheet_name
        else:
            wb.create_sheet(sheet_name)

        path = Path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))
        return {
            "success": True,
            "message": f"Created workbook: {filepath}",
            "active_sheet": sheet_name,
        }
    except Exception as e:
        logger.error(f"Failed to create workbook: {e}")
        return {
            "success": False,
            "error": f"Failed to create workbook: {str(e)}"
        }

def get_workbook_metadata(filepath: str, include_ranges: bool = False) -> Dict[str, Any]:
    """Get detailed metadata about an Excel workbook"""
    try:
        path = Path(filepath)
        if not path.exists():
            return {
                "success": False,
                "error": f"File not found: {filepath}"
            }
            
        wb = load_workbook(filepath, read_only=True)
        
        sheets_info = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_info = {
                "name": sheet_name,
                "rows": ws.max_row,
                "columns": ws.max_column,
            }
            
            if include_ranges and ws.max_row > 0 and ws.max_column > 0:
                sheet_info["used_range"] = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
                
                # Try to get column headers
                try:
                    headers = []
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=1, column=col)
                        header = cell.value if cell.value is not None else f"Column {get_column_letter(col)}"
                        headers.append(str(header))
                    sheet_info["headers"] = headers
                except:
                    pass  # Skip headers if there's an error
            
            sheets_info.append(sheet_info)
        
        info = {
            "success": True,
            "filename": path.name,
            "file_size": path.stat().st_size,
            "file_size_human": f"{path.stat().st_size / 1024:.2f} KB",
            "modified": path.stat().st_mtime,
            "modified_date": datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
            "sheets_count": len(wb.sheetnames),
            "sheets": sheets_info
        }
        
        wb.close()
        return info
        
    except Exception as e:
        logger.error(f"Failed to get workbook metadata: {e}")
        return {
            "success": False,
            "error": f"Failed to get workbook metadata: {str(e)}"
        }

def delete_excel_workbook(filepath: str) -> Dict[str, Any]:
    """Delete an Excel workbook file.
    
    Args:
        filepath: Path to the Excel workbook to delete
        
    Returns:
        Dictionary with success status and message
    """
    try:
        path = Path(filepath)
        
        # Check if file exists
        if not path.exists():
            return {
                "success": False,
                "error": f"File not found: {filepath}"
            }
            
        # Check if it's an Excel file
        file_extension = path.suffix.lower()
        if file_extension != ".xlsx":
            return {
                "success": False, 
                "error": f"Not an Excel file: {filepath}"
            }
            
        # Delete the workbook file
        path.unlink()
        
        return {
            "success": True,
            "message": f"Workbook {path.name} deleted successfully"
        }
    except Exception as e:
        logger.error(f"Failed to delete workbook: {e}")
        return {
            "success": False,
            "error": f"Failed to delete workbook: {str(e)}"
        }

# Worksheet Operations
def create_worksheet(filepath: str, sheet_name: str) -> Dict[str, Any]:
    """Create a new worksheet in an Excel workbook"""
    try:
        wb = load_workbook(filepath)

        # Check if sheet already exists
        if sheet_name in wb.sheetnames:
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' already exists"
            }

        # Create new sheet
        wb.create_sheet(sheet_name)
        wb.save(filepath)
        wb.close()
        return {
            "success": True,
            "message": f"Sheet '{sheet_name}' created successfully"
        }
    except Exception as e:
        logger.error(f"Failed to create worksheet: {e}")
        return {
            "success": False,
            "error": f"Failed to create worksheet: {str(e)}"
        }

def copy_worksheet(filepath: str, source_sheet: str, target_sheet: str) -> Dict[str, Any]:
    """Copy a worksheet within the same workbook"""
    try:
        wb = load_workbook(filepath)
        
        if source_sheet not in wb.sheetnames:
            return {
                "success": False,
                "error": f"Source sheet '{source_sheet}' not found"
            }
            
        if target_sheet in wb.sheetnames:
            return {
                "success": False,
                "error": f"Target sheet '{target_sheet}' already exists"
            }
            
        source = wb[source_sheet]
        target = wb.copy_worksheet(source)
        target.title = target_sheet
        
        wb.save(filepath)
        wb.close()
        return {
            "success": True,
            "message": f"Sheet '{source_sheet}' copied to '{target_sheet}'"
        }
    except Exception as e:
        logger.error(f"Failed to copy worksheet: {e}")
        return {
            "success": False,
            "error": f"Failed to copy worksheet: {str(e)}"
        }

def delete_worksheet(filepath: str, sheet_name: str) -> Dict[str, Any]:
    """Delete a worksheet from an Excel workbook"""
    try:
        wb = load_workbook(filepath)
        
        if sheet_name not in wb.sheetnames:
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found"
            }
            
        if len(wb.sheetnames) == 1:
            return {
                "success": False,
                "error": "Cannot delete the only sheet in workbook"
            }
            
        del wb[sheet_name]
        wb.save(filepath)
        wb.close()
        return {
            "success": True,
            "message": f"Sheet '{sheet_name}' deleted successfully"
        }
    except Exception as e:
        logger.error(f"Failed to delete worksheet: {e}")
        return {
            "success": False,
            "error": f"Failed to delete worksheet: {str(e)}"
        }

def rename_worksheet(filepath: str, old_name: str, new_name: str) -> Dict[str, Any]:
    """Rename a worksheet in an Excel workbook"""
    try:
        wb = load_workbook(filepath)
        
        if old_name not in wb.sheetnames:
            return {
                "success": False,
                "error": f"Sheet '{old_name}' not found"
            }
            
        if new_name in wb.sheetnames:
            return {
                "success": False,
                "error": f"Sheet '{new_name}' already exists"
            }
            
        sheet = wb[old_name]
        sheet.title = new_name
        wb.save(filepath)
        wb.close()
        return {
            "success": True,
            "message": f"Sheet renamed from '{old_name}' to '{new_name}'"
        }
    except Exception as e:
        logger.error(f"Failed to rename worksheet: {e}")
        return {
            "success": False,
            "error": f"Failed to rename worksheet: {str(e)}"
        }

# Range Operations
def copy_excel_range(
    filepath: str, 
    sheet_name: str, 
    source_range: str, 
    target_start: str, 
    target_sheet: Optional[str] = None
) -> Dict[str, Any]:
    """Copy a range of cells to another location in Excel"""
    try:
        wb = load_workbook(filepath)
        
        # Validate source sheet
        if sheet_name not in wb.sheetnames:
            return {
                "success": False,
                "error": f"Source sheet '{sheet_name}' not found"
            }
        
        source_ws = wb[sheet_name]
        
        # Validate target sheet
        if target_sheet and target_sheet not in wb.sheetnames:
            return {
                "success": False,
                "error": f"Target sheet '{target_sheet}' not found"
            }
        
        target_ws = wb[target_sheet] if target_sheet else source_ws
        
        # Parse source range
        try:
            if ':' in source_range:
                source_start, source_end = source_range.split(':')
                src_start_row, src_start_col, src_end_row, src_end_col = parse_cell_range(
                    source_start, source_end
                )
            else:
                src_start_row, src_start_col, src_end_row, src_end_col = parse_cell_range(
                    source_range
                )
                src_end_row = src_start_row
                src_end_col = src_start_col
        except ValidationError as e:
            return {
                "success": False,
                "error": f"Invalid source range: {str(e)}"
            }
            
        # Parse target start
        try:
            tgt_start_row, tgt_start_col, _, _ = parse_cell_range(target_start)
        except ValidationError as e:
            return {
                "success": False,
                "error": f"Invalid target cell: {str(e)}"
            }
            
        # Copy range
        for i, row in enumerate(range(src_start_row, src_end_row + 1)):
            for j, col in enumerate(range(src_start_col, src_end_col + 1)):
                try:
                    source_cell = source_ws.cell(row=row, column=col)
                    target_cell = target_ws.cell(row=tgt_start_row + i, column=tgt_start_col + j)
                    
                    # Copy value
                    target_cell.value = source_cell.value
                    
                    # Copy style if possible
                    if hasattr(source_cell, '_style') and source_cell._style:
                        target_cell._style = source_cell._style.copy()
                except Exception as e:
                    logger.warning(f"Could not copy cell at row {row}, column {col}: {e}")
                    continue
                    
        # Save workbook
        wb.save(filepath)
        wb.close()
        
        # Prepare success message
        source_range_str = source_range if ':' in source_range else source_range
        target_info = f" to {target_sheet}" if target_sheet else ""
        
        return {
            "success": True,
            "message": f"Range {source_range_str} copied from {sheet_name} to {target_start}{target_info}"
        }
    except Exception as e:
        logger.error(f"Failed to copy range: {e}")
        return {
            "success": False,
            "error": f"Failed to copy range: {str(e)}"
        }

def delete_excel_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    shift_direction: str = "up"
) -> Dict[str, Any]:
    """Delete a range of cells and shift remaining cells in Excel"""
    try:
        wb = load_workbook(filepath)
        
        # Validate sheet
        if sheet_name not in wb.sheetnames:
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found"
            }
            
        worksheet = wb[sheet_name]
        
        # Validate range
        try:
            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
            
            if end_row is None:
                end_row = start_row
                end_col = start_col
                
            if end_row > worksheet.max_row:
                return {
                    "success": False,
                    "error": f"End row {end_row} out of bounds (1-{worksheet.max_row})"
                }
                
            if end_col > worksheet.max_column:
                return {
                    "success": False,
                    "error": f"End column {end_col} out of bounds (1-{worksheet.max_column})"
                }
        except ValidationError as e:
            return {
                "success": False,
                "error": f"Invalid range: {str(e)}"
            }
            
        # Validate shift direction
        if shift_direction not in ["up", "left"]:
            return {
                "success": False,
                "error": f"Invalid shift direction: {shift_direction}. Must be 'up' or 'left'"
            }
            
        # Format range string for the message
        range_string = format_range_string(start_row, start_col, end_row, end_col)
        
        # Clear range contents first
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.value = None
                
        # Shift cells if requested
        if shift_direction == "up":
            # Calculate how many rows to delete
            num_rows = end_row - start_row + 1
            worksheet.delete_rows(start_row, num_rows)
        elif shift_direction == "left":
            # Calculate how many columns to delete
            num_cols = end_col - start_col + 1
            worksheet.delete_cols(start_col, num_cols)
            
        # Save workbook
        wb.save(filepath)
        wb.close()
        
        action = "deleted and cells shifted" if shift_direction else "cleared"
        return {
            "success": True,
            "message": f"Range {range_string} {action} in sheet '{sheet_name}'"
        }
    except Exception as e:
        logger.error(f"Failed to delete range: {e}")
        return {
            "success": False,
            "error": f"Failed to delete range: {str(e)}"
        }

def merge_excel_cells(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str
) -> Dict[str, Any]:
    """Merge a range of cells in Excel"""
    try:
        wb = load_workbook(filepath)
        
        # Validate sheet
        if sheet_name not in wb.sheetnames:
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found"
            }
            
        worksheet = wb[sheet_name]
        
        # Validate range
        try:
            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
            
            if end_row is None or end_col is None:
                return {
                    "success": False,
                    "error": "Both start and end cells must be specified for merging"
                }
                
            if end_row > worksheet.max_row:
                return {
                    "success": False,
                    "error": f"End row {end_row} out of bounds (1-{worksheet.max_row})"
                }
                
            if end_col > worksheet.max_column:
                return {
                    "success": False,
                    "error": f"End column {end_col} out of bounds (1-{worksheet.max_column})"
                }
                
        except ValidationError as e:
            return {
                "success": False,
                "error": f"Invalid range: {str(e)}"
            }
            
        # Format range string
        range_string = format_range_string(start_row, start_col, end_row, end_col)
        
        # Check if already merged
        for merged_range in worksheet.merged_cells.ranges:
            if str(merged_range) == range_string:
                return {
                    "success": False,
                    "error": f"Range {range_string} is already merged"
                }
            
        # Merge cells
        worksheet.merge_cells(range_string)
        
        # Save workbook
        wb.save(filepath)
        wb.close()
        
        return {
            "success": True,
            "message": f"Range {range_string} merged in sheet '{sheet_name}'"
        }
    except Exception as e:
        logger.error(f"Failed to merge cells: {e}")
        return {
            "success": False,
            "error": f"Failed to merge cells: {str(e)}"
        }

def unmerge_excel_cells(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str
) -> Dict[str, Any]:
    """Unmerge a previously merged range of cells in Excel"""
    try:
        wb = load_workbook(filepath)
        
        # Validate sheet
        if sheet_name not in wb.sheetnames:
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found"
            }
            
        worksheet = wb[sheet_name]
        
        # Validate range
        try:
            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
            
            if end_row is None or end_col is None:
                return {
                    "success": False,
                    "error": "Both start and end cells must be specified for unmerging"
                }
                
        except ValidationError as e:
            return {
                "success": False,
                "error": f"Invalid range: {str(e)}"
            }
            
        # Format range string
        range_string = format_range_string(start_row, start_col, end_row, end_col)
        
        # Check if the range is actually merged
        merged_ranges = worksheet.merged_cells.ranges
        target_range = range_string.upper()
        
        if not any(str(merged_range).upper() == target_range for merged_range in merged_ranges):
            return {
                "success": False,
                "error": f"Range {range_string} is not merged"
            }
            
        # Unmerge cells
        worksheet.unmerge_cells(range_string)
        
        # Save workbook
        wb.save(filepath)
        wb.close()
        
        return {
            "success": True,
            "message": f"Range {range_string} unmerged in sheet '{sheet_name}'"
        }
    except Exception as e:
        logger.error(f"Failed to unmerge cells: {e}")
        return {
            "success": False,
            "error": f"Failed to unmerge cells: {str(e)}"
        }

# Data Operations
def write_excel_data(
    filepath: str,
    sheet_name: str,
    data: List[List[Any]],
    start_cell: str = "A1",
    headers: bool = True,
    auto_adjust_width: bool = False
) -> Dict[str, Any]:
    """Write data to an Excel worksheet"""
    try:
        # Validate input data
        if not data or not isinstance(data, list):
            return {
                "success": False,
                "error": "Data must be a non-empty list"
            }
            
        # Load workbook
        wb = load_workbook(filepath)
        
        # Validate sheet
        if sheet_name not in wb.sheetnames:
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found"
            }
            
        worksheet = wb[sheet_name]
        
        # Parse start cell
        try:
            start_row, start_col, _, _ = parse_cell_range(start_cell)
        except ValidationError as e:
            return {
                "success": False,
                "error": f"Invalid start cell: {str(e)}"
            }
            
        # Write data
        rows_written = 0
        max_col_width = {}  # Track max width for each column
        
        for row_idx, row_data in enumerate(data):
            if not isinstance(row_data, list):
                # Try to convert to list if possible (e.g., for dictionaries)
                if hasattr(row_data, 'values'):
                    row_data = list(row_data.values())
                else:
                    row_data = [row_data]  # Single value
                    
            for col_idx, value in enumerate(row_data):
                cell = worksheet.cell(
                    row=start_row + row_idx,
                    column=start_col + col_idx
                )
                cell.value = value
                
                # Add bold formatting for headers
                if headers and row_idx == 0:
                    cell.font = Font(bold=True)
                
                # Track max column width if auto-adjusting
                if auto_adjust_width:
                    # Calculate the display length of the value
                    str_value = str(value) if value is not None else ""
                    display_length = len(str_value)
                    
                    # Get column letter for this column
                    col_letter = get_column_letter(start_col + col_idx)
                    
                    # Update max width if needed
                    if col_letter not in max_col_width or display_length > max_col_width[col_letter]:
                        max_col_width[col_letter] = display_length
                    
            rows_written += 1
        
        # Adjust column widths if requested
        if auto_adjust_width and max_col_width:
            for col_letter, width in max_col_width.items():
                # Add padding and set a minimum width
                adjusted_width = min(max(width + 2, 8), 75)  # Min 8, max 75 characters
                worksheet.column_dimensions[col_letter].width = adjusted_width
        
        # Save workbook
        wb.save(filepath)
        wb.close()
        
        return {
            "success": True,
            "message": f"Data written to sheet '{sheet_name}' starting at {start_cell}",
            "rows_written": rows_written,
            "columns_written": len(data[0]) if data and data[0] else 0
        }
    except Exception as e:
        logger.error(f"Failed to write data: {e}")
        return {
            "success": False,
            "error": f"Failed to write data: {str(e)}"
        }

def format_excel_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    bold: bool = False,
    italic: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False,
    border_style: Optional[str] = None,
    auto_adjust_width: bool = False
) -> Dict[str, Any]:
    """Apply formatting to a range of cells in Excel"""
    try:
        # Load workbook
        wb = load_workbook(filepath)
        
        # Validate sheet
        if sheet_name not in wb.sheetnames:
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found"
            }
            
        worksheet = wb[sheet_name]
        
        # Parse cell range
        try:
            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
            
            if end_row is None:
                end_row = start_row
                end_col = start_col
                
        except ValidationError as e:
            return {
                "success": False,
                "error": f"Invalid range: {str(e)}"
            }
            
        # Set up font
        font_args = {}
        if bold:
            font_args["bold"] = True
        if italic:
            font_args["italic"] = True
        if font_size:
            font_args["size"] = font_size
        if font_color:
            # Handle color format (with or without #)
            if font_color.startswith("#"):
                font_color = font_color[1:]
            font_args["color"] = font_color
            
        # Set up fill
        fill = None
        if bg_color:
            # Handle color format (with or without #)
            if bg_color.startswith("#"):
                bg_color = bg_color[1:]
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            
        # Set up alignment
        align = None
        if alignment or wrap_text:
            align_horz = None
            if alignment:
                align_horz = {
                    "left": "left",
                    "center": "center",
                    "right": "right"
                }.get(alignment.lower())
                
            align = Alignment(horizontal=align_horz, wrap_text=wrap_text)
            
        # Set up border
        border = None
        if border_style:
            border_styles = {
                "thin": Side(style="thin"),
                "medium": Side(style="medium"),
                "thick": Side(style="thick"),
                "dashed": Side(style="dashed"),
                "dotted": Side(style="dotted"),
                "double": Side(style="double")
            }
            if border_style.lower() in border_styles:
                side = border_styles[border_style.lower()]
                border = Border(left=side, right=side, top=side, bottom=side)
        
        # Track max column widths if auto-adjusting
        max_col_width = {} if auto_adjust_width else None
        
        # Apply formatting to range
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                
                # Apply font if any font properties set
                if font_args:
                    # Start with existing font and update properties
                    new_font = Font(
                        name=cell.font.name,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        size=cell.font.size,
                        color=cell.font.color
                    )
                    # Update with specified properties
                    for key, value in font_args.items():
                        setattr(new_font, key, value)
                    cell.font = new_font
                
                # Apply fill if specified
                if fill:
                    cell.fill = fill
                    
                # Apply alignment if specified
                if align:
                    cell.alignment = align
                    
                # Apply border if specified
                if border:
                    cell.border = border
                
                # Track max column width if auto-adjusting
                if auto_adjust_width and cell.value is not None:
                    # Calculate display length
                    str_value = str(cell.value)
                    display_length = len(str_value)
                    
                    # Get column letter
                    col_letter = get_column_letter(col)
                    
                    # Update max width if needed
                    if col_letter not in max_col_width or display_length > max_col_width[col_letter]:
                        max_col_width[col_letter] = display_length
        
        # Adjust column widths if requested
        if auto_adjust_width and max_col_width:
            for col_letter, width in max_col_width.items():
                # Add padding based on content (more padding for wider content)
                padding = 2 if width < 20 else 4
                adjusted_width = min(width + padding, 75)  # Max 75 characters
                worksheet.column_dimensions[col_letter].width = adjusted_width
        
        # Save workbook
        wb.save(filepath)
        wb.close()
        
        # Format range string for the message
        range_string = format_range_string(start_row, start_col, end_row, end_col)
        
        return {
            "success": True,
            "message": f"Formatting applied to range {range_string} in sheet '{sheet_name}'"
        }
    except Exception as e:
        logger.error(f"Failed to apply formatting: {e}")
        return {
            "success": False,
            "error": f"Failed to apply formatting: {str(e)}"
        }

def adjust_column_widths(
    filepath: str,
    sheet_name: str,
    column_range: Optional[str] = None,
    auto_fit: bool = True,
    custom_widths: Optional[Dict[str, int]] = None
) -> Dict[str, Any]:
    """Adjust column widths in an Excel worksheet
    
    Args:
        filepath: Path to the Excel workbook
        sheet_name: Name of the worksheet to modify
        column_range: Optional range of columns to adjust (e.g. 'A:D')
        auto_fit: Whether to automatically fit column widths to content
        custom_widths: Optional dictionary mapping column letters to widths
    
    Returns:
        Dictionary with operation result
    """
    try:
        # Load workbook (can't use read_only mode when modifying)
        wb = load_workbook(filepath)
        
        # Validate sheet
        if sheet_name not in wb.sheetnames:
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found"
            }
            
        worksheet = wb[sheet_name]
        
        # Parse column range if provided
        start_col, end_col = 1, worksheet.max_column
        if column_range:
            try:
                if ':' in column_range:
                    start_col_str, end_col_str = column_range.split(':')
                    start_col = column_index_from_string(start_col_str)
                    end_col = column_index_from_string(end_col_str)
                else:
                    # Single column
                    start_col = end_col = column_index_from_string(column_range)
            except Exception as e:
                return {
                    "success": False,
                    "error": f"Invalid column range format: {str(e)}"
                }
        
        adjusted_columns = []
        
        # Apply custom widths if provided
        if custom_widths:
            for col_letter, width in custom_widths.items():
                try:
                    worksheet.column_dimensions[col_letter].width = width
                    adjusted_columns.append(col_letter)
                except Exception as e:
                    logger.warning(f"Failed to adjust column {col_letter}: {e}")
        
        # Auto-fit columns based on content
        if auto_fit:
            # For each column in range, find max content width
            for col_idx in range(start_col, end_col + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                
                # Check each cell in the column
                for row_idx in range(1, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        # Get string representation and its length
                        cell_text = str(cell.value)
                        
                        # Adjust for bold text (takes more space)
                        if hasattr(cell, 'font') and cell.font and cell.font.bold:
                            length = len(cell_text) * 1.2  # Add 20% for bold
                        else:
                            length = len(cell_text)
                            
                        max_length = max(max_length, length)
                
                # Set column width (with some padding) if content exists
                if max_length > 0:
                    # Add padding based on content length
                    padding = 2 if max_length < 20 else 4
                    
                    # Set width (with minimum and maximum limits)
                    width = min(max(max_length + padding, 8), 75)  # Min 8, max 75 characters
                    worksheet.column_dimensions[col_letter].width = width
                    adjusted_columns.append(col_letter)
        
        # Save workbook
        wb.save(filepath)
        wb.close()
        
        return {
            "success": True,
            "message": f"Adjusted {len(adjusted_columns)} column widths in sheet '{sheet_name}'",
            "adjusted_columns": adjusted_columns
        }
    except Exception as e:
        logger.error(f"Failed to adjust column widths: {e}")
        return {
            "success": False,
            "error": f"Failed to adjust column widths: {str(e)}"
        }

def validate_excel_formula(formula: str) -> Dict[str, Any]:
    """Validate Excel formula syntax and check for common errors.
    
    Args:
        formula: Excel formula to validate
        
    Returns:
        Dictionary with validation result and potential issues
    """
    try:
        # Ensure formula starts with '='
        formula_str = formula.strip()
        if not formula_str.startswith("="):
            formula_str = "=" + formula_str
        
        # Check for division by zero risk
        has_division = "/" in formula_str
        has_division_safety = False
        
        if has_division:
            # Check if formula contains IFERROR, IF to handle division
            division_safety_patterns = ["IFERROR", "IF(", "IFNA"]
            for pattern in division_safety_patterns:
                if pattern in formula_str.upper():
                    has_division_safety = True
                    break
        
        # Check for common syntax errors
        syntax_issues = []
        
        # Check for unbalanced parentheses
        open_count = formula_str.count("(") 
        close_count = formula_str.count(")")
        if open_count != close_count:
            syntax_issues.append(f"Unbalanced parentheses: {open_count} opening vs {close_count} closing")
        
        # Check for incomplete functions (function names followed directly by operators)
        incomplete_function = re.search(r'([A-Z]+)([+\-*/])', formula_str.upper())
        if incomplete_function:
            syntax_issues.append(f"Possible incomplete function: {incomplete_function.group(1)}")
        
        # Expanded array formula detection with better pattern matching
        is_array_formula = False
        array_formula_type = None
        spill_direction = None
        
        # Enhanced detection of array formulas
        for array_type in ARRAY_FORMULA_TYPES:
            # Improved regex to better detect the actual function usage
            # This pattern matches function name followed by opening parenthesis,
            # ensuring we're matching actual function calls, not just occurrences of the name
            pattern = rf'\b{array_type}\s*\('
            if re.search(pattern, formula_str.upper()):
                is_array_formula = True
                array_formula_type = array_type
                
                # Determine likely spill direction
                if array_type in ["UNIQUE", "FILTER", "SORT", "SORTBY"]:
                    spill_direction = "vertical"  # These typically spill downward
                elif array_type == "TRANSPOSE":
                    spill_direction = "horizontal"  # TRANSPOSE spills horizontally
                else:
                    spill_direction = "undetermined"  # Default
                    
                break
                
        # Enhanced detection of array references (like A2#)
        array_references = []
        array_ref_matches = re.findall(ARRAY_REFERENCE_PATTERN, formula_str)
        for match in array_ref_matches:
            array_references.append(match)
                
        # Check for external references
        external_refs = []
        ext_matches = re.findall(EXTERNAL_REF_PATTERN, formula_str)
        for match in ext_matches:
            if match.lower().endswith('.xlsx'):
                external_refs.append(match)
        
        result = {
            "is_valid": len(syntax_issues) == 0,
            "formula": formula_str,
            "issues": syntax_issues,
            "has_division": has_division,
            "has_division_safety": has_division_safety,
            "needs_safety_wrapper": has_division and not has_division_safety,
            "is_array_formula": is_array_formula,
            "array_formula_type": array_formula_type,
            "spill_direction": spill_direction,
            "array_references": array_references,
            "external_references": external_refs
        }
        
        return result
    except Exception as e:
        return {
            "is_valid": False,
            "formula": formula,
            "issues": [str(e)]
        }

def apply_excel_formula(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
    protect_from_errors: bool = True,
    handle_arrays: bool = True,
    clear_spill_range: bool = True,
    spill_rows: int = 200  # Increased default spill rows to ensure adequate space
) -> Dict[str, Any]:
    """Apply any Excel formula to a specific cell, with error handling.
    
    This function applies Excel formulas and helps protect against common errors like 
    division by zero by optionally wrapping formulas in IFERROR when needed.
    Also supports modern array formulas that spill results.
    
    Args:
        filepath: Path to the Excel workbook
        sheet_name: Name of the worksheet
        cell: Target cell reference (e.g., 'A1')
        formula: Excel formula to apply (with or without leading '=')
        protect_from_errors: Whether to auto-protect against errors like division by zero
        handle_arrays: Whether to properly handle modern array formulas (XLOOKUP, FILTER, etc.)
        clear_spill_range: Whether to automatically clear potential spill range for array formulas
        spill_rows: Number of rows to clear below for vertical spill formulas (default 200)
        
    Returns:
        Dictionary with success status and message
    """
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found"
            }
        ws = wb[sheet_name]
        
        # Validate formula
        validation = validate_excel_formula(formula)
        formula_str = validation["formula"]
        
        # Provide warnings for issues found
        warnings = []
        if not validation["is_valid"]:
            warnings.extend(validation["issues"])
        
        # Apply division by zero protection if needed and requested
        if protect_from_errors and validation["needs_safety_wrapper"]:
            original_formula = formula_str
            formula_str = f"=IFERROR({formula_str[1:]},\"\")"
            warnings.append(f"Added error protection: {original_formula} → {formula_str}")
        
        # Check for external references
        if validation["external_references"]:
            ext_refs = validation["external_references"]
            ext_warnings = []
            for ext_ref in ext_refs:
                # Check if external file exists relative to current file
                ext_path = os.path.join(os.path.dirname(filepath), ext_ref)
                if not os.path.exists(ext_path):
                    ext_warnings.append(f"External workbook not found: {ext_ref}")
            if ext_warnings:
                warnings.extend(ext_warnings)

        # Check for array formula references (like A2#)
        has_array_reference = bool(re.search(ARRAY_REFERENCE_PATTERN, formula_str))
        if has_array_reference:
            warnings.append(
                "Formula contains spilled array references (using # notation). "
                "Ensure the referenced array formulas are correctly set up in the worksheet."
            )
                
        # Handle array formulas specially if requested
        if handle_arrays and validation["is_array_formula"]:
            # Parse the target cell to get coordinates
            try:
                row_idx, col_idx = parse_cell_reference(cell)
            except ValidationError as e:
                return {
                    "success": False,
                    "error": f"Invalid cell reference: {str(e)}"
                }
                
            # For array formulas, especially UNIQUE, clear potential spill range
            # to ensure formula has room to display all results
            if clear_spill_range:
                # Clear more rows for UNIQUE, FILTER, etc. to ensure adequate space for spilling
                if validation["array_formula_type"] in ["UNIQUE", "FILTER", "SORT", "SORTBY"]:
                    max_potential_rows = min(spill_rows, ws.max_row - row_idx + 1)
                    
                    # Clear cells below
                    cells_cleared = 0
                    for r in range(row_idx + 1, row_idx + max_potential_rows + 1):
                        if ws.cell(row=r, column=col_idx).value is not None:
                            ws.cell(row=r, column=col_idx).value = None
                            cells_cleared += 1
                    
                    warnings.append(
                        f"Cleared {cells_cleared} cells below {cell} to ensure array formula can spill all results."
                    )
                
                # If it's a formula that might spill horizontally (like TRANSPOSE)
                if validation["array_formula_type"] in ["TRANSPOSE"]:
                    max_potential_cols = min(30, ws.max_column - col_idx + 1)
                    for c in range(col_idx + 1, col_idx + max_potential_cols + 1):
                        if ws.cell(row=row_idx, column=c).value is not None:
                            ws.cell(row=row_idx, column=c).value = None
            
            # Apply the array formula to the target cell
            try:
                ws[cell].value = formula_str
                # Additional sleep to allow Excel to process the formula completely
                # This can help with complex array formulas especially during rapid updates
                time.sleep(0.1)
            except Exception as e:
                return {
                    "success": False,
                    "error": f"Error applying array formula: {str(e)}"
                }
            
            # Add specific message for UNIQUE formula to help user understand behavior
            if validation["array_formula_type"] == "UNIQUE":
                warnings.append(
                    "UNIQUE formula will dynamically spill results downward as needed. "
                    f"Cleared up to {spill_rows} cells below to ensure proper display."
                )
            else:
                warnings.append(
                    f"Applied as array formula ({validation['array_formula_type']}). "
                    f"Space below/right has been cleared for results to spill properly."
                )
        else:
            # Assign regular formula to cell
            ws[cell].value = formula_str
            
        # Save workbook
        try:
            wb.save(filepath)
            wb.close()
        except Exception as e:
            return {
                "success": False,
                "error": f"Error saving workbook after applying formula: {str(e)}"
            }
        
        result = {
            "success": True,
            "message": f"Formula '{formula_str}' applied to {cell} in sheet '{sheet_name}'"
        }
        
        if warnings:
            result["warnings"] = warnings
            result["message"] += f" (with {len(warnings)} warning{'s' if len(warnings) > 1 else ''})"
            
        # Add additional data when appropriate
        if validation["is_array_formula"]:
            result["array_formula_type"] = validation["array_formula_type"]
            result["is_spill_formula"] = True
            result["cells_cleared_for_spill"] = clear_spill_range
            result["spill_rows_cleared"] = spill_rows if clear_spill_range else 0
        if validation["external_references"]:
            result["external_references"] = validation["external_references"]
            
        return result
    except Exception as e:
        logger.error(f"Failed to apply formula: {e}")
        if 'wb' in locals() and wb is not None:
            try:
                wb.close()
            except:
                pass
        return {
            "success": False,
            "error": f"Failed to apply formula: {str(e)}"
        }

def apply_excel_formula_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    formula_template: str,
    protect_from_errors: bool = True,
    dynamic_calculation: bool = True,
    chunk_size: int = 1000,
    clear_spill_range: bool = True
) -> Dict[str, Any]:
    """Apply an Excel formula to a range of cells, with error handling.
    
    This function takes a template formula with placeholders and applies it to each cell
    in the specified range, with support for large datasets and modern array formulas.
    
    Args:
        filepath: Path to the Excel workbook
        sheet_name: Name of the worksheet
        start_cell: Top-left cell of the range (e.g., 'L2')
        end_cell: Bottom-right cell of the range (e.g., 'L36')
        formula_template: Formula template with {row} and {col} placeholders 
                          (e.g., 'IF(F{row}>=90,"Excellent",IF(F{row}>=80,"Good",...))')
        protect_from_errors: Whether to auto-protect against errors like division by zero
        dynamic_calculation: Whether to use dynamic array functions for better performance
        chunk_size: How many cells to process at once (for large ranges)
        clear_spill_range: Whether to automatically clear potential spill range for array formulas
        
    Returns:
        Dictionary with success status and message
    """
    try:
        wb = load_workbook(filepath)
        if sheet_name not in wb.sheetnames:
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found"
            }
        
        ws = wb[sheet_name]
        
        # Parse cell range
        try:
            min_col, min_row, max_col, max_row = range_boundaries(f"{start_cell}:{end_cell}")
        except Exception as e:
            return {
                "success": False,
                "error": f"Invalid cell range: {start_cell}:{end_cell} - {str(e)}"
            }
        
        # Calculate range size to determine if we need chunking for large datasets
        total_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
        use_chunking = total_cells > chunk_size
        
        # Ensure formula template has placeholders or warn
        warnings = []
        has_row_placeholder = "{row}" in formula_template
        has_col_placeholder = "{col}" in formula_template
        
        if not (has_row_placeholder or has_col_placeholder):
            warnings.append("No {row} or {col} placeholder found in formula template. Using static formula.")
        
        # Check if this is a modern array formula
        validation = validate_excel_formula(formula_template.replace("{row}", str(min_row)).replace("{col}", get_column_letter(min_col)))
        is_array_formula = validation["is_array_formula"]
        
        # For array formulas with dynamic calculation, we can optimize by applying once
        if is_array_formula and dynamic_calculation:
            # Apply just to the top-left cell and let it spill
            top_left_cell = f"{get_column_letter(min_col)}{min_row}"
            cell_formula = formula_template.replace("{row}", str(min_row)).replace("{col}", get_column_letter(min_col))
            
            # For array formulas like UNIQUE, we need to handle the spill behavior properly
            # by calling our enhanced apply_excel_formula function
            result = apply_excel_formula(
                filepath, sheet_name, top_left_cell, cell_formula, 
                protect_from_errors, handle_arrays=True, clear_spill_range=clear_spill_range
            )
            
            if not result["success"]:
                return result
                
            # Add array formula info
            result["is_array_formula"] = True
            result["cells_affected"] = total_cells
            result["applied_as"] = "single_array_formula"
            
            # Add specific note about array formula behavior based on the formula type
            if validation["array_formula_type"] == "UNIQUE":
                warnings.append(
                    "UNIQUE formula is designed to spill results automatically. "
                    "The formula has been applied to the top cell only, and will display all unique values below."
                )
            elif validation["array_formula_type"] in ["FILTER", "SORT", "SORTBY"]:
                warnings.append(
                    f"{validation['array_formula_type']} formula is designed to spill results automatically. "
                    "The formula has been applied to the top cell only."
                )
                
            if warnings:
                if "warnings" not in result:
                    result["warnings"] = []
                result["warnings"].extend(warnings)
                
            return result
        
        # Process in chunks for large datasets
        formulas_applied = 0
        errors = []
        
        # Process cells in chunks if needed
        if use_chunking:
            # Process cells in chunks
            for row_start in range(min_row, max_row + 1, chunk_size):
                row_end = min(row_start + chunk_size - 1, max_row)
                
                # Apply formulas to this chunk
                for row in range(row_start, row_end + 1):
                    for col in range(min_col, max_col + 1):
                        # Generate cell-specific formula
                        cell_formula = formula_template.replace("{row}", str(row)).replace("{col}", get_column_letter(col))
                        
                        # Apply formula directly
                        cell = ws.cell(row=row, column=col)
                        
                        try:
                            # Validate and protect if needed
                            cell_validation = validate_excel_formula(cell_formula)
                            final_formula = cell_validation["formula"]
                            
                            # Apply division by zero protection if needed and requested
                            if protect_from_errors and cell_validation["needs_safety_wrapper"]:
                                final_formula = f"=IFERROR({final_formula[1:]},\"\")"
                            
                            # Apply formula to cell
                            cell.value = final_formula
                            formulas_applied += 1
                        except Exception as cell_error:
                            errors.append(f"Error at {get_column_letter(col)}{row}: {str(cell_error)}")
                
                # Save after each chunk to minimize memory usage
                wb.save(filepath)
        else:
            # Apply formulas to all cells in the range
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    # Generate cell-specific formula
                    cell_formula = formula_template.replace("{row}", str(row)).replace("{col}", get_column_letter(col))
                    
                    # Validate and protect if needed
                    cell_validation = validate_excel_formula(cell_formula)
                    final_formula = cell_validation["formula"]
                    
                    # Apply division by zero protection if needed and requested
                    if protect_from_errors and cell_validation["needs_safety_wrapper"]:
                        final_formula = f"=IFERROR({final_formula[1:]},\"\")"
                    
                    # Apply formula to cell
                    cell = ws.cell(row=row, column=col)
                    cell.value = final_formula
                    formulas_applied += 1
            
            # Save workbook once for small ranges
            wb.save(filepath)
        
        wb.close()
        
        # Return success message
        range_str = f"{start_cell}:{end_cell}"
        result = {
            "success": True,
            "message": f"Applied formula to {formulas_applied} cells in range {range_str} of sheet '{sheet_name}'",
            "cells_affected": formulas_applied,
            "processed_with_chunking": use_chunking
        }
        
        if warnings:
            result["warnings"] = warnings
            result["message"] += f" (with {len(warnings)} warning{'s' if len(warnings) > 1 else ''})"
            
        if errors:
            result["errors"] = errors[:10]  # Limit to first 10 errors
            if len(errors) > 10:
                result["errors"].append(f"...and {len(errors) - 10} more errors")
            
        return result
        
    except Exception as e:
        logger.error(f"Failed to apply formula range: {e}")
        if 'wb' in locals() and wb is not None:
            wb.close()
        return {
            "success": False,
            "error": f"Failed to apply formula range: {str(e)}"
        }

def add_excel_column(
    filepath: str,
    sheet_name: str,
    column_name: str,
    column_position: Optional[str] = None,
    data: Optional[List[Any]] = None,
    header_style: Optional[Dict[str, Any]] = None
) -> Dict[str, Any]:
    """Add a new column to an existing Excel worksheet without rewriting the entire sheet.
    
    This function intelligently inserts a new column at a specific position or at the end of existing data.
    It handles formatting the header and can optionally insert data for the new column.
    
    Args:
        filepath: Path to the Excel workbook
        sheet_name: Name of the worksheet to modify
        column_name: Name for the new column (header text)
        column_position: Optional column letter where to insert (e.g., 'C' to insert as third column).
                         If not provided, adds to the end of existing data.
        data: Optional list of values for the column. Length should match the data rows in the sheet.
        header_style: Optional dictionary of styling parameters for the header
                     (e.g., {'bold': True, 'bg_color': 'FFFF00'})
    
    Returns:
        Dictionary with success status and message
    """
    try:
        # Load the workbook
        wb = load_workbook(filepath)
        
        # Validate sheet exists
        if sheet_name not in wb.sheetnames:
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found"
            }
        
        ws = wb[sheet_name]
        
        # Find the current dimensions
        min_col, min_row, max_col, max_row = range_boundaries(ws.calculate_dimension())
        
        # Default header style
        default_header_style = {
            "bold": True,
            "bg_color": None,  # No background color by default
            "font_size": None,  # Keep default font size
            "alignment": "center"
        }
        
        # Merge provided header style with defaults
        if header_style:
            for key, value in header_style.items():
                default_header_style[key] = value
        
        # Determine insertion column index
        insert_col_idx = max_col + 1  # Default to end of data
        if column_position:
            try:
                insert_col_idx = column_index_from_string(column_position)
                # If inserting within existing data, shift columns
                if insert_col_idx <= max_col:
                    ws.insert_cols(insert_col_idx)
            except ValueError:
                return {
                    "success": False,
                    "error": f"Invalid column position: {column_position}"
                }
        
        # Add the header
        header_cell = ws.cell(row=min_row, column=insert_col_idx)
        header_cell.value = column_name
        
        # Apply header styling
        header_cell.font = Font(bold=default_header_style["bold"])
        if default_header_style["font_size"]:
            header_cell.font = Font(bold=default_header_style["bold"], size=default_header_style["font_size"])
        
        if default_header_style["bg_color"]:
            bg_color = default_header_style["bg_color"]
            if bg_color.startswith("#"):
                bg_color = bg_color[1:]  # Remove # if present
            header_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
        if default_header_style["alignment"]:
            header_cell.alignment = Alignment(horizontal=default_header_style["alignment"])
        
        # Add data if provided
        if data:
            for i, value in enumerate(data, start=1):
                # Skip the header row (min_row), start from the next row
                row_idx = min_row + i
                # Don't exceed existing data rows
                if max_row >= row_idx:
                    cell = ws.cell(row=row_idx, column=insert_col_idx)
                    cell.value = value
        
        # Save the workbook
        wb.save(filepath)
        wb.close()
        
        # Prepare success message
        col_letter = get_column_letter(insert_col_idx)
        return {
            "success": True,
            "message": f"Added column '{column_name}' at position {col_letter} in sheet '{sheet_name}'",
            "column_position": col_letter,
            "data_rows_affected": len(data) if data else 0
        }
    
    except Exception as e:
        logger.error(f"Failed to add column: {e}")
        if 'wb' in locals():
            wb.close()
        return {
            "success": False,
            "error": f"Failed to add column: {str(e)}"
        }

def add_data_validation(
    filepath: str,
    sheet_name: str,
    cell_range: str,
    validation_type: str,
    validation_criteria: Dict[str, Any],
    error_message: Optional[str] = None
) -> Dict[str, Any]:
    """Add data validation rules to Excel cells.
    
    Args:
        filepath: Path to Excel workbook
        sheet_name: Target worksheet name
        cell_range: Range to apply validation to (e.g., "A1:A10")
        validation_type: Type of validation ("list", "decimal", "date", "textLength", "custom")
        validation_criteria: Dictionary with validation parameters
            For list: {"source": ["Option1", "Option2"] or "=Sheet2!A1:A10"}
            For decimal: {"operator": "between", "minimum": 1, "maximum": 100}
            For custom: {"formula": "=AND(A1>0,A1<100)"}
        error_message: Optional custom error message
        
    Returns:
        Result dictionary with success status
    """
    try:
        wb = load_workbook(filepath)
        
        # Validate sheet
        if sheet_name not in wb.sheetnames:
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found"
            }
            
        ws = wb[sheet_name]
        
        # Create appropriate validation rule based on type
        if validation_type == "list":
            source = validation_criteria.get("source")
            if isinstance(source, list):
                source_str = ",".join(f'"{item}"' for item in source)
                dv = DataValidation(type="list", formula1=f"{source_str}")
            else:
                dv = DataValidation(type="list", formula1=source)
                
        elif validation_type == "decimal":
            operator = validation_criteria.get("operator", "between")
            minimum = validation_criteria.get("minimum")
            maximum = validation_criteria.get("maximum")
            
            if operator == "between":
                dv = DataValidation(type="decimal", operator="between", 
                                   formula1=str(minimum), formula2=str(maximum))
            else:
                dv = DataValidation(type="decimal", operator=operator, 
                                   formula1=str(minimum))
                                   
        elif validation_type == "custom":
            formula = validation_criteria.get("formula")
            dv = DataValidation(type="custom", formula1=formula)
        
        # Set error message if provided
        if error_message:
            dv.errorTitle = "Invalid Input"
            dv.error = error_message
            dv.errorStyle = "stop"
        
        # Add the validation to the worksheet
        dv.add(cell_range)
        ws.add_data_validation(dv)
        
        wb.save(filepath)
        return {
            "success": True,
            "message": f"Data validation added to {cell_range} in sheet '{sheet_name}'"
        }
    except Exception as e:
        logger.error(f"Failed to add data validation: {e}")
        return {
            "success": False,
            "error": f"Failed to add data validation: {str(e)}"
        }

def apply_conditional_formatting(
    filepath: str,
    sheet_name: str,
    cell_range: str,
    condition: str,
    bold: bool = False,
    italic: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False,
    border_style: Optional[str] = None,
    condition_column: Optional[str] = None,
    format_entire_row: bool = False,
    columns_to_format: Optional[List[str]] = None
) -> Dict[str, Any]:
    """Apply conditional formatting to cells based on a specified condition.
    
    This function efficiently applies formatting to cells in a range that meet specified criteria,
    eliminating the need to manually filter and format cells one by one.
    
    Args:
        filepath: Path to the Excel workbook
        sheet_name: Name of the worksheet
        cell_range: Range of cells to check and potentially format (e.g., 'A1:D10')
        condition: Condition for formatting, using syntax like ">90", "='Yes'", "CONTAINS('text')"
                   Example formats:
                   - Numeric: ">90", "<=50", "=100", "<>0" (not equal to zero)
                   - Text: "='Yes'", "<>'No'", "CONTAINS('text')", "STARTS_WITH('A')"
                   - Date: ">DATE(2023,1,1)", "<=TODAY()"
                   - Blank: "=ISBLANK()", "<>ISBLANK()"
                   - Compound conditions: ">50 AND <=70", "<20 OR >80", "CONTAINS('Pass') OR =ISBLANK()"
        bold: Whether to apply bold formatting to matching cells
        italic: Whether to apply italic formatting to matching cells
        font_size: Optional font size to apply to matching cells
        font_color: Optional font color (hex code, with or without #) for matching cells
        bg_color: Optional background color (hex code, with or without #) for matching cells 
        alignment: Optional text alignment ('left', 'center', 'right') for matching cells
        wrap_text: Whether to enable text wrapping for matching cells
        border_style: Optional border style ('thin', 'medium', 'thick', 'dashed', 'dotted', 'double')
        condition_column: Optional column letter to evaluate the condition on (e.g., 'D' for Units)
        format_entire_row: Whether to format the entire row when condition is met (default False)
        columns_to_format: Optional list of specific column letters to format when condition is met
                           (e.g. ['A', 'C', 'D']). Only used when format_entire_row is False or
                           when you want to format specific columns in the row
    
    Returns:
        Dictionary with success status, message, and count of formatted cells
    """
    try:
        # Load workbook
        wb = load_workbook(filepath)
        
        # Validate sheet
        if sheet_name not in wb.sheetnames:
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found"
            }
            
        worksheet = wb[sheet_name]
        
        # Parse cell range
        try:
            min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        except Exception as e:
            return {
                "success": False,
                "error": f"Invalid cell range: {cell_range} - {str(e)}"
            }
        
        # If condition_column is specified, convert to column index
        condition_col_idx = None
        if condition_column:
            try:
                condition_col_idx = column_index_from_string(condition_column.upper().strip())
                # Verify the condition column is within the range or at least a valid column
                if condition_col_idx < 1:
                    return {
                        "success": False,
                        "error": f"Invalid condition column: {condition_column}"
                    }
            except Exception as e:
                return {
                    "success": False,
                    "error": f"Invalid condition column: {condition_column} - {str(e)}"
                }
        
        # If columns_to_format is specified, convert to column indices
        format_col_indices = []
        if columns_to_format:
            for col_letter in columns_to_format:
                try:
                    col_idx = column_index_from_string(col_letter.upper().strip())
                    format_col_indices.append(col_idx)
                except Exception as e:
                    return {
                        "success": False,
                        "error": f"Invalid format column: {col_letter} - {str(e)}"
                    }
            
            # If columns_to_format is specified, we'll format only those specific columns
            # regardless of what format_entire_row is set to
            should_format_specific_columns = columns_to_format is not None and len(columns_to_format) > 0
            
            # If we're formatting specific columns, ensure format_entire_row is False
            if should_format_specific_columns:
                format_entire_row = False
        else:
            should_format_specific_columns = False
        
        # Set up font formatting options
        font_args = {}
        if bold:
            font_args["bold"] = True
        if italic:
            font_args["italic"] = True
        if font_size:
            font_args["size"] = font_size
        if font_color:
            # Handle color format (with or without #)
            if font_color.startswith("#"):
                font_color = font_color[1:]
            font_args["color"] = font_color
        
        # Set up fill
        fill = None
        if bg_color:
            # Handle color format (with or without #)
            if bg_color.startswith("#"):
                bg_color = bg_color[1:]
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        
        # Set up alignment
        align = None
        if alignment or wrap_text:
            align_horz = None
            if alignment:
                align_horz = {
                    "left": "left",
                    "center": "center",
                    "right": "right"
                }.get(alignment.lower())
            align = Alignment(horizontal=align_horz, wrap_text=wrap_text)
        
        # Set up border
        border = None
        if border_style:
            border_styles = {
                "thin": Side(style="thin"),
                "medium": Side(style="medium"),
                "thick": Side(style="thick"),
                "dashed": Side(style="dashed"),
                "dotted": Side(style="dotted"),
                "double": Side(style="double")
            }
            if border_style.lower() in border_styles:
                side = border_styles[border_style.lower()]
                border = Border(left=side, right=side, top=side, bottom=side)
        
        # Parse the condition
        formatted_cells_count = 0
        condition = condition.strip()
        
        # Enhanced helper functions to evaluate conditions with support for complex logic
        def evaluate_numeric_condition(cell_value, condition):
            if not isinstance(cell_value, (int, float)) and not str(cell_value).replace('.', '', 1).isdigit():
                return False
                
            try:
                cell_value = float(cell_value) if cell_value is not None else 0
                
                # Check for compound conditions (AND, OR)
                if " AND " in condition.upper():
                    parts = condition.upper().split(" AND ")
                    return all(evaluate_numeric_condition(cell_value, part.strip()) for part in parts)
                    
                if " OR " in condition.upper():
                    parts = condition.upper().split(" OR ")
                    return any(evaluate_numeric_condition(cell_value, part.strip()) for part in parts)
                
                # Single conditions
                if condition.startswith(">="):
                    threshold = float(condition[2:].strip())
                    return cell_value >= threshold
                elif condition.startswith("<="):
                    threshold = float(condition[2:].strip())
                    return cell_value <= threshold
                elif condition.startswith("<>"):
                    threshold = float(condition[2:].strip())
                    return cell_value != threshold
                elif condition.startswith(">"):
                    threshold = float(condition[1:].strip())
                    return cell_value > threshold
                elif condition.startswith("<"):
                    threshold = float(condition[1:].strip())
                    return cell_value < threshold
                elif condition.startswith("="):
                    threshold = float(condition[1:].strip())
                    return cell_value == threshold
                return False
            except (ValueError, TypeError):
                return False
        
        def evaluate_text_condition(cell_value, condition):
            if cell_value is None:
                cell_value = ""
            cell_value = str(cell_value).lower()
            
            # Check for compound conditions (AND, OR)
            if " AND " in condition.upper():
                parts = condition.upper().split(" AND ")
                return all(evaluate_text_condition(cell_value, part.strip()) for part in parts)
                
            if " OR " in condition.upper():
                parts = condition.upper().split(" OR ")
                return any(evaluate_text_condition(cell_value, part.strip()) for part in parts)
            
            if condition.startswith("='"):
                # Exact match (case insensitive)
                text = condition[2:-1].lower() if condition.endswith("'") else condition[2:].lower()
                return cell_value == text
            elif condition.startswith("<>'"):
                # Not equal match (case insensitive)
                text = condition[3:-1].lower() if condition.endswith("'") else condition[3:].lower()
                return cell_value != text
            elif condition.upper().startswith("CONTAINS('"):
                # Contains text (case insensitive)
                text = condition[10:-2].lower() if condition.endswith("')") else condition[10:-1].lower()
                return text in cell_value
            elif condition.upper().startswith("STARTS_WITH('"):
                # Starts with text (case insensitive)
                text = condition[13:-2].lower() if condition.endswith("')") else condition[13:-1].lower()
                return cell_value.startswith(text)
            elif condition.upper().startswith("ENDS_WITH('"):
                # Ends with text (case insensitive)
                text = condition[11:-2].lower() if condition.endswith("')") else condition[11:-1].lower()
                return cell_value.endswith(text)
            
            return False
        
        def evaluate_blank_condition(cell_value, condition):
            is_blank = cell_value is None or str(cell_value).strip() == ""
            
            # Check for compound conditions (AND, OR)
            if " AND " in condition.upper():
                parts = condition.upper().split(" AND ")
                return all(evaluate_blank_condition(cell_value, part.strip()) for part in parts)
                
            if " OR " in condition.upper():
                parts = condition.upper().split(" OR ")
                return any(evaluate_blank_condition(cell_value, part.strip()) for part in parts)
            
            if condition.upper() == "=ISBLANK()":
                return is_blank
            elif condition.upper() == "<>ISBLANK()":
                return not is_blank
            return False
        
        # Master condition evaluator that handles compound conditions across different types
        def evaluate_condition(cell_value, condition):
            # First check for top-level compound conditions
            if " AND " in condition.upper():
                parts = condition.upper().split(" AND ")
                return all(evaluate_condition(cell_value, part.strip()) for part in parts)
                
            if " OR " in condition.upper():
                parts = condition.upper().split(" OR ")
                return any(evaluate_condition(cell_value, part.strip()) for part in parts)
            
            # Then evaluate based on condition type
            # Numeric conditions
            if any(condition.startswith(op) for op in ["=", ">", "<", "<=", ">=", "<>"]) and not condition.startswith("='") and not condition.startswith("<>'"):
                return evaluate_numeric_condition(cell_value, condition)
            
            # Text conditions
            elif condition.startswith("='") or condition.startswith("<>'") or \
                 condition.upper().startswith(("CONTAINS('"), ("STARTS_WITH('"), ("ENDS_WITH(")):
                return evaluate_text_condition(cell_value, condition)
            
            # Blank/not blank conditions
            elif condition.upper() in ["=ISBLANK()", "<>ISBLANK()"]:
                return evaluate_blank_condition(cell_value, condition)
                
            # Fallback
            return False
            
        # Track rows to format based on condition
        rows_to_format = set()
        cells_to_format = []
        
        # Evaluate condition for each row or cell in the range
        for row in range(min_row, max_row + 1):
            # Skip header row if it exists (first row of range)
            if row == min_row and min_row != max_row:  # Skip only if we have multiple rows
                # Assume first row is header and skip evaluation
                continue
                
            # If we're checking a specific column for the condition
            if condition_col_idx:
                cell = worksheet.cell(row=row, column=condition_col_idx)
                cell_value = cell.value
                
                # Use the master condition evaluator
                should_format = evaluate_condition(cell_value, condition)
                
                # If condition is met, add row to formatting list
                if should_format:
                    if format_entire_row or should_format_specific_columns:
                        rows_to_format.add(row)
                    else:
                        cells_to_format.append((row, condition_col_idx))
            else:
                # Original behavior: check each cell in the range
                for col in range(min_col, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell_value = cell.value
                    
                    # Use the master condition evaluator
                    should_format = evaluate_condition(cell_value, condition)
                    
                    # If condition is met, add cell to formatting list
                    if should_format:
                        if format_entire_row or should_format_specific_columns:
                            rows_to_format.add(row)
                            break  # Once we know the row should be formatted, no need to check other cells
                        else:
                            cells_to_format.append((row, col))
        
        # Apply formatting to entire rows if needed
        if format_entire_row and rows_to_format:
            for row in rows_to_format:
                for col in range(min_col, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    
                    # Apply font if any font properties set
                    if font_args:
                        # Start with existing font and update properties
                        new_font = Font(
                            name=cell.font.name,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            size=cell.font.size,
                            color=cell.font.color
                        )
                        # Update with specified properties
                        for key, value in font_args.items():
                            setattr(new_font, key, value)
                        cell.font = new_font
                    
                    # Apply fill if specified
                    if fill:
                        cell.fill = fill
                        
                    # Apply alignment if specified
                    if align:
                        cell.alignment = align
                        
                    # Apply border if specified
                    if border:
                        cell.border = border
                    
                    formatted_cells_count += 1
        # Apply formatting to specific columns in matching rows
        elif should_format_specific_columns and rows_to_format:
            for row in rows_to_format:
                for col_idx in format_col_indices:
                    cell = worksheet.cell(row=row, column=col_idx)
                    
                    # Apply font if any font properties set
                    if font_args:
                        # Start with existing font and update properties
                        new_font = Font(
                            name=cell.font.name,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            size=cell.font.size,
                            color=cell.font.color
                        )
                        # Update with specified properties
                        for key, value in font_args.items():
                            setattr(new_font, key, value)
                        cell.font = new_font
                    
                    # Apply fill if specified
                    if fill:
                        cell.fill = fill
                        
                    # Apply alignment if specified
                    if align:
                        cell.alignment = align
                        
                    # Apply border if specified
                    if border:
                        cell.border = border
                    
                    formatted_cells_count += 1
        else:
            # Apply formatting to individual cells
            for row, col in cells_to_format:
                cell = worksheet.cell(row=row, column=col)
                
                # Apply font if any font properties set
                if font_args:
                    # Start with existing font and update properties
                    new_font = Font(
                        name=cell.font.name,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        size=cell.font.size,
                        color=cell.font.color
                    )
                    # Update with specified properties
                    for key, value in font_args.items():
                        setattr(new_font, key, value)
                    cell.font = new_font
                
                # Apply fill if specified
                if fill:
                    cell.fill = fill
                    
                # Apply alignment if specified
                if align:
                    cell.alignment = align
                    
                # Apply border if specified
                if border:
                    cell.border = border
                
                formatted_cells_count += 1
        
        # Save workbook
        wb.save(filepath)
        wb.close()
        
        # Format message based on formatting mode
        format_mode = "rows" if format_entire_row else "cells"
        if columns_to_format:
            format_mode = f"specified columns ({', '.join(columns_to_format)}) in matching rows"
        
        condition_info = f" in column {condition_column}" if condition_column else ""
        
        return {
            "success": True,
            "message": f"Conditional formatting applied to {formatted_cells_count} {format_mode} matching condition '{condition}'{condition_info} in range {cell_range}",
            "cells_formatted": formatted_cells_count,
            "condition_applied": condition,
            "formatted_entire_rows": format_entire_row
        }
    except Exception as e:
        logger.error(f"Failed to apply conditional formatting: {e}")
        if 'wb' in locals():
            wb.close()
        return {
            "success": False,
            "error": f"Failed to apply conditional formatting: {str(e)}"
        } 

def evaluate_excel_formula(formula_str: str, context_values: dict = None) -> Dict[str, Any]:
    """Evaluate an Excel formula using the formulas library.
    
    Args:
        formula_str: The Excel formula string to evaluate (with or without leading =)
        context_values: Optional dictionary mapping cell references to values for variables in the formula
        
    Returns:
        Dictionary with evaluation result and metadata
    """
    if not has_formula_engine:
        return {
            "success": False,
            "error": "Formulas library not installed. Install with: pip install formulas",
            "formula": formula_str,
            "value": None
        }
        
    try:
        # Ensure formula starts with '='
        if not formula_str.startswith('='):
            formula_str = f"={formula_str}"
            
        # Create a parser and parse the formula
        parser = Parser()
        ast = parser.ast(formula_str)[1].compile()
        
        # If we have context values, use them for evaluation
        if context_values:
            inputs = context_values
        else:
            inputs = {}
            
        # Evaluate the formula
        result = ast(inputs)
        
        # Handle special types for JSON serialization
        try:
            import numpy as np
            if isinstance(result, np.ndarray):
                result = result.item()  # Convert numpy array to Python scalar
        except (ImportError, AttributeError, Exception):
            pass
            
        # Convert other types if needed
        if not isinstance(result, (int, float, str, bool, type(None))):
            try:
                result = float(result)
            except (TypeError, ValueError):
                try:
                    result = str(result)
                except Exception:
                    result = None
            
        return {
            "success": True,
            "formula": formula_str,
            "value": result,
            "value_type": type(result).__name__
        }
        
    except Exception as e:
        # For simple multiplication formulas, try manual evaluation
        if "*" in formula_str and formula_str.count("*") == 1 and context_values:
            try:
                # Extract cell references
                formula_parts = formula_str.strip("=").split("*")
                if len(formula_parts) == 2:
                    ref1, ref2 = formula_parts[0].strip(), formula_parts[1].strip()
                    val1 = context_values.get(ref1)
                    val2 = context_values.get(ref2)
                    
                    if val1 is not None and val2 is not None:
                        try:
                            result = float(val1) * float(val2)
                            return {
                                "success": True,
                                "formula": formula_str,
                                "value": result,
                                "value_type": type(result).__name__,
                                "evaluated_by": "manual_calculation"
                            }
                        except (TypeError, ValueError):
                            pass
            except Exception:
                pass
        
        return {
            "success": False,
            "error": f"Formula evaluation error: {str(e)}",
            "formula": formula_str,
            "value": None
        }

def read_excel_with_formulas(filepath: str, sheet_name: str = None, cell_range: str = None) -> Dict[str, Any]:
    """Read Excel data with enhanced formula handling.
    
    This function reads Excel data and handles formulas properly, returning both the 
    formula text and evaluated values. It uses a combination of openpyxl and the formulas
    library to provide complete formula information.
    
    Args:
        filepath: Path to the Excel file
        sheet_name: Optional specific sheet to read. If None, returns file info only
        cell_range: Optional cell range to read (e.g. 'A1:D10'). If None, reads all cells
        
    Returns:
        Dictionary containing file info and optionally sheet data with formula handling
    """
    try:
        # First load without data_only to get formulas
        workbook_formulas = openpyxl.load_workbook(filepath, data_only=False)
        
        # Then load with data_only to get calculated values
        workbook_values = openpyxl.load_workbook(filepath, data_only=True)
        
        # Basic file info
        file_info = {
            "file": os.path.basename(filepath),
            "sheets": []
        }
        
        # Get info for all sheets
        for ws_name in workbook_formulas.sheetnames:
            ws_formulas = workbook_formulas[ws_name]
            ws_values = workbook_values[ws_name]
            
            # Get sheet dimensions
            min_col, min_row, max_col, max_row = range_boundaries(ws_formulas.calculate_dimension())
            
            # Get column headers (first row)
            columns = []
            column_refs = []
            for col in range(min_col, max_col + 1):
                cell = ws_formulas.cell(min_row, col)
                header = cell.value if cell.value is not None else f"Column {get_column_letter(col)}"
                columns.append(header)
                column_refs.append(get_column_letter(col))
            
            sheet_info = {
                "name": ws_name,
                "dimensions": ws_formulas.calculate_dimension(),
                "row_count": ws_formulas.max_row,
                "column_count": max_col - min_col + 1,
                "columns": columns,
                "column_refs": column_refs
            }
            file_info["sheets"].append(sheet_info)
            
            # If this is the requested sheet, collect detailed data
            if sheet_name and ws_name == sheet_name:
                # Parse cell range if provided
                if cell_range:
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
                    except ValueError:
                        return {
                            "success": False,
                            "error": f"Invalid cell range format: {cell_range}"
                        }
                
                # Get column headers (first row)
                columns = []
                for col in range(min_col, max_col + 1):
                    cell = ws_formulas.cell(min_row, col)
                    columns.append(cell.value if cell.value is not None else "")
                
                # Build context values for formula evaluation - collect all cell values from data_only workbook
                context_values = {}
                all_values = {}
                
                # First pass - collect all values from the data_only workbook
                for r in range(1, ws_values.max_row + 1):
                    for c in range(1, ws_values.max_column + 1):
                        cell_ref = f"{get_column_letter(c)}{r}"
                        cell_value = ws_values.cell(row=r, column=c).value
                        
                        # Skip None values to avoid formula evaluation issues
                        if cell_value is not None:
                            # Handle datetime objects for formula calculations
                            if isinstance(cell_value, datetime):
                                import datetime as dt
                                excel_epoch = dt.datetime(1899, 12, 30)
                                delta = cell_value - excel_epoch
                                excel_date = delta.days + (delta.seconds / 86400)
                                all_values[cell_ref] = excel_date
                            else:
                                all_values[cell_ref] = cell_value
                
                # Collect all cells including formula info
                records = []
                formula_cells = []
                
                for row in range(min_row + 1, max_row + 1):
                    values = []
                    formulas = []
                    cell_refs = []
                    row_has_data = False
                    row_has_formula = False
                    
                    for col in range(min_col, max_col + 1):
                        # Get both formula and value versions
                        cell_formula = ws_formulas.cell(row, col)
                        cell_value = ws_values.cell(row, col)
                        cell_ref = f"{get_column_letter(col)}{row}"
                        cell_refs.append(cell_ref)
                        
                        # Check if cell has formula
                        has_formula = cell_formula.data_type == 'f' if hasattr(cell_formula, 'data_type') else False
                        formula = cell_formula.value if has_formula else None
                        
                        # Get calculated value
                        value = cell_value.value
                        
                        # Convert datetime objects to ISO format for JSON serialization
                        if isinstance(value, datetime):
                            value = value.isoformat()
                        
                        # If openpyxl couldn't calculate the formula value, try with formulas library
                        if has_formula and (value is None) and has_formula_engine and formula:
                            try:
                                # Try to evaluate the formula using all collected context values
                                eval_result = evaluate_excel_formula(formula, all_values)
                                if eval_result.get("success", False):
                                    value = eval_result["value"]
                                    
                                    # If value is numpy array or other complex type, convert to Python native type
                                    try:
                                        import numpy as np
                                        if isinstance(value, np.ndarray):
                                            value = value.item()
                                    except (ImportError, AttributeError, Exception):
                                        try:
                                            value = float(value)
                                        except (TypeError, ValueError):
                                            try:
                                                value = str(value)
                                            except Exception:
                                                value = None
                            except Exception as e:
                                logger.warning(f"Error evaluating formula in {cell_ref}: {str(e)}")
                        
                        # Store formula info for cells with formulas
                        if has_formula:
                            # Try to manually calculate simple formulas like D*E or F*0.25
                            if formula and formula.startswith("="):
                                # First check for simple multiplication with a constant
                                if "*" in formula and formula.count("*") == 1 and any(c.isdigit() for c in formula):
                                    try:
                                        # Extract the parts of the formula
                                        parts = formula[1:].split("*")
                                        
                                        # Check if one part is a cell reference and the other is a number
                                        cell_ref_part = None
                                        number_part = None
                                        
                                        for part in parts:
                                            part = part.strip()
                                            # Check if part looks like a cell reference
                                            if any(c.isalpha() for c in part) and any(c.isdigit() for c in part):
                                                cell_ref_part = part
                                            # Check if part is a number
                                            elif part.replace('.', '', 1).isdigit():
                                                number_part = part
                                                
                                        # If we have both parts, get the cell value and calculate
                                        if cell_ref_part and number_part:
                                            cell_value = all_values.get(cell_ref_part)
                                            
                                            if cell_value is not None:
                                                try:
                                                    cell_value = float(cell_value)
                                                    number_value = float(number_part)
                                                    manually_calculated = cell_value * number_value
                                                    value = manually_calculated
                                                except (ValueError, TypeError):
                                                    pass
                                        # Otherwise try the D*E pattern for two cell references
                                        elif "*" in formula and formula.count("*") == 1:
                                            # Try the cell reference pattern like D2*E2
                                            cell_refs_in_formula = formula[1:].split("*")
                                            if len(cell_refs_in_formula) == 2:
                                                ref1, ref2 = cell_refs_in_formula
                                                
                                                # Get values directly - handle both D*E pattern and D2*E2 pattern
                                                cell1_val = None
                                                cell2_val = None
                                                
                                                # For D2*E2 pattern, extract row and column
                                                if any(c.isdigit() for c in ref1) and any(c.isdigit() for c in ref2):
                                                    # First get values from our context
                                                    cell1_val = all_values.get(ref1)
                                                    cell2_val = all_values.get(ref2)
                                                else:
                                                    # For D*E pattern in the current row, look up values
                                                    # Extract column letter from each reference
                                                    col1_letter = ''.join(c for c in ref1 if c.isalpha())
                                                    col2_letter = ''.join(c for c in ref2 if c.isalpha())
                                                    
                                                    # Get cell references for current row
                                                    new_ref1 = f"{col1_letter}{row}"
                                                    new_ref2 = f"{col2_letter}{row}"
                                                    
                                                    # Get values
                                                    cell1_val = all_values.get(new_ref1)
                                                    cell2_val = all_values.get(new_ref2)
                                                
                                                # Calculate if we have both values
                                                if cell1_val is not None and cell2_val is not None:
                                                    try:
                                                        cell1_val = float(cell1_val)
                                                        cell2_val = float(cell2_val)
                                                        manually_calculated = cell1_val * cell2_val
                                                        value = manually_calculated
                                                    except (ValueError, TypeError):
                                                        pass
                                    except Exception as calc_error:
                                        logger.warning(f"Error calculating formula: {calc_error}")
                            
                            formula_cells.append({
                                "cell": cell_ref,
                                "formula": formula,
                                "calculated_value": value
                            })
                            row_has_formula = True
                        
                        # Append value to the row data
                        values.append(value)
                        formulas.append(formula)
                        
                        if value is not None and value != "" or has_formula:
                            row_has_data = True
                    
                    if row_has_data:
                        record = {
                            "row": row,
                            "values": values,
                            "cell_refs": cell_refs,
                        }
                        
                        # Only include formulas when they exist
                        if row_has_formula:
                            record["formulas"] = formulas
                            
                        records.append(record)
                
                # Count charts and images
                chart_count = len([drawing for drawing in ws_values._charts])
                image_count = len([drawing for drawing in ws_values._images])
                
                # Prepare sheet data
                sheet_data = {
                    "sheet_name": sheet_name,
                    "dimensions": cell_range or ws_formulas.calculate_dimension(),
                    "non_empty_cells": len(records),
                    "charts": chart_count,
                    "images": image_count,
                    "columns": columns,
                    "column_refs": [get_column_letter(col) for col in range(min_col, max_col + 1)],
                    "records": records,
                    "has_formulas": len(formula_cells) > 0,
                    "formula_cells": formula_cells
                }
                
                # Add sheet data to response
                file_info["sheet"] = sheet_data
                file_info["success"] = True
        
        # Handle requested sheet not found
        if sheet_name and not file_info.get("sheet"):
            return {
                "success": False,
                "error": f"Sheet '{sheet_name}' not found in workbook"
            }
            
        # Clean up
        workbook_formulas.close()
        workbook_values.close()
        
        return file_info
        
    except Exception as e:
        logger.error(f"Failed to read Excel with formulas: {e}")
        return {
            "success": False,
            "error": f"Failed to read Excel file: {str(e)}"
        }