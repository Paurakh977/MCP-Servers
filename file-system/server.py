from mcp.server import Server
import mcp.types as types
import asyncio
import os, sys
from typing import Optional, Dict, Any, List
from mcp.server.stdio import stdio_server
from pydantic import AnyUrl
import urllib.parse
import mimetypes
import json
import time
from datetime import datetime


# checking some crusial imports because in the client config file the env might not have the following dependencies
try:
    import PyPDF2
except ImportError:
    print("PyPDF2 not installed. To read PDF files: pip install PyPDF2")

try:
    import docx  # for .docx files
except ImportError:
    print("python-docx not installed. To read Word files: pip install python-docx")

try:
    import fitz  # PyMuPDF

    has_pymupdf = True
except ImportError:
    has_pymupdf = False
    print("PyMuPDF not installed. For better PDF image detection: pip install pymupdf")

try:
    import openpyxl  # for .xlsx files
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import range_boundaries

    has_openpyxl = True
except ImportError:
    has_openpyxl = False
    print("openpyxl not installed. To read Excel files: pip install openpyxl")

try:
    from pptx import Presentation  # for .pptx files
except ImportError:
    print(
        "python-pptx not installed. To read PowerPoint files: pip install python-pptx"
    )

try:
    import pdfplumber

    has_pdfplumber = True
except ImportError:
    has_pdfplumber = False
    print(
        "pdfplumber not installed. For better PDF text extraction: pip install pdfplumber"
    )

try:
    import tabula

    has_tabula = True
except ImportError:
    has_tabula = False
    print(
        "Tabula-py not installed. For better PDF table extraction: pip install tabula-py"
    )

try:
    from PIL import Image

    has_pil = True
except ImportError:
    print("Pillow not installed. For better image analysis: pip install Pillow")
    has_pil = False

try:
    import ebooklib
    from ebooklib import epub
    from bs4 import BeautifulSoup

    has_epub_support = True
except ImportError:
    has_epub_support = False
    print(
        "EPUB support not available. To read EPUB files: pip install ebooklib beautifulsoup4"
    )

try:
    import striprtf.striprtf as striprtf

    has_rtf_support = True
except ImportError:
    has_rtf_support = False
    print("RTF support not available. To read RTF files: pip install striprtf")

from resources.core import (
    # Feature flags
    has_pymupdf,
    has_pdfplumber,
    has_tabula,
    has_pil,
    has_epub_support,
    has_rtf_support,
    check_path_security,
    check_against_single_base,
    get_mime_type,
    parse_query_params,
    get_file_extraction_options,
)

from resources import (
    extract_file_content,
    read_file,
    find_file_in_allowed_dirs,
    get_directory_listing,
    get_file_info,
    get_resource_definitions,
    get_tool_definitions,
    PROMPTS,
)

from resources.readers import (
    read_text_file,
    read_pdf_file,
    read_docx_file,
    read_xlsx_file,
    read_pptx_file,
    read_csv_file,
    read_epub_file,
    read_rtf_file,
)
from resources.utils.formatters import summarize_content

# Import Excel tools
from resources.excel_tools import (
    create_excel_workbook,
    get_workbook_metadata,
    create_worksheet,
    copy_worksheet,
    delete_worksheet,
    rename_worksheet,
    copy_excel_range,
    delete_excel_range,
    merge_excel_cells,
    unmerge_excel_cells,
    write_excel_data,
    format_excel_range,
    adjust_column_widths,
    apply_excel_formula,
    apply_excel_formula_range,
    delete_excel_workbook,
    add_excel_column,
    validate_excel_formula,
    add_data_validation,
    apply_conditional_formatting,
    read_excel_with_formulas,
    evaluate_excel_formula
)

# Import new table and pivot table tools
from resources.excel_table_pivot_tools import (
    create_excel_table,
    sort_excel_table,
    filter_excel_table,
    create_pivot_table,
    modify_pivot_table_fields,
    sort_pivot_table_field,
    filter_pivot_table_items,
    set_pivot_table_value_field_calculation,
    refresh_pivot_table,
    add_pivot_table_calculated_field,
    add_pivot_table_calculated_item,
    create_pivot_table_slicer,
    modify_pivot_table_slicer,
    set_pivot_table_layout,
    configure_pivot_table_totals,
    format_pivot_table_part,
    change_pivot_table_data_source,
    group_pivot_field_items,
    ungroup_pivot_field_items,
    apply_pivot_table_conditional_formatting,
    create_timeline_slicer,
    connect_slicer_to_pivot_tables,
    setup_power_pivot_data_model,
    create_power_pivot_measure
)

from resources.excel_charts_pivot_tools import ExcelChartsCore, create_dashboard_charts

# Initialize allowed directories list - will be populated with command-line arguments
allowed_directories = [os.getcwd()]  # Default to current directory

server = Server(
    name="file-system",
    instructions="You are a file system MCP server. You can read and extract content from various file types including text files, PDFs, Office documents (Word, Excel, PowerPoint), CSV, EPUB, and RTF. The server provides advanced extraction capabilities for tables, images, and metadata. PDF documents will include information about tables, images, and annotations when available. Office documents will extract embedded images, charts, tables, and preserve formatting where possible.",
    version="0.1",
)


@server.list_resources()
async def list_resources() -> list[types.Resource]:
    """
    Expose resources for file reading and information:
      1. file:///path - Read file content with advanced extraction
      2. excel-info:///path - Get Excel workbook metadata and sheet information
      3. excel-sheet:///path - Read specific Excel sheet content
      4. file-info:///path - Get file metadata and capabilities
      5. directory:///path - List directory contents
    """
    return get_resource_definitions()


@server.list_prompts()
async def list_prompts() -> list[types.Prompt]:
    """List all the available prompts for reading and writing files

    Returns:
        list[types.Prompt]: prompts' name, theor descriptions and arguments
    """
    return list(PROMPTS.values())


@server.list_tools()
async def list_tools() -> list[types.Tool]:
    """
    Returns a list of available tools for file operations.
    """
    return get_tool_definitions()


@server.read_resource()
async def read_resource(uri: AnyUrl) -> str:
    """Handle resource URIs, returning file content, information, or directory listings."""
    try:
        s = str(uri)

        # Parse file:/// URI
        if s.startswith("file:///"):
            try:
                uri_path = s[len("file:///") :]

                # Extract path and query parameters
                query_string = ""
                if "?" in uri_path:
                    uri_path, query_string = uri_path.split("?", 1)

                path = uri_path

                # Parse query parameters
                params = parse_query_params(query_string)

                # Extract parameters with defaults
                summarize = params.get("summarize", False)
                max_length = params.get("max_length", 500)
                metadata_only = params.get("metadata_only", False)

                # Verify path security
                security_check = check_path_security(allowed_directories, path)
                if not security_check["is_allowed"]:
                    return json.dumps(
                        {"error": f"Access denied: {security_check['message']}"}
                    )

                # Standard file content extraction
                result = read_file(
                    path, summarize=summarize, max_summary_length=max_length
                )

                if not result["success"]:
                    return json.dumps({"error": result["error"]})

                return json.dumps(
                    {
                        "content": result["content"],
                        "file_type": result["file_type"],
                        "file_path": result["file_path"],
                    }
                )

            except Exception as e:
                return json.dumps({"error": f"Failed to process file: {str(e)}"})

        # Parse excel-info:/// URI
        elif s.startswith("excel-info:///"):
            try:
                uri_path = s[len("excel-info:///") :]
                path = uri_path

                # Verify path security
                security_check = check_path_security(allowed_directories, path)
                if not security_check["is_allowed"]:
                    return json.dumps(
                        {"error": f"Access denied: {security_check['message']}"}
                    )

                # Verify file exists and is Excel
                if not os.path.exists(path):
                    return json.dumps({"error": f"File not found: {path}"})

                file_extension = os.path.splitext(path)[1].lower()
                if file_extension != ".xlsx":
                    return json.dumps({"error": f"Not an Excel file: {path}"})

                # Get Excel workbook info
                try:
                    workbook = openpyxl.load_workbook(
                        path, data_only=True, read_only=True
                    )

                    # Format the output in a more readable way
                    output_lines = []
                    output_lines.append(f"Excel File: {os.path.basename(path)}")
                    output_lines.append(f"Number of Sheets: {len(workbook.worksheets)}")
                    output_lines.append("\nSheet Information:")

                    for sheet in workbook.worksheets:
                        # Get sheet dimensions
                        min_col, min_row, max_col, max_row = range_boundaries(
                            sheet.calculate_dimension()
                        )

                        output_lines.append(f"\n📄 Sheet: {sheet.title}")
                        output_lines.append(
                            f"   Dimensions: {sheet.calculate_dimension()}"
                        )
                        output_lines.append(f"   Rows: {sheet.max_row}")
                        output_lines.append(f"   Columns: {max_col - min_col + 1}")

                        # Get column headers
                        columns = []
                        column_refs = []
                        for col in range(min_col, max_col + 1):
                            cell = sheet.cell(min_row, col)
                            header = (
                                cell.value
                                if cell.value is not None
                                else f"Column {get_column_letter(col)}"
                            )
                            columns.append(header)
                            column_refs.append(get_column_letter(col))

                        # Format column information
                        output_lines.append("   Columns:")
                        for i, (col_ref, col_name) in enumerate(
                            zip(column_refs, columns)
                        ):
                            output_lines.append(f"     {col_ref}: {col_name}")

                    workbook.close()
                    return json.dumps({"content": "\n".join(output_lines)})

                except Exception as e:
                    return json.dumps({"error": f"Failed to read Excel file: {str(e)}"})

            except Exception as e:
                return json.dumps({"error": f"Failed to process Excel info: {str(e)}"})

        # Parse excel-sheet:/// URI
        elif s.startswith("excel-sheet:///") :
            try:
                uri_path = s[len("excel-sheet:///") :]

                # Extract path and query parameters
                query_string = ""
                if "?" in uri_path:
                    uri_path, query_string = uri_path.split("?", 1)

                path = uri_path

                # Parse query parameters
                params = parse_query_params(query_string)

                # Get required sheet_name and optional cell_range
                sheet_name = params.get("sheet_name")
                if not sheet_name:
                    return json.dumps({"error": "sheet_name parameter is required"})

                cell_range = params.get("cell_range")

                # Verify path security
                security_check = check_path_security(allowed_directories, path)
                if not security_check["is_allowed"]:
                    return json.dumps(
                        {"error": f"Access denied: {security_check['message']}"}
                    )

                # Verify file exists and is Excel
                if not os.path.exists(path):
                    return json.dumps({"error": f"File not found: {path}"})

                file_extension = os.path.splitext(path)[1].lower()
                if file_extension != ".xlsx":
                    return json.dumps({"error": f"Not an Excel file: {path}"})

                # Use the formula-aware Excel reader
                try:
                    result = read_excel_with_formulas(path, sheet_name, cell_range)
                    
                    if not result.get("success", True):
                        return json.dumps({"error": result.get("error", "Failed to read Excel sheet")})
                        
                    return json.dumps(result)
                except Exception as e:
                    return json.dumps({"error": f"Failed to read Excel sheet: {str(e)}"})

            except Exception as e:
                return json.dumps({"error": f"Failed to process Excel sheet: {str(e)}"})

        # Parse file-info:/// URI
        elif s.startswith("file-info:///") :
            try:
                uri_path = s[len("file-info:///") :]
                path = uri_path

                # Verify path security
                security_check = check_path_security(allowed_directories, path)
                if not security_check["is_allowed"]:
                    return json.dumps(
                        {"error": f"Access denied: {security_check['message']}"}
                    )

                # Get file info
                if not os.path.exists(path):
                    return json.dumps({"error": f"File not found: {path}"})

                # Get file info as JSON
                file_info = get_file_info(path)
                return json.dumps(file_info)

            except Exception as e:
                return json.dumps({"error": f"Failed to get file info: {str(e)}"})

        # Parse directory:/// URI
        elif s.startswith("directory:///") :
            try:
                uri_path = s[len("directory:///") :]
                path = uri_path

                # Verify path security
                security_check = check_path_security(allowed_directories, path)
                if not security_check["is_allowed"]:
                    return json.dumps(
                        {"error": f"Access denied: {security_check['message']}"}
                    )

                # List directory contents
                if not os.path.isdir(path):
                    return json.dumps({"error": f"Directory not found: {path}"})

                # Get directory listing as JSON
                dir_listing = get_directory_listing(path)
                return json.dumps(dir_listing)

            except Exception as e:
                return json.dumps({"error": f"Failed to list directory: {str(e)}"})

        # Invalid URI
        else:
            return json.dumps({"error": f"Unsupported resource URI: {uri}"})

    except Exception as e:
        return json.dumps({"error": f"Invalid URI or unexpected error: {str(e)}"})


@server.get_prompt()
async def get_prompt(
    name: str, arguments: dict[str, str] | None = None
) -> types.GetPromptResult:
    """
    call the prompt by name with arguments
    Args:
        name (str): The name of the prompt to call.
        arguments (dict[str, str], optional): The arguments to pass to the prompt. Defaults to None.
    """
    print(f"PROMPT REQUESTED: {name} with arguments: {arguments}", file=sys.stderr)
    if name not in PROMPTS:
        raise ValueError(f"Prompt not found: {name}")

    # Common system-level instructions
    system_msg = """
You are a File Assistant. Whenever the user asks to read or inspect a file by name:
1. Retrieve the list of allowed directories by calling `list_allowed_directories`.
2. Recursively search each allowed directory (including subdirectories) for the best match to the requested filename or pattern.
3. If multiple candidates are found, ask the user to disambiguate.
4. Use the appropriate tool to read or inspect the file:
   - Use `read_file` or `read_excel_sheet` or whatever tool is best as per the context for reading content.
   - Use `get_excel_info` or `get_file_info` or whatever tool is best as per the context for metadata.
5. If summarization is requested, truncate or summarize to the specified `max_length`.
6. Always handle typos or partial names by finding the closest match.
7.**Critically**, format the result as a **markdown table** (with headers and rows) or valid CSV if large—never as prose.
8. If summarization is requested, truncate to the given character/token limit *after* tabular output.
"""

    # Dispatch per prompt
    if name == "find-and-read-file":
        filename = arguments.get("filename", "") if arguments else ""
        summarize = arguments.get("summarize", "false") if arguments else "false"
        max_length = arguments.get("max_length", "500") if arguments else "500"

        user_msg = (
            f"Find the file named or matching pattern `{filename}` and read its content. "
            f"Summarize: {summarize}. Max length: {max_length}."
        )

    elif name == "read-excel-as-table":
        system_msg += """\n
        **IMPORTANT - YOU MUST FORMAT ALL OUTPUT AS A MARKDOWN TABLE**
        You MUST format ALL excel/tabular data as a markdown table with headers and rows using | delimiters. 
        Example format:
        | Header1 | Header2 | Header3 |
        |---------|---------|---------|
        | Value1  | Value2  | Value3  |
        
        Only if the sheet is extremely wide (>10 columns), you may output valid CSV instead.
        If `summarize=true`, append a brief summary truncated to `max_length` *after* the table.
        """
        filename   = arguments.get("filename", "")   if arguments else ""
        sheet_name = arguments.get("sheet_name", "") if arguments else ""
        summarize  = arguments.get("summarize", "false") if arguments else "false"
        max_length = arguments.get("max_length", "reuturn full content ")    if arguments else "reuturn full content "
        
        user_msg = (
        f"Read the Excel/CSV file named \"{filename}\""
        + (f", sheet \"{sheet_name}\"" if sheet_name else "")
        + f" and display it as a markdown table. Summarize: {summarize}. Max length: {max_length}."
    )
    
    elif name == "find-files":
        pattern = arguments.get("pattern", "") if arguments else ""
        user_msg = f"Search for all files matching the pattern `{pattern}`."

    elif name == "find-file-info":
        filename = arguments.get("filename", "") if arguments else ""
        user_msg = f"Get detailed metadata and supported extraction capabilities for the file `{filename}`."

    else:
        # Should never happen
        raise ValueError("Prompt implementation not found")

    return types.GetPromptResult(
        messages=[
            types.PromptMessage(
                role="assistant",
                content=types.TextContent(type="text", text=system_msg.strip()),
            ),
            types.PromptMessage(
                role="user",
                content=types.TextContent(type="text", text=user_msg.strip()),
            ),
        ]
    )


@server.call_tool()
async def call_tool(tool_name: str, arguments: dict) -> list[types.TextContent]:
    """Call a tool by name with arguments."""
    try:
        if tool_name == "read_file":
            path = arguments["path"]
            summarize = arguments.get("summarize", False)
            max_length = arguments.get("max_length", 500)
            metadata_only = arguments.get("metadata_only", False)

            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                    print(f"Found file at: {path}", file=sys.stderr)
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find file matching '{path}' in allowed directories: {allowed_directories}",
                        )
                    ]

            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]

            # Standard file content extraction
            result = read_file(path, summarize=summarize, max_summary_length=max_length)

            if not result["success"]:
                return [
                    types.TextContent(
                        type="text", text=f"Failed to read file: {result['error']}"
                    )
                ]

            return [types.TextContent(type="text", text=result["content"])]

        elif tool_name == "get_excel_info":
            path = arguments["path"]

            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                    print(f"Found Excel file at: {path}", file=sys.stderr)
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories: {allowed_directories}",
                        )
                    ]

            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]

            # Verify file exists and is Excel
            if not os.path.exists(path):
                return [types.TextContent(type="text", text=f"File not found: {path}")]

            file_extension = os.path.splitext(path)[1].lower()
            if file_extension != ".xlsx":
                return [
                    types.TextContent(type="text", text=f"Not an Excel file: {path}")
                ]

            # Get Excel workbook info
            try:
                workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)

                # Format the output in a more readable way
                output_lines = []
                output_lines.append(f"Excel File: {os.path.basename(path)}")
                output_lines.append(f"Number of Sheets: {len(workbook.worksheets)}")
                output_lines.append("\nSheet Information:")

                for sheet in workbook.worksheets:
                    # Get sheet dimensions
                    min_col, min_row, max_col, max_row = range_boundaries(
                        sheet.calculate_dimension()
                    )

                    output_lines.append(f"\n📄 Sheet: {sheet.title}")
                    output_lines.append(f"   Dimensions: {sheet.calculate_dimension()}")
                    output_lines.append(f"   Rows: {sheet.max_row}")
                    output_lines.append(f"   Columns: {max_col - min_col + 1}")

                    # Get column headers
                    columns = []
                    column_refs = []
                    for col in range(min_col, max_col + 1):
                        cell = sheet.cell(min_row, col)
                        header = (
                            cell.value
                            if cell.value is not None
                            else f"Column {get_column_letter(col)}"
                        )
                        columns.append(header)
                        column_refs.append(get_column_letter(col))

                    # Format column information
                    output_lines.append("   Columns:")
                    for i, (col_ref, col_name) in enumerate(zip(column_refs, columns)):
                        output_lines.append(f"     {col_ref}: {col_name}")

                workbook.close()
                return [types.TextContent(type="text", text="\n".join(output_lines))]

            except Exception as e:
                return [
                    types.TextContent(
                        type="text", text=f"Failed to read Excel file: {str(e)}"
                    )
                ]

        elif tool_name == "read_excel_sheet":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            cell_range = arguments.get("cell_range")

            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                    print(f"Found Excel file at: {path}", file=sys.stderr)
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories: {allowed_directories}",
                        )
                    ]

            # Verify file exists and is Excel
            if not os.path.exists(path):
                return [types.TextContent(type="text", text=f"File not found: {path}")]

            file_extension = os.path.splitext(path)[1].lower()
            if file_extension != ".xlsx":
                return [
                    types.TextContent(type="text", text=f"Not an Excel file: {path}")
                ]

            # Use the formula-aware Excel reader
            try:
                result = read_excel_with_formulas(path, sheet_name, cell_range)
                
                if not result.get("success", True):
                    return [
                        types.TextContent(
                            type="text", 
                            text=f"Failed to read Excel sheet: {result.get('error', 'Unknown error')}"
                        )
                    ]
                    
                # Format the output to be more informative about formulas
                response_text = json.dumps(result, indent=2)
                
                # If there are formulas, add a note at the top
                if result.get("sheet", {}).get("has_formulas", False):
                    formula_info = "\n".join([
                        "NOTICE: This sheet contains formulas!",
                        "Formula cells are included in the 'formula_cells' section.",
                        "Each record with formulas contains both the formula and calculated value.",
                        "--------------------------------------------------------------\n"
                    ])
                    response_text = formula_info + response_text
                
                return [types.TextContent(type="text", text=response_text)]
            except Exception as e:
                return [types.TextContent(type="text", text=f"Failed to read Excel sheet: {str(e)}")]

        elif tool_name == "list_directory":
            path = arguments["path"]

            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                try:
                    # Try to find a directory matching the name
                    for base_dir in allowed_directories:
                        for root, dirs, _ in os.walk(base_dir):
                            if path.lower() in [d.lower() for d in dirs]:
                                path = os.path.join(
                                    root,
                                    next(d for d in dirs if path.lower() in d.lower()),
                                )
                                print(f"Found directory at: {path}", file=sys.stderr)
                                break
                        if os.path.exists(path):
                            break
                except Exception as e:
                    print(f"Error while searching for directory: {e}", file=sys.stderr)

            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]

            # List directory contents
            if not os.path.isdir(path):
                return [
                    types.TextContent(type="text", text=f"Directory not found: {path}")
                ]

            # Get directory contents with metadata
            contents = []
            for item in os.listdir(path):
                item_path = os.path.join(path, item)
                is_dir = os.path.isdir(item_path)

                # Get basic item info
                item_info = {
                    "name": item,
                    "path": item_path,
                    "is_dir": is_dir,
                }

                # For files, add additional info
                if not is_dir:
                    try:
                        file_size = os.path.getsize(item_path)
                        item_info.update(
                            {
                                "size": file_size,
                                "size_human": f"{file_size / 1024:.2f} KB",
                                "extension": os.path.splitext(item)[1].lower(),
                                "mime_type": get_mime_type(item_path),
                            }
                        )
                    except:
                        # If we can't get file info, provide minimal data
                        item_info.update(
                            {
                                "size": None,
                                "extension": os.path.splitext(item)[1].lower(),
                            }
                        )

                contents.append(item_info)

            # Sort contents: directories first, then files alphabetically
            contents.sort(key=lambda x: (not x["is_dir"], x["name"].lower()))

            formatted_listing = f"Directory listing for {path}:\n\n"

            # Add directories
            formatted_listing += "Directories:\n"
            dirs = [item for item in contents if item["is_dir"]]
            if dirs:
                for dir_item in dirs:
                    formatted_listing += f"  📁 {dir_item['name']}/\n"
            else:
                formatted_listing += "  (No directories)\n"

            # Add files
            formatted_listing += "\nFiles:\n"
            files = [item for item in contents if not item["is_dir"]]
            if files:
                for file_item in files:
                    size_info = f" ({file_item.get('size_human', 'unknown size')})"
                    formatted_listing += f"  📄 {file_item['name']}{size_info}\n"
            else:
                formatted_listing += "  (No files)\n"

            return [types.TextContent(type="text", text=formatted_listing)]

        elif tool_name == "get_file_info":
            path = arguments["path"]

            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                    print(f"Found file at: {path}", file=sys.stderr)
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find file matching '{path}' in allowed directories: {allowed_directories}",
                        )
                    ]

            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]

            # Get file info
            if not os.path.exists(path):
                return [types.TextContent(type="text", text=f"File not found: {path}")]

            # Get file info as JSON
            file_info = get_file_info(path)

            # Generate human-friendly output
            formatted_info = f"File Information for {path}:\n\n"
            formatted_info += f"File Size: {file_info['size_human']}\n"
            formatted_info += (
                f"Type: {file_info['file_type']} ({file_info['mime_type']})\n"
            )
            formatted_info += f"Modified: {time.ctime(file_info['modified'])}\n"
            formatted_info += f"Created: {time.ctime(file_info['created'])}\n\n"

            formatted_info += "Available Extraction Features:\n"

            for cap, available in file_info["capabilities"].items():
                if available:
                    formatted_info += f"✅ {cap.replace('_', ' ').title()}\n"
                else:
                    formatted_info += f"❌ {cap.replace('_', ' ').title()}\n"

            return [types.TextContent(type="text", text=formatted_info)]

        elif tool_name == "list_allowed_directories":
            return [
                types.TextContent(
                    type="text",
                    text=f"Allowed directories:\n{json.dumps(allowed_directories, indent=2)}",
                )
            ]

        # Excel Workbook Operations
        elif tool_name == "create_excel_workbook":
            path = arguments["path"]
            sheet_name = arguments.get("sheet_name", "Sheet1")
            
            # For new file creation, we need to check if the parent directory exists and is within allowed dirs
            parent_dir = os.path.dirname(os.path.abspath(path))
            
            # Create parent directories if they don't exist, but only if within allowed dirs
            is_path_allowed = False
            for base_dir in allowed_directories:
                # Check if parent_dir is within this allowed directory
                try:
                    rel_path = os.path.relpath(parent_dir, base_dir)
                    # If the relative path starts with '..' it means it's outside the base_dir
                    if not rel_path.startswith('..'):
                        is_path_allowed = True
                        break
                except ValueError:
                    # This happens on Windows if paths are on different drives
                    continue
            
            if not is_path_allowed:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: Path not within any allowed directory"
                    )
                ]
                
            # Ensure the parent directory exists
            os.makedirs(parent_dir, exist_ok=True)
                
            # Create workbook
            result = create_excel_workbook(path, sheet_name)
            
            if result["success"]:
                return [types.TextContent(type="text", text=result["message"])]
            else:
                return [types.TextContent(type="text", text=result["error"])]
                
        elif tool_name == "get_workbook_metadata":
            path = arguments["path"]
            include_ranges = arguments.get("include_ranges", False)
            
            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories"
                        )
                    ]
                    
            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]
                
            # Get workbook metadata
            result = get_workbook_metadata(path, include_ranges)
            
            if result["success"]:
                formatted_result = f"Excel Workbook: {result['filename']}\n"
                formatted_result += f"Size: {result['file_size_human']}\n"
                formatted_result += f"Modified: {result['modified_date']}\n"
                formatted_result += f"Sheets: {result['sheets_count']}\n\n"
                
                # Add sheet information
                for sheet in result["sheets"]:
                    formatted_result += f"Sheet: {sheet['name']}\n"
                    formatted_result += f"  Rows: {sheet['rows']}\n"
                    formatted_result += f"  Columns: {sheet['columns']}\n"
                    
                    if include_ranges and sheet.get("used_range"):
                        formatted_result += f"  Used Range: {sheet['used_range']}\n"
                        
                    if sheet.get("headers"):
                        formatted_result += f"  Headers: {', '.join(sheet['headers'])}\n"
                        
                    formatted_result += "\n"
                
                return [types.TextContent(type="text", text=formatted_result)]
            else:
                return [types.TextContent(type="text", text=result["error"])]
                
        # Excel Worksheet Operations
        elif tool_name == "create_worksheet":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            
            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories"
                        )
                    ]
                    
            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]
                
            # Create worksheet
            result = create_worksheet(path, sheet_name)
            
            if result["success"]:
                return [types.TextContent(type="text", text=result["message"])]
            else:
                return [types.TextContent(type="text", text=result["error"])]
                
        elif tool_name == "copy_worksheet":
            path = arguments["path"]
            source_sheet = arguments["source_sheet"]
            target_sheet = arguments["target_sheet"]
            
            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories"
                        )
                    ]
                    
            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]
                
            # Copy worksheet
            result = copy_worksheet(path, source_sheet, target_sheet)
            
            if result["success"]:
                return [types.TextContent(type="text", text=result["message"])]
            else:
                return [types.TextContent(type="text", text=result["error"])]
                
        elif tool_name == "delete_worksheet":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            
            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories"
                        )
                    ]
                    
            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]
                
            # Delete worksheet
            result = delete_worksheet(path, sheet_name)
            
            if result["success"]:
                return [types.TextContent(type="text", text=result["message"])]
            else:
                return [types.TextContent(type="text", text=result["error"])]
                
        elif tool_name == "rename_worksheet":
            path = arguments["path"]
            old_name = arguments["old_name"]
            new_name = arguments["new_name"]
            
            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories"
                        )
                    ]
                    
            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]
                
            # Rename worksheet
            result = rename_worksheet(path, old_name, new_name)
            
            if result["success"]:
                return [types.TextContent(type="text", text=result["message"])]
            else:
                return [types.TextContent(type="text", text=result["error"])]
                
        # Excel Range Operations
        elif tool_name == "copy_excel_range":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            source_range = arguments["source_range"]
            target_start = arguments["target_start"]
            target_sheet = arguments.get("target_sheet")
            
            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories"
                        )
                    ]
                    
            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]
                
            # Copy range
            result = copy_excel_range(path, sheet_name, source_range, target_start, target_sheet)
            
            if result["success"]:
                return [types.TextContent(type="text", text=result["message"])]
            else:
                return [types.TextContent(type="text", text=result["error"])]
                
        elif tool_name == "delete_excel_range":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            start_cell = arguments["start_cell"]
            end_cell = arguments.get("end_cell")
            shift_direction = arguments.get("shift_direction", "up")
            
            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories"
                        )
                    ]
                    
            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]
                
            # Delete range
            result = delete_excel_range(path, sheet_name, start_cell, end_cell, shift_direction)
            
            if result["success"]:
                return [types.TextContent(type="text", text=result["message"])]
            else:
                return [types.TextContent(type="text", text=result["error"])]
                
        elif tool_name == "merge_excel_cells":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            start_cell = arguments["start_cell"]
            end_cell = arguments["end_cell"]
            
            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories"
                        )
                    ]
                    
            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]
                
            # Merge cells
            result = merge_excel_cells(path, sheet_name, start_cell, end_cell)
            
            if result["success"]:
                return [types.TextContent(type="text", text=result["message"])]
            else:
                return [types.TextContent(type="text", text=result["error"])]
                
        elif tool_name == "unmerge_excel_cells":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            start_cell = arguments["start_cell"]
            end_cell = arguments["end_cell"]
            
            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories"
                        )
                    ]
                    
            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]
                
            # Unmerge cells
            result = unmerge_excel_cells(path, sheet_name, start_cell, end_cell)
            
            if result["success"]:
                return [types.TextContent(type="text", text=result["message"])]
            else:
                return [types.TextContent(type="text", text=result["error"])]
                
        # Excel Data Operations
        elif tool_name == "write_excel_data":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            data = arguments["data"]
            start_cell = arguments.get("start_cell", "A1")
            headers = arguments.get("headers", True)
            auto_adjust_width = arguments.get("auto_adjust_width", False)
            
            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories"
                        )
                    ]
                    
            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]
                
            # Write data
            result = write_excel_data(path, sheet_name, data, start_cell, headers, auto_adjust_width)
            
            if result["success"]:
                return [types.TextContent(type="text", text=result["message"])]
            else:
                return [types.TextContent(type="text", text=result["error"])]
                
        elif tool_name == "format_excel_range":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            start_cell = arguments["start_cell"]
            end_cell = arguments.get("end_cell")
            bold = arguments.get("bold", False)
            italic = arguments.get("italic", False)
            font_size = arguments.get("font_size")
            font_color = arguments.get("font_color")
            bg_color = arguments.get("bg_color")
            alignment = arguments.get("alignment")
            wrap_text = arguments.get("wrap_text", False)
            border_style = arguments.get("border_style")
            auto_adjust_width = arguments.get("auto_adjust_width", False)
            
            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories"
                        )
                    ]
                    
            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]
                
            # Apply formatting
            result = format_excel_range(
                path, sheet_name, start_cell, end_cell,
                bold, italic, font_size, font_color, bg_color, 
                alignment, wrap_text, border_style, auto_adjust_width
            )
            
            if result["success"]:
                return [types.TextContent(type="text", text=result["message"])]
            else:
                return [types.TextContent(type="text", text=result["error"])]

        elif tool_name == "adjust_column_widths":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            column_range = arguments.get("column_range")
            auto_fit = arguments.get("auto_fit", True)
            custom_widths = arguments.get("custom_widths")
            
            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories"
                        )
                    ]
                    
            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]
                
            # Adjust column widths
            result = adjust_column_widths(path, sheet_name, column_range, auto_fit, custom_widths)
            
            if result["success"]:
                return [types.TextContent(type="text", text=result["message"])]
            else:
                return [types.TextContent(type="text", text=result["error"])]

        elif tool_name == "apply_excel_formula":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            cell = arguments["cell"]
            formula = arguments["formula"]
            protect_from_errors = arguments.get("protect_from_errors", True)
            handle_arrays = arguments.get("handle_arrays", True)
            clear_spill_range = arguments.get("clear_spill_range", True)
            spill_rows = arguments.get("spill_rows", 200)

            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories"
                        )
                    ]

            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]

            # Apply formula
            result = apply_excel_formula(path, sheet_name, cell, formula, protect_from_errors, handle_arrays, clear_spill_range, spill_rows)

            if result["success"]:
                response_text = result["message"]
                # Add warnings if any
                if "warnings" in result:
                    response_text += "\n\nWarnings:"
                    for warning in result["warnings"]:
                        response_text += f"\n- {warning}"
                
                # Add array formula info if applicable
                if "array_formula_type" in result:
                    response_text += f"\n\nApplied as {result['array_formula_type']} array formula."
                    if "is_spill_formula" in result and result["is_spill_formula"]:
                        response_text += f" Cells below/right have been cleared to ensure proper spill behavior."
                
                # Add external reference info if applicable
                if "external_references" in result:
                    response_text += "\n\nExternal references used:"
                    for ref in result["external_references"]:
                        response_text += f"\n- {ref}"
                
                return [types.TextContent(type="text", text=response_text)]
            else:
                return [types.TextContent(type="text", text=result["error"])]

        elif tool_name == "apply_excel_formula_range":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            start_cell = arguments["start_cell"]
            end_cell = arguments["end_cell"]
            formula_template = arguments["formula_template"]
            protect_from_errors = arguments.get("protect_from_errors", True)
            dynamic_calculation = arguments.get("dynamic_calculation", True)
            clear_spill_range = arguments.get("clear_spill_range", True)
            chunk_size = arguments.get("chunk_size", 1000)

            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories"
                        )
                    ]

            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]

            # Apply formula to range
            result = apply_excel_formula_range(
                path, sheet_name, start_cell, end_cell, formula_template, 
                protect_from_errors, dynamic_calculation, chunk_size, clear_spill_range
            )

            if result["success"]:
                response_text = result["message"]
                # Add warnings if any
                if "warnings" in result:
                    response_text += "\n\nWarnings:"
                    for warning in result["warnings"]:
                        response_text += f"\n- {warning}"
                
                # Add errors if any
                if "errors" in result:
                    response_text += "\n\nErrors encountered:"
                    for error in result["errors"]:
                        response_text += f"\n- {error}"
                        
                # Add array formula info if applicable
                if result.get("is_array_formula"):
                    response_text += f"\n\nApplied as array formula with dynamic calculation."
                    if result.get("applied_as") == "single_array_formula":
                        response_text += " Formula will spill results automatically."
                        
                return [types.TextContent(type="text", text=response_text)]
            else:
                return [types.TextContent(type="text", text=result["error"])]

        elif tool_name == "delete_excel_workbook":
            path = arguments["path"]
            
            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories"
                        )
                    ]
                    
            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]
                
            # Delete workbook
            result = delete_excel_workbook(path)
            
            if result["success"]:
                return [types.TextContent(type="text", text=result["message"])]
            else:
                return [types.TextContent(type="text", text=result["error"])]

        elif tool_name == "add_excel_column":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            column_name = arguments["column_name"]
            column_position = arguments.get("column_position")
            data = arguments.get("data")
            header_style = arguments.get("header_style")
            
            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories"
                        )
                    ]
                    
            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]
                
            # Add column
            result = add_excel_column(path, sheet_name, column_name, column_position, data, header_style)

            if result["success"]:
                return [types.TextContent(type="text", text=result["message"])]
            else:
                return [types.TextContent(type="text", text=result["error"])]

        elif tool_name == "add_data_validation":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            cell_range = arguments["cell_range"]
            validation_type = arguments["validation_type"]
            validation_criteria = arguments["validation_criteria"]
            error_message = arguments.get("error_message")
            
            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories"
                        )
                    ]
                    
            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]
                
            # Add data validation
            result = add_data_validation(path, sheet_name, cell_range, validation_type, validation_criteria, error_message)
            
            if result["success"]:
                return [types.TextContent(type="text", text=result["message"])]
            else:
                return [types.TextContent(type="text", text=result["error"])]

        # Add the new tool handler
        elif tool_name == "apply_conditional_formatting":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            cell_range = arguments["cell_range"]
            condition = arguments["condition"]
            bold = arguments.get("bold", False)
            italic = arguments.get("italic", False)
            font_size = arguments.get("font_size")
            font_color = arguments.get("font_color")
            bg_color = arguments.get("bg_color")
            alignment = arguments.get("alignment")
            wrap_text = arguments.get("wrap_text", False)
            border_style = arguments.get("border_style")
            condition_column = arguments.get("condition_column")
            format_entire_row = arguments.get("format_entire_row", False)
            columns_to_format = arguments.get("columns_to_format")
            handle_formulas = arguments.get("handle_formulas", True)
            outside_range_columns = arguments.get("outside_range_columns")
            compare_columns = arguments.get("compare_columns")
            date_format = arguments.get("date_format")
            icon_set = arguments.get("icon_set")

            # If path doesn't exist, try to find it in allowed directories
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    path = found_path
                else:
                    return [
                        types.TextContent(
                            type="text",
                            text=f"Could not find Excel file matching '{path}' in allowed directories"
                        )
                    ]
                    
            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                return [
                    types.TextContent(
                        type="text", text=f"Access denied: {security_check['message']}"
                    )
                ]
                
            # Apply conditional formatting with enhanced parameters
            result = apply_conditional_formatting(
                path, sheet_name, cell_range, condition,
                bold, italic, font_size, font_color, bg_color,
                alignment, wrap_text, border_style, condition_column, 
                format_entire_row, columns_to_format, handle_formulas,
                outside_range_columns, compare_columns, date_format, icon_set
            )
            
            if result["success"]:
                response_text = result["message"]
                
                # Add additional information if available
                if "icon_set_message" in result:
                    response_text += f"\n\n{result['icon_set_message']}"
                
                return [types.TextContent(type="text", text=response_text)]
            else:
                return [types.TextContent(type="text", text=result["error"])]

        # --- Excel Table Tool Implementations ---
        elif tool_name == "create_excel_table":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            data_range = arguments["data_range"]
            table_name = arguments["table_name"]
            table_style = arguments.get("table_style", "TableStyleMedium9")

            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}'")]
            
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]: return [types.TextContent(type="text", text=f"Access denied: {security_check['message']}")]

            result = create_excel_table(path, sheet_name, data_range, table_name, table_style)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]

        elif tool_name == "sort_excel_table":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            table_name = arguments["table_name"]
            sort_column_name = arguments["sort_column_name"]
            sort_order = arguments.get("sort_order", "ascending")

            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}'")]
            
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]: return [types.TextContent(type="text", text=f"Access denied: {security_check['message']}")]

            result = sort_excel_table(path, sheet_name, table_name, sort_column_name, sort_order)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]

        elif tool_name == "filter_excel_table":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            table_name = arguments["table_name"]
            column_name = arguments["column_name"]
            criteria1 = arguments["criteria1"]
            operator = arguments.get("operator", "equals")
            criteria2 = arguments.get("criteria2")

            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}'")]
            
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]: return [types.TextContent(type="text", text=f"Access denied: {security_check['message']}")]

            result = filter_excel_table(path, sheet_name, table_name, column_name, criteria1, operator, criteria2)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]

        # --- PivotTable Tool Implementations ---
        elif tool_name == "create_pivot_table":
            path = arguments["path"]
            source_sheet_name = arguments["source_sheet_name"]
            source_data_range = arguments["source_data_range"]
            target_sheet_name = arguments["target_sheet_name"]
            target_cell_address = arguments["target_cell_address"]
            pivot_table_name = arguments["pivot_table_name"]
            row_fields = arguments.get("row_fields")
            column_fields = arguments.get("column_fields")
            value_fields = arguments.get("value_fields")
            filter_fields = arguments.get("filter_fields")
            pivot_style = arguments.get("pivot_style", "PivotStyleMedium9")

            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}'")]
            
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]: return [types.TextContent(type="text", text=f"Access denied: {security_check['message']}")]
            
            result = create_pivot_table(
                filepath=path, 
                source_sheet_name=source_sheet_name, 
                source_data_range=source_data_range,
                target_sheet_name=target_sheet_name, 
                target_cell_address=target_cell_address, 
                pivot_table_name=pivot_table_name,
                row_fields=row_fields, 
                column_fields=column_fields, 
                value_fields=value_fields, 
                filter_fields=filter_fields, 
                pivot_style=pivot_style
            )
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]

        elif tool_name == "modify_pivot_table_fields":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            pivot_table_name = arguments["pivot_table_name"]
            add_row_fields = arguments.get("add_row_fields")
            add_column_fields = arguments.get("add_column_fields")
            add_value_fields = arguments.get("add_value_fields")
            add_filter_fields = arguments.get("add_filter_fields")
            remove_fields = arguments.get("remove_fields")

            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}'")]
            
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]: return [types.TextContent(type="text", text=f"Access denied: {security_check['message']}")]

            result = modify_pivot_table_fields(
                filepath=path, 
                sheet_name=sheet_name, 
                pivot_table_name=pivot_table_name,
                add_row_fields=add_row_fields, 
                add_column_fields=add_column_fields,
                add_value_fields=add_value_fields, 
                add_filter_fields=add_filter_fields, 
                remove_fields=remove_fields
            )
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]

        elif tool_name == "sort_pivot_table_field":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            pivot_table_name = arguments["pivot_table_name"]
            field_name = arguments["field_name"]
            sort_on_field = arguments["sort_on_field"]
            sort_order = arguments.get("sort_order", "ascending")
            sort_type = arguments.get("sort_type", "data")

            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}'")]
            
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]: return [types.TextContent(type="text", text=f"Access denied: {security_check['message']}")]

            result = sort_pivot_table_field(path, sheet_name, pivot_table_name, field_name, sort_on_field, sort_order, sort_type)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]

        elif tool_name == "filter_pivot_table_items":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            pivot_table_name = arguments["pivot_table_name"]
            field_name = arguments["field_name"]
            visible_items = arguments.get("visible_items")
            hidden_items = arguments.get("hidden_items")
            # filter_type = arguments.get("filter_type", "value") # For future expansion

            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}'")]
            
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]: return [types.TextContent(type="text", text=f"Access denied: {security_check['message']}")]

            result = filter_pivot_table_items(path, sheet_name, pivot_table_name, field_name, visible_items, hidden_items)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]
            
        elif tool_name == "refresh_pivot_table":
            path = arguments["path"]
            sheet_name = arguments["sheet_name"]
            pivot_table_name = arguments["pivot_table_name"]

            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}'")]
            
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]: return [types.TextContent(type="text", text=f"Access denied: {security_check['message']}")]

            result = refresh_pivot_table(path, sheet_name, pivot_table_name)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]

        elif tool_name == "set_pivot_table_value_field_calculation":
            path = arguments.get("path")
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}' in allowed directories.")]
            
            security_check_result = check_path_security(allowed_directories, path)
            if not security_check_result["is_allowed"]:
                return [types.TextContent(type="text", text=f"Access denied: {security_check_result['message']}")]
            
            arguments["filepath"] = path
            arguments.pop("path", None)

            result = set_pivot_table_value_field_calculation(**arguments)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]

        elif tool_name == "group_pivot_field_items":
            path = arguments.get("path")
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}' in allowed directories.")]

            security_check_result = check_path_security(allowed_directories, path)
            if not security_check_result["is_allowed"]:
                return [types.TextContent(type="text", text=f"Access denied: {security_check_result['message']}")]
            
            arguments["filepath"] = path
            arguments.pop("path", None)

            result = group_pivot_field_items(**arguments)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]
            
        elif tool_name == "ungroup_pivot_field_items":
            path = arguments.get("path")
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}' in allowed directories.")]

            security_check_result = check_path_security(allowed_directories, path)
            if not security_check_result["is_allowed"]:
                return [types.TextContent(type="text", text=f"Access denied: {security_check_result['message']}")]
            
            arguments["filepath"] = path
            arguments.pop("path", None)

            result = ungroup_pivot_field_items(**arguments)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]
            
        elif tool_name == "apply_pivot_table_conditional_formatting":
            path = arguments.get("path")
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}' in allowed directories.")]

            security_check_result = check_path_security(allowed_directories, path)
            if not security_check_result["is_allowed"]:
                return [types.TextContent(type="text", text=f"Access denied: {security_check_result['message']}")]
            
            arguments["filepath"] = path
            arguments.pop("path", None)

            result = apply_pivot_table_conditional_formatting(**arguments)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]

        elif tool_name == "configure_pivot_table_totals":
            path = arguments.get("path")
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}' in allowed directories.")]

            security_check_result = check_path_security(allowed_directories, path)
            if not security_check_result["is_allowed"]:
                return [types.TextContent(type="text", text=f"Access denied: {security_check_result['message']}")]
            
            arguments["filepath"] = path
            arguments.pop("path", None)

            result = configure_pivot_table_totals(**arguments)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]

        elif tool_name == "format_pivot_table_part":
            path = arguments.get("path")
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}' in allowed directories.")]
            
            security_check_result = check_path_security(allowed_directories, path)
            if not security_check_result["is_allowed"]:
                return [types.TextContent(type="text", text=f"Access denied: {security_check_result['message']}")]

            arguments["filepath"] = path
            arguments.pop("path", None)

            result = format_pivot_table_part(**arguments)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]

        elif tool_name == "change_pivot_table_data_source":
            path = arguments.get("path")
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}' in allowed directories.")]

            security_check_result = check_path_security(allowed_directories, path)
            if not security_check_result["is_allowed"]:
                return [types.TextContent(type="text", text=f"Access denied: {security_check_result['message']}")]
            
            arguments["filepath"] = path # Ensure 'filepath' is used as the parameter name
            arguments.pop("path", None) # Remove original 'path'

            result = change_pivot_table_data_source(**arguments)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]
            
        elif tool_name == "add_pivot_table_calculated_field":
            path = arguments.get("path")
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}' in allowed directories.")]

            security_check_result = check_path_security(allowed_directories, path)
            if not security_check_result["is_allowed"]:
                return [types.TextContent(type="text", text=f"Access denied: {security_check_result['message']}")]
            
            arguments["filepath"] = path
            arguments.pop("path", None)

            result = add_pivot_table_calculated_field(**arguments)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]
            
        elif tool_name == "add_pivot_table_calculated_item":
            path = arguments.get("path")
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}' in allowed directories.")]

            security_check_result = check_path_security(allowed_directories, path)
            if not security_check_result["is_allowed"]:
                return [types.TextContent(type="text", text=f"Access denied: {security_check_result['message']}")]
            
            arguments["filepath"] = path
            arguments.pop("path", None)

            result = add_pivot_table_calculated_item(**arguments)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]
            
        elif tool_name == "set_pivot_table_layout":
            path = arguments.get("path")
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}' in allowed directories.")]

            security_check_result = check_path_security(allowed_directories, path)
            if not security_check_result["is_allowed"]:
                return [types.TextContent(type="text", text=f"Access denied: {security_check_result['message']}")]
            
            arguments["filepath"] = path
            arguments.pop("path", None)

            result = set_pivot_table_layout(**arguments)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]
            
        elif tool_name == "create_pivot_table_slicer":
            path = arguments.get("path")
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}' in allowed directories.")]

            security_check_result = check_path_security(allowed_directories, path)
            if not security_check_result["is_allowed"]:
                return [types.TextContent(type="text", text=f"Access denied: {security_check_result['message']}")]
            
            arguments["filepath"] = path
            arguments.pop("path", None)

            result = create_pivot_table_slicer(**arguments)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]
            
        elif tool_name == "modify_pivot_table_slicer":
            path = arguments.get("path")
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}' in allowed directories.")]

            security_check_result = check_path_security(allowed_directories, path)
            if not security_check_result["is_allowed"]:
                return [types.TextContent(type="text", text=f"Access denied: {security_check_result['message']}")]
            
            arguments["filepath"] = path
            arguments.pop("path", None)

            result = modify_pivot_table_slicer(**arguments)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]
            
        elif tool_name == "create_timeline_slicer":
            path = arguments.get("path")
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}' in allowed directories.")]

            security_check_result = check_path_security(allowed_directories, path)
            if not security_check_result["is_allowed"]:
                return [types.TextContent(type="text", text=f"Access denied: {security_check_result['message']}")]
            
            arguments["filepath"] = path
            arguments.pop("path", None)

            result = create_timeline_slicer(**arguments)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]
            
        elif tool_name == "connect_slicer_to_pivot_tables":
            path = arguments.get("path")
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}' in allowed directories.")]

            security_check_result = check_path_security(allowed_directories, path)
            if not security_check_result["is_allowed"]:
                return [types.TextContent(type="text", text=f"Access denied: {security_check_result['message']}")]
            
            arguments["filepath"] = path
            arguments.pop("path", None)

            result = connect_slicer_to_pivot_tables(**arguments)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]
            
        elif tool_name == "setup_power_pivot_data_model":
            path = arguments.get("path")
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}' in allowed directories.")]

            security_check_result = check_path_security(allowed_directories, path)
            if not security_check_result["is_allowed"]:
                return [types.TextContent(type="text", text=f"Access denied: {security_check_result['message']}")]
            
            arguments["filepath"] = path
            arguments.pop("path", None)

            result = setup_power_pivot_data_model(**arguments)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]
            
        elif tool_name == "create_power_pivot_measure":
            path = arguments.get("path")
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path: path = found_path
                else: return [types.TextContent(type="text", text=f"Could not find Excel file '{path}' in allowed directories.")]

            security_check_result = check_path_security(allowed_directories, path)
            if not security_check_result["is_allowed"]:
                return [types.TextContent(type="text", text=f"Access denied: {security_check_result['message']}")]
            
            arguments["filepath"] = path
            arguments.pop("path", None)

            result = create_power_pivot_measure(**arguments)
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]
        # --- END ADVANCED PIVOTTABLE TOOL HANDLERS ---

        elif tool_name == "create_pivot_chart":
            workbook_path = get_validated_path(arguments["workbook_path"])
            source_type = arguments["source_type"]
            chart_type = arguments.get("chart_type", "COLUMN")
            chart_title = arguments.get("chart_title")
            position = arguments.get("position", [100, 100])
            sheet_name = arguments.get("sheet_name")
            
            core = ExcelChartsCore()
            try:
                # Open workbook
                wb_result = core.open_workbook(workbook_path, sheet_name)
                if wb_result["status"] != "success":
                    return [types.TextContent(type="text", text=f"Failed to open workbook: {wb_result.get('message', 'Unknown error')}")]
                
                result = {}
                
                if source_type == "pivot_table":
                    # Create chart from existing pivot table
                    pivot_table_name = arguments.get("pivot_table_name")
                    if not pivot_table_name:
                        return [types.TextContent(type="text", text="pivot_table_name is required for 'pivot_table' source type")]
                    
                    result = core.create_pivot_chart_from_table(
                        pivot_table_name=pivot_table_name,
                        chart_type=chart_type,
                        chart_title=chart_title,
                        position=tuple(position)
                    )
                
                elif source_type == "data_range":
                    # Create chart directly from data range
                    data_range = arguments.get("data_range")
                    if not data_range:
                        return [types.TextContent(type="text", text="data_range is required for 'data_range' source type")]
                    
                    result = core.create_chart_from_range(
                        data_range=data_range,
                        chart_type=chart_type,
                        chart_title=chart_title,
                        position=tuple(position)
                    )
                
                elif source_type == "new_pivot":
                    # Create new pivot table then chart
                    data_range = arguments.get("data_range")
                    pivot_config = arguments.get("pivot_config", {})
                    
                    if not data_range:
                        return [types.TextContent(type="text", text="data_range is required for 'new_pivot' source type")]
                    if not pivot_config:
                        return [types.TextContent(type="text", text="pivot_config is required for 'new_pivot' source type")]
                    
                    # Create pivot table first
                    pt_name = f"PivotTable_{int(time.time())}"
                    pt_result = core.create_pivot_table(
                        data_range=data_range,
                        pivot_table_name=pt_name,
                        destination=pivot_config.get("destination", "H1"),
                        row_fields=pivot_config.get("row_fields", []),
                        column_fields=pivot_config.get("column_fields", []),
                        value_fields=pivot_config.get("value_fields", [])
                    )
                    
                    if pt_result["status"] != "success":
                        return [types.TextContent(type="text", text=f"Failed to create pivot table: {pt_result.get('message', 'Unknown error')}")]
                    
                    # Create chart from new pivot table
                    result = core.create_pivot_chart_from_table(
                        pivot_table_name=pt_name,
                        chart_type=chart_type,
                        chart_title=chart_title,
                        position=tuple(position)
                    )
                
                return [types.TextContent(type="text", text=f"✅ Pivot chart creation result:\n{json.dumps(result, indent=2)}")]
                
            finally:
                core.close_workbook(save=True)

        # MANAGE CHART ELEMENTS TOOL
        elif tool_name == "manage_chart_elements":
            workbook_path = get_validated_path(arguments["workbook_path"])
            chart_name = arguments["chart_name"]
            sheet_name = arguments.get("sheet_name")
            
            core = ExcelChartsCore()
            try:
                wb_result = core.open_workbook(workbook_path, sheet_name)
                if wb_result["status"] != "success":
                    return [types.TextContent(type="text", text=f"Failed to open workbook: {wb_result.get('message', 'Unknown error')}")]
                
                results = []
                
                # Handle title configuration
                if "title_config" in arguments:
                    title_config = arguments["title_config"]
                    result = core.set_chart_title(
                        chart_name=chart_name,
                        title=title_config.get("title_text", ""),
                        show_title=title_config.get("show_title", True)
                    )
                    results.append(f"Title: {result}")
                
                # Handle axis configuration
                if "axis_config" in arguments:
                    axis_config = arguments["axis_config"]
                    if "x_axis_title" in axis_config:
                        result = core.set_axis_title(
                            chart_name=chart_name,
                            axis_type="X",
                            title=axis_config["x_axis_title"],
                            show_title=axis_config.get("show_x_title", True)
                        )
                        results.append(f"X-axis: {result}")
                    
                    if "y_axis_title" in axis_config:
                        result = core.set_axis_title(
                            chart_name=chart_name,
                            axis_type="Y",
                            title=axis_config["y_axis_title"],
                            show_title=axis_config.get("show_y_title", True)
                        )
                        results.append(f"Y-axis: {result}")
                
                # Handle legend configuration
                if "legend_config" in arguments:
                    legend_config = arguments["legend_config"]
                    result = core.set_legend_properties(
                        chart_name=chart_name,
                        show_legend=legend_config.get("show_legend", True),
                        position=legend_config.get("position", "RIGHT")
                    )
                    results.append(f"Legend: {result}")
                
                # Handle data labels
                if "data_labels" in arguments:
                    data_labels = arguments["data_labels"]
                    result = core.toggle_data_labels(
                        chart_name=chart_name,
                        show_labels=data_labels.get("show_labels", False),
                        series_index=data_labels.get("series_index", 1)
                    )
                    results.append(f"Data labels: {result}")
                
                # Handle gridlines
                if "gridlines" in arguments:
                    gridlines = arguments["gridlines"]
                    for axis, major in [("X", True), ("X", False), ("Y", True), ("Y", False)]:
                        key = f"{axis.lower()}_{'major' if major else 'minor'}"
                        if key in gridlines:
                            result = core.toggle_gridlines(
                                chart_name=chart_name,
                                axis_type=axis,
                                major=major,
                                show=gridlines[key]
                            )
                            results.append(f"Gridlines {axis} {'major' if major else 'minor'}: {result}")
                
                return [types.TextContent(type="text", text=f"✅ Chart elements management results:\n" + "\n".join(results))]
                
            finally:
                core.close_workbook(save=True)

        # APPLY CHART STYLING TOOL
        elif tool_name == "apply_chart_styling":
            workbook_path = get_validated_path(arguments["workbook_path"])
            chart_name = arguments["chart_name"]
            sheet_name = arguments.get("sheet_name")
            
            core = ExcelChartsCore()
            try:
                wb_result = core.open_workbook(workbook_path, sheet_name)
                if wb_result["status"] != "success":
                    return [types.TextContent(type="text", text=f"Failed to open workbook: {wb_result.get('message', 'Unknown error')}")]
                
                results = []
                
                # Apply style
                if "style_id" in arguments:
                    result = core.set_chart_style(chart_name, arguments["style_id"])
                    results.append(f"Style: {result}")
                
                # Apply layout
                if "layout_id" in arguments:
                    result = core.apply_chart_layout(chart_name, arguments["layout_id"])
                    results.append(f"Layout: {result}")
                
                # Change chart type
                if "new_chart_type" in arguments:
                    result = core.change_chart_type(chart_name, arguments["new_chart_type"])
                    results.append(f"Chart type: {result}")
                
                return [types.TextContent(type="text", text=f"✅ Chart styling results:\n" + "\n".join(results))]
                
            finally:
                core.close_workbook(save=True)

        # MANAGE PIVOT FIELDS TOOL
        elif tool_name == "manage_pivot_fields":
            workbook_path = get_validated_path(arguments["workbook_path"])
            pivot_table_name = arguments["pivot_table_name"]
            sheet_name = arguments.get("sheet_name")
            
            core = ExcelChartsCore()
            try:
                wb_result = core.open_workbook(workbook_path, sheet_name)
                if wb_result["status"] != "success":
                    return [types.TextContent(type="text", text=f"Failed to open workbook: {wb_result.get('message', 'Unknown error')}")]
                
                results = []
                
                # Handle field operations
                if "field_operations" in arguments:
                    for field_op in arguments["field_operations"]:
                        result = core.modify_pivot_fields(
                            pivot_table_name=pivot_table_name,
                            field_name=field_op["field_name"],
                            orientation=field_op["orientation"],
                            summary_function=field_op.get("summary_function")
                        )
                        results.append(f"Field {field_op['field_name']}: {result}")
                
                # Handle calculated fields
                if "calculated_fields" in arguments:
                    for calc_field in arguments["calculated_fields"]:
                        result = core.create_calculated_field(
                            pivot_table_name=pivot_table_name,
                            field_name=calc_field["field_name"],
                            formula=calc_field["formula"]
                        )
                        results.append(f"Calculated field {calc_field['field_name']}: {result}")
                
                return [types.TextContent(type="text", text=f"✅ Pivot fields management results:\n" + "\n".join(results))]
                
            finally:
                core.close_workbook(save=True)

        # CREATE COMBO CHART TOOL
        elif tool_name == "create_combo_chart":
            workbook_path = get_validated_path(arguments["workbook_path"])
            data_range = arguments["data_range"]
            primary_series = arguments["primary_series"]
            secondary_series = arguments["secondary_series"]
            primary_type = arguments.get("primary_type", "COLUMN")
            secondary_type = arguments.get("secondary_type", "LINE")
            chart_title = arguments.get("chart_title")
            position = arguments.get("position", [100, 100])
            sheet_name = arguments.get("sheet_name")
            
            core = ExcelChartsCore()
            try:
                wb_result = core.open_workbook(workbook_path, sheet_name)
                if wb_result["status"] != "success":
                    return [types.TextContent(type="text", text=f"Failed to open workbook: {wb_result.get('message', 'Unknown error')}")]
                
                result = core.create_combo_chart(
                    chart_name=chart_title or "ComboChart",
                    data_range=data_range,
                    primary_series=primary_series,
                    secondary_series=secondary_series,
                    primary_type=primary_type,
                    secondary_type=secondary_type
                )
                
                return [types.TextContent(type="text", text=f"✅ Combo chart creation result:\n{json.dumps(result, indent=2)}")]
                
            finally:
                core.close_workbook(save=True)

        # ADD CHART FILTERS TOOL
        elif tool_name == "add_chart_filters":
            workbook_path = get_validated_path(arguments["workbook_path"])
            pivot_table_name = arguments["pivot_table_name"]
            slicer_fields = arguments["slicer_fields"]
            sheet_name = arguments.get("sheet_name")
            
            core = ExcelChartsCore()
            try:
                wb_result = core.open_workbook(workbook_path, sheet_name)
                if wb_result["status"] != "success":
                    return [types.TextContent(type="text", text=f"Failed to open workbook: {wb_result.get('message', 'Unknown error')}")]
                
                results = []
                for slicer_field in slicer_fields:
                    field_name = slicer_field["field_name"]
                    position = slicer_field.get("position", [500, 100])
                    
                    result = core.add_slicer(
                        pivot_table_name=pivot_table_name,
                        field_name=field_name,
                        position=tuple(position)
                    )
                    results.append(f"Slicer {field_name}: {result}")
                
                return [types.TextContent(type="text", text=f"✅ Chart filters (slicers) results:\n" + "\n".join(results))]
                
            finally:
                core.close_workbook(save=True)

        # REFRESH AND UPDATE TOOL
        elif tool_name == "refresh_and_update":
            workbook_path = get_validated_path(arguments["workbook_path"])
            operation = arguments["operation"]
            sheet_name = arguments.get("sheet_name")
            
            core = ExcelChartsCore()
            try:
                wb_result = core.open_workbook(workbook_path, sheet_name)
                if wb_result["status"] != "success":
                    return [types.TextContent(type="text", text=f"Failed to open workbook: {wb_result.get('message', 'Unknown error')}")]
                
                if operation == "refresh_all":
                    result = core.refresh_pivot_data()
                
                elif operation == "refresh_pivot":
                    pivot_table_name = arguments.get("pivot_table_name")
                    if not pivot_table_name:
                        return [types.TextContent(type="text", text="pivot_table_name is required for 'refresh_pivot' operation")]
                    result = core.refresh_pivot_data(pivot_table_name)
                
                elif operation == "update_chart_source":
                    chart_name = arguments.get("chart_name")
                    new_data_range = arguments.get("new_data_range")
                    if not chart_name or not new_data_range:
                        return [types.TextContent(type="text", text="chart_name and new_data_range are required for 'update_chart_source' operation")]
                    result = core.update_chart_data_source(chart_name, new_data_range)
                
                else:
                    return [types.TextContent(type="text", text=f"Unknown operation: {operation}")]
                
                return [types.TextContent(type="text", text=f"✅ Refresh/update result:\n{json.dumps(result, indent=2)}")]
                
            finally:
                core.close_workbook(save=True)

        # EXPORT AND DISTRIBUTE TOOL
        elif tool_name == "export_and_distribute":
            workbook_path = get_validated_path(arguments["workbook_path"])
            operation = arguments["operation"]
            sheet_name = arguments.get("sheet_name")
            
            core = ExcelChartsCore()
            try:
                wb_result = core.open_workbook(workbook_path, sheet_name)
                if wb_result["status"] != "success":
                    return [types.TextContent(type="text", text=f"Failed to open workbook: {wb_result.get('message', 'Unknown error')}")]
                
                if operation == "export_chart":
                    chart_name = arguments.get("chart_name")
                    export_path = arguments.get("export_path")
                    file_format = arguments.get("file_format", "PNG")
                    
                    if not chart_name or not export_path:
                        return [types.TextContent(type="text", text="chart_name and export_path are required for 'export_chart' operation")]
                    
                    result = core.export_chart(chart_name, export_path, file_format)
                    return [types.TextContent(type="text", text=f"✅ Chart export result:\n{json.dumps(result, indent=2)}")]
                
                elif operation == "create_dashboard":
                    dashboard_config = arguments.get("dashboard_config", [])
                    if not dashboard_config:
                        return [types.TextContent(type="text", text="dashboard_config is required for 'create_dashboard' operation")]
                    
                    result = create_dashboard_charts(workbook_path, dashboard_config)
                    return [types.TextContent(type="text", text=f"✅ Dashboard creation result:\n{json.dumps(result, indent=2)}")]
                
                else:
                    return [types.TextContent(type="text", text=f"Unknown operation: {operation}")]
                
            finally:
                core.close_workbook(save=True)

        # GET CHART INFO TOOL
        elif tool_name == "get_chart_info":
            workbook_path = get_validated_path(arguments["workbook_path"])
            info_type = arguments["info_type"]
            sheet_name = arguments.get("sheet_name")
            
            core = ExcelChartsCore()
            try:
                wb_result = core.open_workbook(workbook_path, sheet_name)
                if wb_result["status"] != "success":
                    return [types.TextContent(type="text", text=f"Failed to open workbook: {wb_result.get('message', 'Unknown error')}")]
                
                if info_type == "list_charts":
                    result = core.list_all_charts()
                
                elif info_type == "list_pivot_tables":
                    result = core.list_pivot_tables()
                
                elif info_type == "chart_details":
                    chart_name = arguments.get("chart_name")
                    if not chart_name:
                        return [types.TextContent(type="text", text="chart_name is required for 'chart_details' info type")]
                    result = core.get_chart_info(chart_name)
                
                elif info_type == "workbook_overview":
                    charts_result = core.list_all_charts()
                    pivot_result = core.list_pivot_tables()
                    result = {
                        "status": "success",
                        "workbook": wb_result["workbook"],
                        "sheets": wb_result["sheets"],
                        "charts": charts_result.get("charts", []),
                        "pivot_tables": pivot_result.get("pivot_tables", [])
                    }
                
                else:
                    return [types.TextContent(type="text", text=f"Unknown info_type: {info_type}")]
                
                return [types.TextContent(type="text", text=f"📊 Chart information:\n{json.dumps(result, indent=2)}")]
                
            finally:
                core.close_workbook(save=False)
       
        else:
            return [types.TextContent(type="text", text=f"Unknown tool: {tool_name}")]
        

    except Exception as e:
        return [types.TextContent(type="text", text=f"Error: {str(e)}")]


async def main() -> None:
    # Parse command-line arguments for allowed directories
    global allowed_directories
    if len(sys.argv) > 1:
        allowed_directories = [
            os.path.abspath(os.path.normpath(dir_path)) for dir_path in sys.argv[1:]
        ]

    print(f"Starting MCP file-system server with allowed directories:", file=sys.stderr)
    for directory in allowed_directories:
        print(f"  - {directory}", file=sys.stderr)

    async with stdio_server() as streams:
        await server.run(
            read_stream=streams[0],
            write_stream=streams[1],
            initialization_options=server.create_initialization_options(),
        )

def get_validated_path(path: str) -> str:
            if not os.path.exists(path):
                found_path = find_file_in_allowed_dirs(path, allowed_directories)
                if found_path:
                    print(f"Found Excel file at: {found_path}", file=sys.stderr)
                    return found_path
                else:
                    raise FileNotFoundError(f"Could not find Excel file matching '{path}' in allowed directories: {allowed_directories}")
            
            # Verify path security
            security_check = check_path_security(allowed_directories, path)
            if not security_check["is_allowed"]:
                raise PermissionError(f"Access denied: {security_check['message']}")
            
            # Verify file is Excel format
            file_extension = os.path.splitext(path)[1].lower()
            if file_extension not in ['.xlsx', '.xlsm', '.xls']:
                raise ValueError(f"Not an Excel file: {path}")
            
            return path

if __name__ == "__main__":
    asyncio.run(main())
